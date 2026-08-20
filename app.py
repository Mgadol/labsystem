# ── ЦАГИЙН БҮС ──────────────────────────────────────────
# Сервер UTC-д ажилладаг тул системийн бүх цаг (шинжилгээ, бэлтгэл, орчны
# бүртгэл, нөөцлөлт) 8 цагаар хоцордог байсан. Процессийн цагийн бүсийг
# тохируулснаар datetime.now() болон SQLite-ийн 'localtime' хоёулаа
# Улаанбаатарын цагаар явна. Шаардлагатай бол TZ хувьсагчаар дарж болно.
import os, time as _time
os.environ.setdefault('TZ', 'Asia/Ulaanbaatar')
try:
    _time.tzset()          # Unix; Windows дээр байхгүй
except AttributeError:
    pass

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
from models import get_db, init_db, hash_password, check_password
from datetime import datetime, date
from functools import wraps
import uuid, io, secrets, time
from werkzeug.utils import secure_filename

app = Flask(__name__)

# ── ХУВИЛБАР ──────────────────────────────────────────────────
# Тохиргоо хуудсанд харагдана. Өөр лабораторид алдаа гарахад "ямар
# хувилбар ажиллаж байна вэ" гэдгийг хариулах, шинэчлэлт үнэхээр
# буусан эсэхийг батлах боломжтой болгоно.
VERSION = '1.0.0'

# ── ХЭЛ ───────────────────────────────────────────────────────
# Одоогоор зөвхөн монгол. Кодод англи орчуулга үлдсэн тул хожим
# (жишээ нь гадаад түншид зориулж) энэ тугийг True болгоход л сэргэнэ.
LANGUAGES_ENABLED = False

# ── SECRET KEY: instance/-д хадгалагдана, автоматаар үүснэ ────
_KEY_FILE = os.path.join(os.path.dirname(__file__), 'instance', 'secret_key')
os.makedirs(os.path.dirname(_KEY_FILE), exist_ok=True)
if os.path.exists(_KEY_FILE):
    with open(_KEY_FILE, 'rb') as _f:
        app.secret_key = _f.read()
else:
    _k = secrets.token_bytes(32)
    with open(_KEY_FILE, 'wb') as _f:
        _f.write(_k)
    app.secret_key = _k

# ── SESSION COOKIE АЮУЛГҮЙ ТОХИРГОО ───────────────────────────
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE']   = os.environ.get('HTTPS_ENABLED', 'false').lower() == 'true'
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# ── LOGIN BRUTE-FORCE ХАМГААЛАЛТ ──────────────────────────────
_login_attempts = {}   # {ip: [timestamp, ...]}
_LOGIN_MAX   = 5       # дээд тоо
_LOGIN_WINDOW = 300    # 5 минут (секундээр)
_LOGIN_BLOCK  = 600    # 10 минут блоклоно

def _get_ip():
    return request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()

def _is_blocked(ip):
    now = time.time()
    attempts = [t for t in _login_attempts.get(ip, []) if now - t < _LOGIN_WINDOW]
    _login_attempts[ip] = attempts
    return len(attempts) >= _LOGIN_MAX

def _record_fail(ip):
    now = time.time()
    _login_attempts.setdefault(ip, []).append(now)

def _clear_attempts(ip):
    _login_attempts.pop(ip, None)

# ── SECURITY HEADERS ──────────────────────────────────────────
@app.after_request
def security_headers(resp):
    resp.headers['X-Content-Type-Options']  = 'nosniff'
    resp.headers['X-Frame-Options']          = 'SAMEORIGIN'
    resp.headers['X-XSS-Protection']         = '1; mode=block'
    resp.headers['Referrer-Policy']          = 'strict-origin-when-cross-origin'
    resp.headers['Permissions-Policy']       = 'geolocation=(), microphone=(), camera=()'
    # HTML хуудсыг хөтөч кэшлэхийг хориглоно. Кэшилсэн байвал серверт шинэ
    # хувилбар татсан ч хэрэглэгч хуучин хуудас (хуучин JS) ажиллуулсаар
    # байдаг тул засвар хэрэгжээгүй мэт харагддаг. static/ файлууд хэвээр
    # кэшлэгдэнэ (зураг, лого).
    if resp.mimetype == 'text/html':
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma']        = 'no-cache'
        resp.headers['Expires']       = '0'
    return resp

ALLOWED = {'png','jpg','jpeg','gif','webp','pdf','doc','docx'}

def allowed(filename):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in ALLOWED

def save_file(file, subfolder):
    if file and file.filename and allowed(file.filename):
        ext  = file.filename.rsplit('.',1)[1].lower()
        name = f"{uuid.uuid4().hex}.{ext}"
        path = os.path.join(app.config['UPLOAD_FOLDER'], subfolder)
        os.makedirs(path, exist_ok=True)
        dest = os.path.join(path, name)
        # Зураг бол чанарыг хадгалан оновчтой хэмжээнд resize хийнэ
        if ext in ('jpg','jpeg','png','webp'):
            try:
                from PIL import Image as PILImage, ExifTags
                img = PILImage.open(file.stream)
                # EXIF rotation засах
                try:
                    for key, val in ExifTags.TAGS.items():
                        if val == 'Orientation':
                            exif = img._getexif()
                            if exif and key in exif:
                                ori = exif[key]
                                if ori == 3:   img = img.rotate(180, expand=True)
                                elif ori == 6: img = img.rotate(270, expand=True)
                                elif ori == 8: img = img.rotate(90,  expand=True)
                            break
                except Exception:
                    pass
                # Профайл зураг: max 800px хадгалах (чанар алдахгүй)
                if subfolder == 'staff':
                    img.thumbnail((800, 1000), PILImage.LANCZOS)
                else:
                    img.thumbnail((1200, 1200), PILImage.LANCZOS)
                save_ext = 'JPEG' if ext in ('jpg','jpeg') else ext.upper()
                img.save(dest, save_ext, quality=92, optimize=True)
            except Exception:
                file.stream.seek(0)
                file.save(dest)
        else:
            file.save(dest)
        return f"{subfolder}/{name}"
    return None

def _validate_guest():
    """Зочны token DB-д байгаа эсэх, хугацаа дууссан эсэхийг шалгана"""
    token = session.get('guest_token')
    if not token:
        session.clear()
        return False
    conn = get_db()
    now = datetime.now().isoformat()
    row = conn.execute(
        "SELECT id FROM guest_tokens WHERE token=? AND expires_at >= ?", (token, now)
    ).fetchone()
    conn.close()
    if not row:
        session.clear()
        return False
    return True

def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'user_id' not in session and session.get('role') != 'guest':
            return redirect(url_for('login'))
        if session.get('role') == 'guest':
            if not _validate_guest():
                return redirect(url_for('login'))
            if request.method == 'POST':
                flash('Зочин горимд өөрчлөлт хийх боломжгүй.', 'error')
                return redirect(request.referrer or url_for('dashboard'))
        return f(*a, **kw)
    return dec

def guest_block(f):
    """Guest горимд POST үйлдлийг хаах"""
    @wraps(f)
    def dec(*a, **kw):
        if session.get('role') == 'guest' and request.method == 'POST':
            flash('Зочин горимд өөрчлөлт хийх боломжгүй.', 'error')
            return redirect(request.referrer or url_for('dashboard'))
        return f(*a, **kw)
    return dec

def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'user_id' not in session: return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Админы эрх шаардлагатай.', 'error')
            return redirect(url_for('dashboard'))
        return f(*a, **kw)
    return dec

def _has_perm(perm):
    """Тухайн хэрэглэгч тусгай эрхтэй эсэхийг шалгана (senior/admin автоматаар эрхтэй)"""
    role = session.get('role')
    if role in ('admin', 'senior'): return True
    uid = session.get('user_id')
    if not uid: return False
    conn = get_db()
    u = conn.execute(f"SELECT {perm} FROM users WHERE id=?", (uid,)).fetchone()
    conn.close()
    return bool(u and u[perm])

def perm_required(perm):
    """Тусгай эрх шаардлагатай decorator"""
    def decorator(f):
        @wraps(f)
        def dec(*a, **kw):
            if session.get('role') == 'guest':
                if request.method == 'POST':
                    flash('Зочин горимд өөрчлөлт хийх боломжгүй.', 'error')
                    return redirect(request.referrer or url_for('dashboard'))
                return f(*a, **kw)
            if 'user_id' not in session: return redirect(url_for('login'))
            if not _has_perm(perm):
                flash('Энэ үйлдэл хийх эрх байхгүй байна.', 'error')
                return redirect(url_for('dashboard'))
            return f(*a, **kw)
        return dec
    return decorator

def senior_required(f):
    """Админ + Ахлах химич хоёулан нэвтэрч болно"""
    @wraps(f)
    def dec(*a, **kw):
        if session.get('role') == 'guest':
            if request.method == 'POST':
                flash('Зочин горимд өөрчлөлт хийх боломжгүй.', 'error')
                return redirect(request.referrer or url_for('dashboard'))
            return f(*a, **kw)
        if 'user_id' not in session: return redirect(url_for('login'))
        if session.get('role') not in ('admin', 'senior'):
            flash('Ахлах химич эсвэл админы эрх шаардлагатай.', 'error')
            return redirect(url_for('dashboard'))
        return f(*a, **kw)
    return dec

def lab_required(f):
    """Лабын бүх ажилтан (senior, staff, preparer)"""
    @wraps(f)
    def dec(*a, **kw):
        if session.get('role') == 'guest':
            if request.method == 'POST':
                flash('Зочин горимд өөрчлөлт хийх боломжгүй.', 'error')
                return redirect(request.referrer or url_for('dashboard'))
            return f(*a, **kw)
        if 'user_id' not in session: return redirect(url_for('login'))
        if session.get('role') not in ('admin', 'senior', 'staff', 'preparer'):
            flash('Лабын ажилтны эрх шаардлагатай.', 'error')
            return redirect(url_for('dashboard'))
        return f(*a, **kw)
    return dec

def preparer_required(f):
    """Дээж бэлтгэгч + дээш"""
    @wraps(f)
    def dec(*a, **kw):
        if 'user_id' not in session: return redirect(url_for('login'))
        if session.get('role') not in ('admin', 'senior', 'staff', 'preparer'):
            flash('Дээж бэлтгэгчийн эрх шаардлагатай.', 'error')
            return redirect(url_for('dashboard'))
        return f(*a, **kw)
    return dec

def get_user(user_id):
    conn = get_db()
    u = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    conn.close()
    return u

@app.context_processor
def inject_user():
    user = None
    if session.get('role') == 'guest':
        user = {'id': 0, 'name': 'Зочин', 'role': 'guest', 'photo': None,
                'employee_id': 'GUEST', 'position': 'Зочин', 'phone': None,
                'email': None, 'joined_date': None, 'is_active': 1,
                'can_report': 0, 'can_approve': 0, 'can_reopen': 0,
                'can_export': 0, 'can_view_result': 0, 'can_register': 0}
    elif 'user_id' in session:
        user = get_user(session.get('user_id', 0))
    return dict(current_user=user, now=datetime.now())

# ── AUTH ────────────────────────────────────────────────
@app.route('/', methods=['GET','POST'])
def login():
    if 'user_id' in session: return redirect(url_for('dashboard'))
    lang = request.args.get('lang', session.get('lang','mn'))
    session['lang'] = lang
    error = None
    if request.method == 'POST':
        lang = request.form.get('lang','mn')
        session['lang'] = lang
        ip = _get_ip()
        if _is_blocked(ip):
            error = 'Хэт олон удаа буруу оруулсан. 10 минутын дараа дахин оролдоно уу.'
            return render_template('auth/login.html', error=error, lang=lang)
        emp_id = request.form.get('employee_id','').strip()
        pw     = request.form.get('password','')
        conn   = get_db()
        u = conn.execute("SELECT * FROM users WHERE (employee_id=? OR name=?) AND is_active=1",(emp_id,emp_id)).fetchone()
        conn.close()
        if u and check_password(u['password_hash'], pw):
            _clear_attempts(ip)
            session['user_id'] = u['id']
            session['role']    = u['role']
            return redirect(url_for('dashboard'))
        _record_fail(ip)
        remaining = _LOGIN_MAX - len(_login_attempts.get(ip, []))
        error = f'Нэвтрэх нэр эсвэл нууц үг буруу. ({max(0,remaining)} оролдлого үлдлээ)'
    return render_template('auth/login.html', error=error, lang=lang)

@app.route('/guest/<token>')
def guest_login(token):
    conn = get_db()
    now = datetime.now().isoformat()
    # Хугацаа дууссан токенуудыг устгана
    conn.execute("DELETE FROM guest_tokens WHERE expires_at < ?", (now,))
    conn.commit()
    row = conn.execute("SELECT * FROM guest_tokens WHERE token=? AND expires_at >= ?", (token.upper(), now)).fetchone()
    conn.close()
    if not row:
        return render_template('auth/login.html', error='Зочны код хүчингүй болсон эсвэл буруу байна.', lang='mn'), 403
    session['user_id'] = 0
    session['role'] = 'guest'
    session['lang'] = 'mn'
    session['guest_label'] = row['label'] or ''
    session['guest_token'] = token.upper()
    return redirect(url_for('dashboard'))

@app.route('/guest/token/delete/<int:tid>', methods=['POST'])
@admin_required
def guest_token_delete(tid):
    conn = get_db()
    conn.execute("DELETE FROM guest_tokens WHERE id=?", (tid,))
    conn.commit(); conn.close()
    flash('Зочны код устгагдлаа.', 'success')
    return redirect(url_for('lab_settings') + '?tab=guest')

@app.route('/guest/generate', methods=['POST'])
@admin_required
def guest_token_generate():
    import random, string
    label = request.form.get('label', '').strip() or 'Зочин'
    token = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    from datetime import timedelta
    expires_at = (datetime.now().replace(microsecond=0) + timedelta(hours=24)).isoformat()
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS guest_tokens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        token TEXT UNIQUE NOT NULL,
        label TEXT,
        created_by INTEGER REFERENCES users(id),
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        expires_at TEXT NOT NULL
    )""")
    try:
        conn.execute("INSERT INTO guest_tokens (token, label, created_by, expires_at) VALUES (?,?,?,?)",
                     (token, label, session['user_id'], expires_at))
        conn.commit()
        flash(f'Зочны код үүслээ: {token}  (24 цаг хүчинтэй, {expires_at[:16]} хүртэл)', 'success')
    except Exception as e:
        flash(f'Алдаа гарлаа: {e}', 'error')
    conn.close()
    return redirect(url_for('lab_settings') + '?tab=guest')

@app.route('/logout')
def logout():
    # Тоногийн ашиглалт хаах
    if 'user_id' in session:
        try:
            now = datetime.now().isoformat()
            conn = get_db()
            conn.execute("""UPDATE device_usage_log
                           SET ended_at=?,
                               duration_min=ROUND((JULIANDAY(?)-JULIANDAY(started_at))*1440,1)
                           WHERE user_id=? AND ended_at IS NULL""",
                        (now, now, session['user_id']))
            conn.commit()
            conn.close()
        except: pass
    session.clear()
    return redirect(url_for('login'))

@app.before_request
def _force_language():
    """Хэл идэвхгүй үед бүх сешнийг монгол руу буцаана.

    Урьд нь англи болгосон сешн үлдсэн байвал хэрэглэгч гацахгүй —
    товч байхгүй тул буцаах арга ч байхгүй болно.
    """
    if not LANGUAGES_ENABLED and session.get('lang', 'mn') != 'mn':
        session['lang'] = 'mn'


@app.route('/lang/<lang>')
def set_lang(lang):
    if LANGUAGES_ENABLED and lang in ('mn', 'en'):
        session['lang'] = lang
    else:
        session['lang'] = 'mn'
    ref = request.referrer
    return redirect(ref if ref else url_for('dashboard'))

# ── DASHBOARD ───────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    lang = session.get('lang','mn')
    conn = get_db()
    if session.get('role') in ('admin', 'senior', 'guest'):
        devices  = conn.execute("SELECT d.*, dm.manufacturer, dm.model FROM devices d LEFT JOIN device_marks dm ON d.mark_id=dm.id ORDER BY d.name").fetchall()
        users    = conn.execute("SELECT * FROM users WHERE is_active=1 AND role NOT IN ('geologist','bayjuulach')").fetchall()
        clients  = conn.execute("SELECT * FROM users WHERE is_active=1 AND role IN ('geologist','bayjuulach')").fetchall()
        open_rep = conn.execute("SELECT COUNT(*) as c FROM repairs WHERE status='new'").fetchone()['c']
        today    = date.today().isoformat()
        expiring = conn.execute("""
            SELECT d.id, d.name, c.next_date,
                   CAST(julianday(c.next_date) - julianday(?) AS INTEGER) as days
            FROM devices d
            JOIN calibrations c ON c.device_id=d.id
            WHERE c.id=(SELECT id FROM calibrations WHERE device_id=d.id ORDER BY calibration_date DESC LIMIT 1)
            AND days<=30 ORDER BY days
        """, (today,)).fetchall()
        my_devices, active_map = [], {}
        if session.get('role') == 'senior':
            uid = session.get('user_id', 0)
            my_devices = conn.execute("""
                SELECT d.*, dm.manufacturer, dm.model FROM devices d
                LEFT JOIN device_marks dm ON d.mark_id=dm.id
                JOIN staff_device_permissions p ON p.device_id=d.id
                WHERE p.user_id=? ORDER BY d.name
            """, (uid,)).fetchall()
            active_logs = conn.execute("SELECT * FROM usage_logs WHERE user_id=? AND end_time IS NULL", (uid,)).fetchall()
            active_map = {r['device_id']: r for r in active_logs}
        conn.close()
        return render_template('admin/dashboard.html',
            devices=devices, users=users, clients=clients, open_rep=open_rep, expiring=expiring,
            my_devices=my_devices, active_map=active_map, lang=lang)
    else:
        uid = session.get('user_id', 0)
        role = session.get('role')
        if role in ('bayjuulach', 'geologist'):
            # Харилцагчийн нүүр хуудас — баяжуулагч, геологич хоёулаа ижил
            user = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
            if role == 'bayjuulach':
                thousands = [6]                       # 6000–6999 баяжуулах дээж
            else:
                vr = user['view_ranges'] if user else None
                thousands = (None if vr is None else
                             [int(x) for x in vr.split(',') if x.strip().isdigit()])
            SAMPLE_SQL = """
                SELECT sr.id as receipt_id, sr.lab_number, sr.lab_serial, sr.received_date,
                       g.sample_name, g.sample_type, g.status
                FROM sample_receipt sr
                JOIN geo_samples g ON g.id=sr.geo_sample_id
                WHERE g.status='done' {rng}
                ORDER BY sr.lab_serial DESC LIMIT 100
            """
            samples = []
            if user and user['can_view_result']:
                if thousands is None:
                    samples = conn.execute(SAMPLE_SQL.format(rng='')).fetchall()
                elif thousands:
                    ph = ','.join('?' * len(thousands))
                    samples = conn.execute(
                        SAMPLE_SQL.format(rng=f'AND CAST(sr.lab_serial/1000 AS INTEGER) IN ({ph})'),
                        thousands).fetchall()
            # Хүрээний тайлбар — дээжийн төрлийн тохиргооноос уншина
            names = {int(r['serial_from'] or 0) // 1000: (r['name_mn'] or r['code'])
                     for r in conn.execute(
                         "SELECT code, name_mn, serial_from FROM sample_types WHERE serial_from IS NOT NULL")}
            if thousands is None:
                range_label = 'бүх дугаар'
            elif thousands:
                range_label = ', '.join(f'{t}000 {names.get(t, "")}'.strip() for t in sorted(thousands))
            else:
                range_label = 'муж тохируулаагүй'
            # ── Миний бүртгэсэн, хараахан дуусаагүй дээж ──
            # Дээрх жагсаалт нь ЗӨВХӨН дууссан ажлыг харуулдаг тул дөнгөж
            # бүртгэсэн дээж нүүр хуудсанд огт гардаггүй байв: геологич
            # өөрийн бичсэн нэрээ шалгах, засах газаргүй байсан.
            mine = conn.execute("""
                SELECT g.*, sr.id AS receipt_id, sr.lab_number, sr.received_date
                  FROM geo_samples g
                  LEFT JOIN sample_receipt sr ON sr.geo_sample_id = g.id
                 WHERE g.registered_by = ? AND COALESCE(g.status,'pending') <> 'done'
                 ORDER BY g.created_at DESC, g.id DESC LIMIT 50
            """, (uid,)).fetchall()
            conn.close()
            return render_template('staff/dashboard_client.html',
                user=user, samples=samples, lang=lang, range_label=range_label,
                mine=mine,
                role_label=('Баяжуулах цех' if role == 'bayjuulach' else 'Геологи'))
        my_devices = conn.execute("""
            SELECT d.*, dm.manufacturer, dm.model FROM devices d
            LEFT JOIN device_marks dm ON d.mark_id=dm.id
            JOIN staff_device_permissions p ON p.device_id=d.id
            WHERE p.user_id=? ORDER BY d.name
        """, (uid,)).fetchall()
        active_logs = conn.execute("SELECT * FROM usage_logs WHERE user_id=? AND end_time IS NULL", (uid,)).fetchall()
        conn.close()
        active_map = {r['device_id']: r for r in active_logs}
        return render_template('staff/dashboard.html',
            devices=my_devices, active_map=active_map, lang=lang)

# ── DEVICES ─────────────────────────────────────────────
@app.route('/devices')
@login_required
def devices():
    lang = session.get('lang','mn')
    conn = get_db()
    role = session.get('role')
    if role in ('admin', 'guest'):
        devs = conn.execute("""
            SELECT d.*, dm.manufacturer, dm.model, dm.category FROM devices d
            LEFT JOIN device_marks dm ON d.mark_id=dm.id
            ORDER BY (d.lab_id IS NULL OR d.lab_id=''), d.lab_id, d.name
        """).fetchall()
    else:
        devs = conn.execute("""
            SELECT d.*, dm.manufacturer, dm.model, dm.category FROM devices d
            LEFT JOIN device_marks dm ON d.mark_id=dm.id
            JOIN staff_device_permissions p ON p.device_id=d.id
            WHERE p.user_id=?
            ORDER BY (d.lab_id IS NULL OR d.lab_id=''), d.lab_id, d.name
        """, (session.get('user_id', 0),)).fetchall()
    # Хэрэглэгчийн бүх идэвхтэй ашиглалт (олон төхөөрөмж зэрэг ашиглаж болно)
    active_rows = conn.execute("SELECT id, device_id FROM usage_logs WHERE user_id=? AND end_time IS NULL",
                          (session.get('user_id', 0),)).fetchall()
    # Бүх хэрэглэгчийн идэвхтэй ашиглалт (хэн ямар төхөөрөмж дээр ажиллаж байгааг бусдад харуулах)
    busy_rows = conn.execute("""
        SELECT ul.device_id, ul.user_id, u.name as uname, ul.start_time
        FROM usage_logs ul LEFT JOIN users u ON u.id=ul.user_id
        WHERE ul.end_time IS NULL
    """).fetchall()
    conn.close()
    # device_id → log_id харгалзаа (өөрийн)
    active_map = {r['device_id']: r['id'] for r in active_rows}
    # device_id → {хэрэглэгчийн нэр, эхэлсэн цаг} (бусдын ашиглалт)
    busy_map = {}
    for r in busy_rows:
        busy_map.setdefault(r['device_id'], []).append(
            {'user_id': r['user_id'], 'name': r['uname'], 'start': (r['start_time'] or '')[11:16]})
    return render_template('device/list.html', devices=devs, lang=lang,
        active_map=active_map, busy_map=busy_map)

@app.route('/devices/<int:did>')
@login_required
def device_detail(did):
    lang = session.get('lang','mn')
    conn = get_db()
    device = conn.execute("SELECT d.*, dm.manufacturer, dm.model, dm.category FROM devices d LEFT JOIN device_marks dm ON d.mark_id=dm.id WHERE d.id=?", (did,)).fetchone()
    if not device:
        conn.close(); return redirect(url_for('devices'))
    if session.get('role') != 'admin':
        perm = conn.execute("SELECT 1 FROM staff_device_permissions WHERE user_id=? AND device_id=?", (session.get('user_id', 0), did)).fetchone()
        if not perm:
            conn.close()
            flash('Энэ төхөөрөмжид хандах эрх байхгүй.', 'error')
            return redirect(url_for('devices'))
    cals   = conn.execute("SELECT c.*, u.name as uname FROM calibrations c LEFT JOIN users u ON u.id=c.performed_by WHERE c.device_id=? ORDER BY c.calibration_date DESC LIMIT 20", (did,)).fetchall()
    reps   = conn.execute("SELECT r.*, u.name as uname FROM repairs r LEFT JOIN users u ON u.id=r.reported_by WHERE r.device_id=? ORDER BY r.reported_date DESC LIMIT 20", (did,)).fetchall()
    logs   = conn.execute("SELECT ul.*, u.name as uname FROM usage_logs ul LEFT JOIN users u ON u.id=ul.user_id WHERE ul.device_id=? ORDER BY ul.start_time DESC LIMIT 50", (did,)).fetchall()
    active = conn.execute("SELECT * FROM usage_logs WHERE device_id=? AND end_time IS NULL", (did,)).fetchone()
    # Monthly hours
    now = datetime.now()
    mhours = conn.execute("""
        SELECT COALESCE(SUM(duration_hours),0) as total FROM usage_logs
        WHERE device_id=? AND strftime('%Y-%m', start_time)=?
    """, (did, now.strftime('%Y-%m'))).fetchone()['total']
    analysis_usage = conn.execute("""
        SELECT u.id as user_id, u.name as user_name,
               COUNT(*) as sessions,
               COALESCE(SUM(CAST((julianday(ul.end_time)-julianday(ul.start_time))*1440 AS INTEGER)),0) as total_min,
               MAX(ul.start_time) as last_used
        FROM usage_logs ul
        JOIN users u ON u.id=ul.user_id
        WHERE ul.device_id=? AND ul.end_time IS NOT NULL
        GROUP BY u.id ORDER BY total_min DESC
    """, (did,)).fetchall()
    checks = [dict(r) for r in conn.execute("""
        SELECT ch.*, u.name as uname FROM device_checks ch
        LEFT JOIN users u ON u.id=ch.checked_by
        WHERE ch.device_id=? ORDER BY ch.check_date DESC, ch.id DESC LIMIT 60
    """, (did,)).fetchall()]
    calib_checks = []
    try:
        calib_checks = [dict(r) for r in conn.execute("""
            SELECT cc.*, u.name as uname FROM device_calib_checks cc
            LEFT JOIN users u ON u.id=cc.checked_by
            WHERE cc.device_id=? ORDER BY cc.check_date DESC LIMIT 24
        """, (did,)).fetchall()]
    except Exception: pass
    cal_checks = []
    try:
        cal_checks = [dict(r) for r in conn.execute("""
            SELECT cc.*, u.name as uname FROM device_calorimeter_checks cc
            LEFT JOIN users u ON u.id=cc.checked_by
            WHERE cc.device_id=? ORDER BY cc.check_date DESC LIMIT 24
        """, (did,)).fetchall()]
    except Exception: pass
    conn.close()
    return render_template('device/detail.html',
        device=device, calibrations=cals, repairs=reps,
        usage_logs=logs, active_log=active, checks=checks,
        calib_checks=calib_checks, cal_checks=cal_checks,
        monthly_hours=round(mhours,2),
        analysis_usage=analysis_usage,
        lang=lang, today=date.today().isoformat())

def _resolve_mark(conn, manufacturer, model, category):
    """Үйлдвэрлэгч/загвараар mark олох, байхгүй бол шинээр үүсгэж id буцаана."""
    manufacturer = (manufacturer or '').strip()
    model = (model or '').strip()
    category = (category or '').strip() or None
    if not manufacturer and not model:
        return None
    row = conn.execute(
        "SELECT id FROM device_marks WHERE manufacturer=? AND model=?",
        (manufacturer, model)).fetchone()
    if row:
        return row['id']
    cur = conn.execute(
        "INSERT INTO device_marks(manufacturer,model,category) VALUES(?,?,?)",
        (manufacturer, model, category))
    return cur.lastrowid

@app.route('/devices/add', methods=['GET','POST'])
@senior_required
def device_add():
    lang = session.get('lang','mn')
    conn = get_db()
    marks = conn.execute("SELECT * FROM device_marks ORDER BY manufacturer").fetchall()
    if request.method == 'POST':
        photo = save_file(request.files.get('photo'), 'devices')
        pdf   = save_file(request.files.get('passport_pdf'), 'passports')
        # Үйлдвэрлэгч/загвараас mark олох эсвэл шинээр үүсгэх
        mark_id = _resolve_mark(conn,
            request.form.get('manufacturer'),
            request.form.get('model'),
            request.form.get('category'))
        conn.execute("""
            INSERT INTO devices(name,serial_number,mark_id,location,purchase_date,
            warranty_expiry,calibration_interval,photo,passport_pdf,status,notes,
            lab_id,web_link,method,max_temp,particular,measuring_time,measuring_limit,
            dimension,capacity,weight_kg,other_spec,power,frequency,voltage,
            specification,operating_state,received_date,
            check_param1,check_standard,check_tolerance,
            check_param2,check_standard2,check_tolerance2,
            check_param3,check_standard3,check_tolerance3,
            check_param4,check_standard4,check_tolerance4,
            check_param5,check_standard5,check_tolerance5,
            check_enabled,stage,check_freq)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            request.form['name'],
            request.form.get('serial_number') or None,
            mark_id,
            request.form.get('location'),
            request.form.get('purchase_date') or None,
            request.form.get('warranty_expiry') or None,
            int(request.form.get('calibration_interval') or 90),
            photo, pdf, 'active',
            request.form.get('notes'),
            request.form.get('lab_id') or None,
            request.form.get('web_link') or None,
            request.form.get('method') or None,
            request.form.get('max_temp') or None,
            request.form.get('particular') or None,
            request.form.get('measuring_time') or None,
            request.form.get('measuring_limit') or None,
            request.form.get('dimension') or None,
            request.form.get('capacity') or None,
            request.form.get('weight_kg') or None,
            request.form.get('other_spec') or None,
            request.form.get('power') or None,
            request.form.get('frequency') or None,
            request.form.get('voltage') or None,
            request.form.get('specification') or None,
            request.form.get('operating_state') or None,
            request.form.get('received_date') or None,
            request.form.get('check_param1') or None,
            request.form.get('check_standard') or None,
            request.form.get('check_tolerance') or None,
            request.form.get('check_param2') or None,
            request.form.get('check_standard2') or None,
            request.form.get('check_tolerance2') or None,
            request.form.get('check_param3') or None,
            request.form.get('check_standard3') or None,
            request.form.get('check_tolerance3') or None,
            request.form.get('check_param4') or None,
            request.form.get('check_standard4') or None,
            request.form.get('check_tolerance4') or None,
            request.form.get('check_param5') or None,
            request.form.get('check_standard5') or None,
            request.form.get('check_tolerance5') or None,
            1 if request.form.get('check_enabled') else 0,
            request.form.get('stage') or 'both',
            request.form.get('check_freq') or 'daily',
        ))
        did = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        # Аналитик жин — босоо жингийн загвар (check_group1_cols=0) автоматаар тохируулна
        _nm = (request.form['name'] or '').lower()
        if 'жин' in _nm or 'налитик' in _nm:
            conn.execute("UPDATE devices SET check_group1=COALESCE(NULLIF(check_group1,''),'Туухай'), check_group1_cols=0 WHERE id=?", (did,))
        conn.commit(); conn.close()
        flash('Төхөөрөмж нэмэгдлээ!' if lang=='mn' else 'Device added!', 'success')
        return redirect(url_for('device_detail', did=did))
    conn.close()
    return render_template('device/add.html', marks=marks, lang=lang)

@app.route('/devices/<int:did>/edit', methods=['GET','POST'])
@senior_required
def device_edit(did):
    lang = session.get('lang','mn')
    conn = get_db()
    device = conn.execute("SELECT d.*, dm.manufacturer, dm.model, dm.category FROM devices d LEFT JOIN device_marks dm ON d.mark_id=dm.id WHERE d.id=?", (did,)).fetchone()
    marks  = conn.execute("SELECT * FROM device_marks").fetchall()
    if request.method == 'POST':
        # Зөвхөн шалгалтын тохиргоо хадгалах (check tab-аас дуудагдана)
        if request.form.get('_check_config_only'):
            try:
                conn.execute("""UPDATE devices SET
                    check_group1=?, check_group2=?, check_group3=?,
                    check_param1=?, check_standard=?, check_tolerance=?,
                    check_param2=?, check_standard2=?, check_tolerance2=?,
                    check_param3=?, check_standard3=?, check_tolerance3=?,
                    check_param4=?, check_standard4=?, check_tolerance4=?,
                    check_param5=?, check_standard5=?, check_tolerance5=?,
                    check_group1_cols=?, check_group2_cols=?, check_group3_cols=?
                    WHERE id=?""", (
                    request.form.get('check_group1') or None,
                    request.form.get('check_group2') or None,
                    request.form.get('check_group3') or None,
                    request.form.get('check_param1') or None,
                    request.form.get('check_standard') or None,
                    request.form.get('check_tolerance') or None,
                    request.form.get('check_param2') or None,
                    request.form.get('check_standard2') or None,
                    request.form.get('check_tolerance2') or None,
                    request.form.get('check_param3') or None,
                    request.form.get('check_standard3') or None,
                    request.form.get('check_tolerance3') or None,
                    request.form.get('check_param4') or None,
                    request.form.get('check_standard4') or None,
                    request.form.get('check_tolerance4') or None,
                    request.form.get('check_param5') or None,
                    request.form.get('check_standard5') or None,
                    request.form.get('check_tolerance5') or None,
                    int(request.form.get('check_group1_cols') or 2),
                    int(request.form.get('check_group2_cols') or 1),
                    int(request.form.get('check_group3_cols') or 1),
                    did))
                conn.commit()
            except Exception:
                try: conn.rollback()
                except Exception: pass
                conn.execute("""UPDATE devices SET
                    check_group1=?, check_standard=?, check_tolerance=?,
                    check_param1=?, check_param2=?, check_standard2=?, check_tolerance2=?,
                    check_param3=?, check_standard3=?, check_tolerance3=?,
                    check_param4=?, check_standard4=?, check_tolerance4=?,
                    check_param5=?, check_standard5=?, check_tolerance5=?
                    WHERE id=?""", (
                    request.form.get('check_group1') or None,
                    request.form.get('check_standard') or None,
                    request.form.get('check_tolerance') or None,
                    request.form.get('check_param1') or None,
                    request.form.get('check_param2') or None,
                    request.form.get('check_standard2') or None,
                    request.form.get('check_tolerance2') or None,
                    request.form.get('check_param3') or None,
                    request.form.get('check_standard3') or None,
                    request.form.get('check_tolerance3') or None,
                    request.form.get('check_param4') or None,
                    request.form.get('check_standard4') or None,
                    request.form.get('check_tolerance4') or None,
                    request.form.get('check_param5') or None,
                    request.form.get('check_standard5') or None,
                    request.form.get('check_tolerance5') or None,
                    did))
                conn.commit()
            # Зураг тусад нь хадгалах
            for i in ['1','2','3','4','5']:
                try: conn.execute(f"ALTER TABLE devices ADD COLUMN check_photo{i} TEXT")
                except Exception: pass
                f = request.files.get(f'check_photo{i}')
                if f and f.filename:
                    fn = save_file(f, 'devices')
                    if fn:
                        conn.execute(f"UPDATE devices SET check_photo{i}=? WHERE id=?", (fn, did))
                        conn.commit()
            conn.close()
            flash('Шалгалтын тохиргоо хадгалагдлаа!', 'success')
            return redirect(url_for('device_detail', did=did) + '#tab-check')
        photo = save_file(request.files.get('photo'), 'devices')
        pdf   = save_file(request.files.get('passport_pdf'), 'passports')
        mark_id = _resolve_mark(conn,
            request.form.get('manufacturer'),
            request.form.get('model'),
            request.form.get('category'))
        conn.execute("""
            UPDATE devices SET name=?,serial_number=?,mark_id=?,location=?,
            purchase_date=?,warranty_expiry=?,calibration_interval=?,
            status=?,notes=?,lab_id=?,web_link=?,method=?,max_temp=?,particular=?,
            measuring_time=?,measuring_limit=?,dimension=?,capacity=?,weight_kg=?,
            other_spec=?,power=?,frequency=?,voltage=?,specification=?,
            operating_state=?,received_date=?,
            check_standard=?,check_tolerance=?,check_enabled=?,stage=?,check_freq=?
            WHERE id=?
        """, (
            request.form['name'],
            request.form.get('serial_number') or None,
            mark_id,
            request.form.get('location'),
            request.form.get('purchase_date') or None,
            request.form.get('warranty_expiry') or None,
            int(request.form.get('calibration_interval') or 90),
            request.form.get('status','active'),
            request.form.get('notes'),
            request.form.get('lab_id') or None,
            request.form.get('web_link') or None,
            request.form.get('method') or None,
            request.form.get('max_temp') or None,
            request.form.get('particular') or None,
            request.form.get('measuring_time') or None,
            request.form.get('measuring_limit') or None,
            request.form.get('dimension') or None,
            request.form.get('capacity') or None,
            request.form.get('weight_kg') or None,
            request.form.get('other_spec') or None,
            request.form.get('power') or None,
            request.form.get('frequency') or None,
            request.form.get('voltage') or None,
            request.form.get('specification') or None,
            request.form.get('operating_state') or None,
            request.form.get('received_date') or None,
            request.form.get('check_standard') or None,
            request.form.get('check_tolerance') or None,
            1 if request.form.get('check_enabled') else 0,
            request.form.get('stage') or 'both',
            request.form.get('check_freq') or 'daily',
            did
        ))
        if photo: conn.execute("UPDATE devices SET photo=? WHERE id=?", (photo, did))
        if pdf:   conn.execute("UPDATE devices SET passport_pdf=? WHERE id=?", (pdf, did))
        conn.commit(); conn.close()
        flash('Шинэчлэгдлээ!' if lang=='mn' else 'Updated!', 'success')
        return redirect(url_for('device_detail', did=did))
    conn.close()
    return render_template('device/edit.html', device=device, marks=marks, lang=lang)

# ── USAGE ───────────────────────────────────────────────
@app.route('/usage/start/<int:did>', methods=['POST'])
@login_required
def usage_start(did):
    uid  = session.get('user_id', 0)
    conn = get_db()
    if session.get('role') != 'admin':
        perm = conn.execute("SELECT 1 FROM staff_device_permissions WHERE user_id=? AND device_id=?", (uid, did)).fetchone()
        if not perm:
            conn.close(); return jsonify({'error': 'Эрх байхгүй'}), 403
    # Тухайн төхөөрөмж дээр аль хэдийн идэвхтэй сесси байвал давхардуулахгүй
    already = conn.execute("SELECT id FROM usage_logs WHERE user_id=? AND device_id=? AND end_time IS NULL", (uid, did)).fetchone()
    if already:
        conn.close(); return jsonify({'error': 'Та энэ төхөөрөмж дээр аль хэдийн ажиллаж байна!'}), 400
    # Өөр хүн ашиглаж байвал түгжигдэнэ (нэг зэрэг нэг л хүн)
    busy = conn.execute("""SELECT u.name FROM usage_logs ul LEFT JOIN users u ON u.id=ul.user_id
                           WHERE ul.device_id=? AND ul.end_time IS NULL AND ul.user_id!=?""",
                        (did, uid)).fetchone()
    if busy:
        conn.close(); return jsonify({'error': f'Энэ төхөөрөмжийг {busy["name"]} ашиглаж байна. Дуустал хүлээнэ үү.'}), 400
    now = datetime.now().isoformat()
    conn.execute("INSERT INTO usage_logs(device_id,user_id,start_time) VALUES(?,?,?)", (did, uid, now))
    lid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit(); conn.close()
    return jsonify({'success': True, 'log_id': lid, 'start': now[11:16]})

@app.route('/usage/stop/<int:lid>', methods=['POST'])
@login_required
def usage_stop(lid):
    uid  = session.get('user_id', 0)
    conn = get_db()
    log  = conn.execute("SELECT * FROM usage_logs WHERE id=?", (lid,)).fetchone()
    if not log:
        conn.close(); return jsonify({'success': True})  # аль хэдийн устсан/байхгүй
    if log['user_id'] != uid and session.get('role') != 'admin':
        conn.close(); return jsonify({'error': 'Эрх байхгүй'}), 403
    if log['end_time']:
        conn.close(); return jsonify({'success': True})  # аль хэдийн хаагдсан
    now  = datetime.now()
    try:
        start = datetime.fromisoformat(log['start_time'])
        dur   = round((now - start).total_seconds() / 3600, 2)
    except Exception:
        dur = 0
    notes= request.json.get('notes','') if request.is_json else ''
    conn.execute("UPDATE usage_logs SET end_time=?,duration_hours=?,notes=? WHERE id=?",
                 (now.isoformat(), dur, notes, lid))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'duration': dur})

# ── CALIBRATION ─────────────────────────────────────────
@app.route('/devices/<int:did>/calibration/add', methods=['POST'])
@login_required
def calibration_add(did):
    lang = session.get('lang','mn')
    conn = get_db()
    doc  = save_file(request.files.get('document'), 'calibrations')
    conn.execute("""
        INSERT INTO calibrations(device_id,performed_by,calibration_date,next_date,result,document,notes)
        VALUES(?,?,?,?,?,?,?)
    """, (did, session['user_id'],
          request.form['calibration_date'],
          request.form.get('next_date') or None,
          request.form.get('result','passed'),
          doc, request.form.get('notes')))
    conn.commit(); conn.close()
    flash('Calibration бүртгэгдлээ!' if lang=='mn' else 'Calibration recorded!', 'success')
    return redirect(url_for('device_detail', did=did))

# ── REPAIR ──────────────────────────────────────────────
@app.route('/devices/<int:did>/repair/add', methods=['POST'])
@login_required
def repair_add(did):
    lang  = session.get('lang','mn')
    conn  = get_db()
    photo = save_file(request.files.get('photo'), 'repairs')
    status= request.form.get('status','new')
    conn.execute("""
        INSERT INTO repairs(device_id,reported_by,reported_date,description,company,repair_date,cost,photo,status,notes)
        VALUES(?,?,?,?,?,?,?,?,?,?)
    """, (did, session['user_id'],
          request.form['reported_date'],
          request.form['description'],
          request.form.get('company'),
          request.form.get('repair_date') or None,
          float(request.form.get('cost') or 0),
          photo, status, request.form.get('notes')))
    if status in ('new', 'repair'):
        conn.execute("UPDATE devices SET status='repair' WHERE id=?", (did,))
    conn.commit(); conn.close()
    flash('Засварын бүртгэл нэмэгдлээ!' if lang=='mn' else 'Repair added!', 'success')
    return redirect(url_for('device_detail', did=did))

# ── INTERNAL DAILY CHECK (жин г.м. дотоод шалгалт) ──────
@app.route('/devices/<int:did>/check/add', methods=['POST'])
@login_required
def device_check_add(did):
    lang = session.get('lang','mn')
    conn = get_db()
    photo = save_file(request.files.get('photo'), 'checks') if request.files.get('photo') else None
    conn.execute("""
        INSERT INTO device_checks(device_id,checked_by,check_date,standard_value,
        measured_value,tolerance,result,standard_value2,measured_value2,tolerance2,
        standard_value3,measured_value3,tolerance3,
        standard_value4,measured_value4,tolerance4,
        standard_value5,measured_value5,tolerance5,notes,photo)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (did, session.get('user_id', 0),
          request.form.get('check_date') or date.today().isoformat(),
          request.form.get('standard_value') or None,
          request.form.get('measured_value') or None,
          request.form.get('tolerance') or None,
          request.form.get('result','pass'),
          request.form.get('standard_value2') or None,
          request.form.get('measured_value2') or None,
          request.form.get('tolerance2') or None,
          request.form.get('standard_value3') or None,
          request.form.get('measured_value3') or None,
          request.form.get('tolerance3') or None,
          request.form.get('standard_value4') or None,
          request.form.get('measured_value4') or None,
          request.form.get('tolerance4') or None,
          request.form.get('standard_value5') or None,
          request.form.get('measured_value5') or None,
          request.form.get('tolerance5') or None,
          request.form.get('notes') or None,
          photo))
    conn.commit(); conn.close()
    flash('Дотоод шалгалт бүртгэгдлээ!' if lang=='mn' else 'Internal check recorded!', 'success')
    return redirect(url_for('device_detail', did=did) + '#tab-check')

@app.route('/devices/check/<int:cid>/delete', methods=['POST'])
@senior_required
def device_check_delete(cid):
    conn = get_db()
    row = conn.execute("SELECT device_id FROM device_checks WHERE id=?", (cid,)).fetchone()
    did = row['device_id'] if row else None
    conn.execute("DELETE FROM device_checks WHERE id=?", (cid,))
    conn.commit(); conn.close()
    if did:
        return redirect(url_for('device_detail', did=did) + '#tab-check')
    return redirect(url_for('devices'))

@app.route('/devices/<int:did>/calib-check', methods=['POST'])
@lab_required
def device_calib_check_add(did):
    lang = session.get('lang','mn')
    conn = get_db()
    check_date = request.form.get('check_date') or date.today().isoformat()
    notes = request.form.get('notes') or None
    pairs = []
    for i in range(1, 6):
        xs = request.form.get(f'x{i}','').strip()
        ys = request.form.get(f'y{i}','').strip()
        if xs and ys:
            try: pairs.append((float(xs), float(ys)))
            except ValueError: pass
    result = 'fail'
    slope_b = intercept_a = r_squared = None
    if len(pairs) >= 2:
        n = len(pairs)
        sx  = sum(p[0] for p in pairs)
        sy  = sum(p[1] for p in pairs)
        sxy = sum(p[0]*p[1] for p in pairs)
        sx2 = sum(p[0]**2 for p in pairs)
        sy2 = sum(p[1]**2 for p in pairs)
        denom = n*sx2 - sx**2
        if denom != 0:
            slope_b = (n*sxy - sx*sy) / denom
            intercept_a = (sy - slope_b*sx) / n
            r_num = (n*sxy - sx*sy)**2
            r_den = denom * (n*sy2 - sy**2)
            r_squared = r_num/r_den if r_den != 0 else None
            if r_squared is not None and r_squared >= 0.999:
                result = 'pass'
    xs = [p[0] for p in pairs] + [None]*(5-len(pairs))
    ys = [p[1] for p in pairs] + [None]*(5-len(pairs))
    try: conn.execute("""
        CREATE TABLE IF NOT EXISTS device_calorimeter_checks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_id INTEGER NOT NULL REFERENCES devices(id),
            checked_by INTEGER REFERENCES users(id),
            check_date TEXT NOT NULL,
            bomb_no TEXT,
            old_ee REAL,
            new_ee REAL,
            ee_change_pct REAL,
            mass1 REAL, cv1 REAL,
            mass2 REAL, cv2 REAL,
            mass3 REAL, cv3 REAL,
            mass4 REAL, cv4 REAL,
            mass5 REAL, cv5 REAL,
            mass6 REAL, cv6 REAL,
            mass7 REAL, cv7 REAL,
            cv_avg REAL,
            cv_range REAL,
            cv_range_pct REAL,
            cv_rsd REAL,
            n_samples INTEGER,
            result TEXT DEFAULT 'pass',
            ee_warning INTEGER DEFAULT 0,
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)
    except Exception: pass
    try: conn.execute("ALTER TABLE device_calib_checks ADD COLUMN x1 REAL")
    except Exception: pass
    conn.execute("""
        INSERT INTO device_calib_checks
        (device_id,checked_by,check_date,x1,y1,x2,y2,x3,y3,x4,y4,x5,y5,
         slope_b,intercept_a,r_squared,result,notes)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (did, session.get('user_id',0), check_date,
          xs[0],ys[0], xs[1],ys[1], xs[2],ys[2], xs[3],ys[3], xs[4],ys[4],
          slope_b, intercept_a, r_squared, result, notes))
    conn.commit(); conn.close()
    flash('Калибровкийн шалгалт бүртгэгдлээ!', 'success')
    return redirect(url_for('device_detail', did=did) + '#tab-check')

@app.route('/devices/calib-check/<int:cid>/delete', methods=['POST'])
@senior_required
def device_calib_check_delete(cid):
    conn = get_db()
    row = conn.execute("SELECT device_id FROM device_calib_checks WHERE id=?", (cid,)).fetchone()
    did = row['device_id'] if row else None
    conn.execute("DELETE FROM device_calib_checks WHERE id=?", (cid,))
    conn.commit(); conn.close()
    if did:
        return redirect(url_for('device_detail', did=did) + '#tab-check')
    return redirect(url_for('devices'))

@app.route('/devices/<int:did>/calorimeter-check', methods=['POST'])
@lab_required
def device_calorimeter_check_add(did):
    import math
    conn = get_db()
    check_date = request.form.get('check_date') or date.today().isoformat()
    bomb_no    = request.form.get('bomb_no') or None
    notes      = request.form.get('notes') or None
    old_ee = request.form.get('old_ee') or None
    new_ee = request.form.get('new_ee') or None
    try: old_ee = float(old_ee)
    except: old_ee = None
    try: new_ee = float(new_ee)
    except: new_ee = None
    ee_change_pct = None
    ee_warning = 0
    if old_ee and new_ee and old_ee != 0:
        ee_change_pct = abs(new_ee - old_ee) / old_ee * 100
        if ee_change_pct >= 0.2:
            ee_warning = 1
    masses, cvs = [], []
    for i in range(1, 8):
        ms = request.form.get(f'mass{i}','').strip()
        cs = request.form.get(f'cv{i}','').strip()
        if ms and cs:
            try:
                masses.append(float(ms))
                cvs.append(float(cs))
            except ValueError: pass
    cv_avg = cv_range = cv_range_pct = cv_rsd = None
    result = 'fail'
    n = len(cvs)
    if n >= 2:
        cv_avg = sum(cvs) / n
        cv_range = max(cvs) - min(cvs)
        cv_range_pct = cv_range / cv_avg * 100 if cv_avg else None
        variance = sum((v - cv_avg)**2 for v in cvs) / (n - 1)
        cv_rsd = math.sqrt(variance) / cv_avg * 100 if cv_avg else None
        if cv_range_pct is not None and cv_range_pct <= 0.5:
            result = 'pass'
    def _f(lst, i): return lst[i] if i < len(lst) else None
    conn.execute("""
        INSERT INTO device_calorimeter_checks
        (device_id,checked_by,check_date,bomb_no,old_ee,new_ee,ee_change_pct,
         mass1,cv1,mass2,cv2,mass3,cv3,mass4,cv4,mass5,cv5,mass6,cv6,mass7,cv7,
         cv_avg,cv_range,cv_range_pct,cv_rsd,n_samples,result,ee_warning,notes)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (did, session.get('user_id',0), check_date, bomb_no, old_ee, new_ee, ee_change_pct,
          _f(masses,0),_f(cvs,0), _f(masses,1),_f(cvs,1), _f(masses,2),_f(cvs,2),
          _f(masses,3),_f(cvs,3), _f(masses,4),_f(cvs,4), _f(masses,5),_f(cvs,5),
          _f(masses,6),_f(cvs,6),
          cv_avg, cv_range, cv_range_pct, cv_rsd, n, result, ee_warning, notes))
    conn.commit(); conn.close()
    flash('Калориметрийн шалгалт бүртгэгдлээ!', 'success')
    return redirect(url_for('device_detail', did=did) + '#tab-check')

@app.route('/devices/calorimeter-check/<int:cid>/delete', methods=['POST'])
@senior_required
def device_calorimeter_check_delete(cid):
    conn = get_db()
    row = conn.execute("SELECT device_id FROM device_calorimeter_checks WHERE id=?", (cid,)).fetchone()
    did = row['device_id'] if row else None
    conn.execute("DELETE FROM device_calorimeter_checks WHERE id=?", (cid,))
    conn.commit(); conn.close()
    if did:
        return redirect(url_for('device_detail', did=did) + '#tab-check')
    return redirect(url_for('devices'))

# ── НЭГДСЭН ӨДӨР ТУТМЫН БАТАЛГААЖУУЛАЛТ ──────────────────
@app.route('/checks')
@login_required
def checks_page():
    """Бүх тоног төхөөрөмжийн өдөр тутмын баталгаажуулалтыг нэг хуудсанд."""
    lang = session.get('lang','mn')
    sel_date = request.args.get('date') or date.today().isoformat()
    conn = get_db()
    devices = conn.execute("""
        SELECT d.*, dm.manufacturer, dm.model FROM devices d
        LEFT JOIN device_marks dm ON d.mark_id=dm.id
        WHERE COALESCE(d.check_enabled,1)=1 AND d.status NOT IN ('archived','replaced','decommissioned')
        ORDER BY COALESCE(d.check_freq,'daily'), d.name
    """).fetchall()
    # Weekly-д сонгосон өдрийн долоо хоногийн эхний өдрийг тооцно
    from datetime import datetime, timedelta
    sel_dt = datetime.strptime(sel_date, '%Y-%m-%d').date()
    week_start = (sel_dt - timedelta(days=sel_dt.weekday())).isoformat()
    week_end   = (sel_dt + timedelta(days=6 - sel_dt.weekday())).isoformat()
    # Сонгосон өдрийн шалгалтуудыг device_id-аар индекслэх
    rows = conn.execute("""
        SELECT ch.*, u.name as uname FROM device_checks ch
        LEFT JOIN users u ON u.id=ch.checked_by
        WHERE ch.check_date=? ORDER BY ch.id DESC
    """, (sel_date,)).fetchall()
    week_rows = conn.execute("""
        SELECT ch.*, u.name as uname FROM device_checks ch
        LEFT JOIN users u ON u.id=ch.checked_by
        WHERE ch.check_date BETWEEN ? AND ? ORDER BY ch.check_date DESC, ch.id DESC
    """, (week_start, week_end)).fetchall()
    conn.close()
    today_checks = {}
    for r in rows:
        today_checks.setdefault(r['device_id'], r)
    week_checks = {}
    for r in week_rows:
        week_checks.setdefault(r['device_id'], r)
    return render_template('device/checks.html',
        devices=devices, today_checks=today_checks, week_checks=week_checks,
        sel_date=sel_date, today=date.today().isoformat(),
        week_start=week_start, week_end=week_end, lang=lang)

@app.route('/checks/save', methods=['POST'])
@login_required
def checks_save():
    """Нэг өдрийн бүх төхөөрөмжийн шалгалтыг нэг дор хадгална."""
    lang = session.get('lang','mn')
    if session.get('role') == 'guest':
        return redirect(url_for('checks_page'))
    sel_date = request.form.get('check_date') or date.today().isoformat()
    uid = session.get('user_id', 0)
    conn = get_db()
    saved = 0
    for did in request.form.getlist('device_id'):
        measured = (request.form.get(f'measured_{did}') or '').strip()
        if not measured:
            continue  # хэмжсэн утга оруулаагүй бол алгасна
        # Тухайн өдрийн хуучин бичлэгийг устгаад шинээр бичих (давхардахгүй)
        conn.execute("DELETE FROM device_checks WHERE device_id=? AND check_date=?", (did, sel_date))
        cal_adj = 1 if request.form.get(f'cal_adj_{did}') else 0
        conn.execute("""
            INSERT INTO device_checks(device_id,checked_by,check_date,standard_value,
            measured_value,tolerance,result,calibration_adjusted,
            standard_value2,measured_value2,tolerance2,
            standard_value3,measured_value3,tolerance3,
            standard_value4,measured_value4,tolerance4,
            standard_value5,measured_value5,tolerance5,notes)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (did, uid, sel_date,
              request.form.get(f'standard_{did}') or None,
              measured,
              request.form.get(f'tolerance_{did}') or None,
              request.form.get(f'result_{did}','pass'),
              cal_adj,
              request.form.get(f'standard2_{did}') or None,
              request.form.get(f'measured2_{did}') or None,
              request.form.get(f'tolerance2_{did}') or None,
              request.form.get(f'standard3_{did}') or None,
              request.form.get(f'measured3_{did}') or None,
              request.form.get(f'tolerance3_{did}') or None,
              request.form.get(f'standard4_{did}') or None,
              request.form.get(f'measured4_{did}') or None,
              request.form.get(f'tolerance4_{did}') or None,
              request.form.get(f'standard5_{did}') or None,
              request.form.get(f'measured5_{did}') or None,
              request.form.get(f'tolerance5_{did}') or None,
              request.form.get(f'notes_{did}') or None))
        saved += 1
    conn.commit(); conn.close()
    flash(f'{saved} төхөөрөмжийн баталгаажуулалт хадгалагдлаа!', 'success')
    return redirect(url_for('checks_page', date=sel_date))

@app.route('/checks/monthly')
@login_required
def checks_monthly():
    """Сарын харагдац — бүх жингийн шалгалтыг нэг хүснэгтэд (Excel Sheet 1 загвараар)."""
    import calendar as cal_mod
    lang = session.get('lang', 'mn')
    today = date.today()
    year  = int(request.args.get('year',  today.year))
    month = int(request.args.get('month', today.month))
    _, days_in_month = cal_mod.monthrange(year, month)
    d_from = f"{year}-{month:02d}-01"
    d_to   = f"{year}-{month:02d}-{days_in_month:02d}"
    conn = get_db()
    devices = conn.execute("""
        SELECT d.*, dm.manufacturer, dm.model FROM devices d
        LEFT JOIN device_marks dm ON d.mark_id=dm.id
        WHERE COALESCE(d.check_enabled,1)=1
          AND d.status NOT IN ('archived','replaced','decommissioned')
        ORDER BY COALESCE(d.check_freq,'daily'), d.name
    """).fetchall()
    rows = conn.execute("""
        SELECT ch.*, u.name as uname FROM device_checks ch
        LEFT JOIN users u ON u.id=ch.checked_by
        WHERE ch.check_date BETWEEN ? AND ?
    """, (d_from, d_to)).fetchall()
    conn.close()
    # (device_id, day) → check record
    grid = {}
    for r in rows:
        day = int(r['check_date'].split('-')[2])
        grid[(r['device_id'], day)] = r
    return render_template('device/checks_monthly.html',
        devices=devices, grid=grid,
        year=year, month=month, days_in_month=days_in_month,
        today=today, lang=lang)

@app.route('/checks/export')
@login_required
def checks_export():
    """Баталгаажуулалтын бүртгэлийг Excel болгон татах."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    d_from = request.args.get('from') or date.today().replace(day=1).isoformat()
    d_to   = request.args.get('to') or date.today().isoformat()
    conn = get_db()
    rows = conn.execute("""
        SELECT ch.check_date, d.name as device_name, d.lab_id,
               ch.standard_value, ch.measured_value, ch.tolerance,
               ch.result, ch.calibration_adjusted, ch.notes, u.name as uname
        FROM device_checks ch
        LEFT JOIN devices d ON d.id=ch.device_id
        LEFT JOIN users u ON u.id=ch.checked_by
        WHERE ch.check_date BETWEEN ? AND ?
        ORDER BY ch.check_date DESC, d.name
    """, (d_from, d_to)).fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Баталгаажуулалт'
    headers = ['Огноо','Төхөөрөмж','Лаб дугаар','Стандарт','Хэмжсэн','Зөвшөөрөх зөрүү','Үр дүн','Тохируулга','Шалгасан','Тэмдэглэл']
    ws.append(headers)
    hf = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='1A2744')
    for c in ws[1]:
        c.font = hf; c.fill = fill; c.alignment = Alignment(horizontal='center')
    for r in rows:
        ws.append([r['check_date'], r['device_name'], r['lab_id'] or '',
                   r['standard_value'] or '', r['measured_value'] or '',
                   r['tolerance'] or '',
                   'Тэнцсэн' if r['result']=='pass' else 'Тэнцээгүй',
                   'Тийм' if r['calibration_adjusted'] else '',
                   r['uname'] or '', r['notes'] or ''])
    widths = [12, 24, 12, 14, 14, 16, 12, 10, 18, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f'baталгаажуулалт_{d_from}_{d_to}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/repair/<int:rid>/close', methods=['POST'])
@login_required
def repair_close(rid):
    conn = get_db()
    rep  = conn.execute("SELECT * FROM repairs WHERE id=?", (rid,)).fetchone()
    if rep:
        conn.execute("UPDATE repairs SET status='done' WHERE id=?", (rid,))
        conn.execute("UPDATE devices SET status='active' WHERE id=?", (rep['device_id'],))
        conn.commit()
    conn.close()
    return jsonify({'success': True})

# ── STAFF ───────────────────────────────────────────────
@app.route('/staff')
@login_required
def staff_list():
    lang = session.get('lang','mn')
    conn = get_db()
    is_admin = session.get('role') == 'admin'
    role_filter = '' if is_admin else "AND role != 'admin'"
    # Амарсан хүн (is_active=0, is_dismissed=0) энд идэвхгүй байдлаар ХАРАГДАНА.
    # Ажлаас гарсан хүн (is_dismissed=1) зөвхөн архивт харагдана.
    users_a = conn.execute(f"""
        SELECT * FROM users WHERE shift='A' AND is_dismissed=0 AND is_deleted=0 {role_filter}
        ORDER BY is_active DESC, CASE role
            WHEN 'admin' THEN 1 WHEN 'senior' THEN 2 WHEN 'staff' THEN 3
            WHEN 'preparer' THEN 4 WHEN 'geologist' THEN 5 ELSE 6 END, name
    """).fetchall()
    users_b = conn.execute(f"""
        SELECT * FROM users WHERE shift='B' AND is_dismissed=0 AND is_deleted=0 {role_filter}
        ORDER BY is_active DESC, CASE role
            WHEN 'admin' THEN 1 WHEN 'senior' THEN 2 WHEN 'staff' THEN 3
            WHEN 'preparer' THEN 4 WHEN 'geologist' THEN 5 ELSE 6 END, name
    """).fetchall()
    users_none = conn.execute(f"""
        SELECT * FROM users
        WHERE is_dismissed=0 AND is_deleted=0 AND (shift IS NULL OR shift='') {role_filter}
        ORDER BY is_active DESC, CASE role
            WHEN 'admin' THEN 1 WHEN 'senior' THEN 2 WHEN 'staff' THEN 3
            WHEN 'preparer' THEN 4 WHEN 'geologist' THEN 5 ELSE 6 END, name
    """).fetchall()
    _bayj_raw = conn.execute(
        "SELECT * FROM users WHERE role='bayjuulach' AND is_dismissed=0 AND is_deleted=0").fetchall()
    _pos_rank = ['Үйлдвэрийн дарга','Ашиглалтын ахлах инженер','Ашиглалтын инженер',
                 'Цахилгаан автоматжуулалтын инженер','Удирдлагын оператор',
                 'Орчуулагч, бичиг хэргийн мэргэжилтэн']
    bayj_users = sorted(_bayj_raw, key=lambda u: (
        _pos_rank.index(u['position'].strip()) if u['position'] and u['position'].strip() in _pos_rank else 99,
        u['name']
    ))
    conn.close()
    return render_template('admin/staff_list.html', users_a=users_a, users_b=users_b, users_none=users_none, bayj_users=bayj_users, lang=lang)

@app.route('/staff/add', methods=['GET','POST'])
@senior_required
def staff_add():
    lang    = session.get('lang','mn')
    conn    = get_db()
    devices = conn.execute("SELECT * FROM devices ORDER BY name").fetchall()
    if request.method == 'POST':
        existing = conn.execute("SELECT id FROM users WHERE employee_id=?", (request.form['employee_id'],)).fetchone()
        if existing:
            conn.close()
            flash(f"'{request.form['employee_id']}' ID аль хэдийн бүртгэлтэй байна!", 'error')
            devices2 = get_db().execute("SELECT * FROM devices ORDER BY name").fetchall()
            return render_template('admin/staff_add.html', devices=devices2, lang=lang)
        try:
            photo = save_file(request.files.get('photo'), 'staff')
            pw    = hash_password(request.form['password'])
            # Senior бол зөвхөн 'staff' эрх өгч болно
            role_to_set = request.form.get('role','staff')
            if session.get('role') == 'senior' and role_to_set in ('admin','senior'):
                role_to_set = 'staff'
            is_bayj = role_to_set == 'bayjuulach'
            can_reg = 1 if (is_bayj and request.form.get('can_register')) else 0
            can_view = 1 if (is_bayj and request.form.get('can_view_result')) else 0
            conn.execute("""
                INSERT INTO users(employee_id,name,position,phone,email,photo,role,password_hash,joined_date,shift,can_register,can_view_result)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                request.form['employee_id'], request.form['name'],
                request.form.get('position'), request.form.get('phone'),
                request.form.get('email'), photo,
                role_to_set, pw,
                request.form.get('joined_date') or None,
                request.form.get('shift') or None,
                can_reg, can_view
            ))
            uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            # Геологичид (харилцагч) тоног ашиглах эрх олгохгүй
            if role_to_set != 'geologist':
                for did in request.form.getlist('device_permissions'):
                    conn.execute("INSERT OR IGNORE INTO staff_device_permissions(user_id,device_id) VALUES(?,?)", (uid, did))
            conn.commit(); conn.close()
            flash('Ажилтан нэмэгдлээ!' if lang=='mn' else 'Staff added!', 'success')
            return redirect(url_for('staff_list'))
        except Exception as e:
            conn.close()
            flash(f'Алдаа: {str(e)}', 'error')
            devices2 = get_db().execute("SELECT * FROM devices ORDER BY name").fetchall()
            return render_template('admin/staff_add.html', devices=devices2, lang=lang)
    conn.close()
    return render_template('admin/staff_add.html', devices=devices, lang=lang)


@app.route('/staff/<int:uid>')
@login_required
def staff_detail(uid):
    if session.get('role') not in ('admin','senior','guest') and session.get('user_id') != uid:
        return redirect(url_for('dashboard'))
    lang = session.get('lang','mn')
    conn = get_db()
    target  = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not target:
        conn.close()
        return redirect(url_for('staff_list'))

    # ── ХАРИЛЦАГЧ (геологи + баяжуулах): зөвхөн дээж бүртгэлийн мэдээлэл ──
    if target['role'] in ('geologist', 'bayjuulach'):
        # Тоолол нь АЖЛААР биш ДЭЭЖЭЭР явна: нэг ажилд quantity дээж байдаг
        # (10 дээжтэй ажил = 10, өмнө нь 1 гэж бодогддог байсан).
        QTY = "COALESCE(SUM(COALESCE(quantity,1)),0)"
        geo_total = conn.execute(
            f"SELECT {QTY} FROM geo_samples WHERE registered_by=?", (uid,)).fetchone()[0]
        geo_done  = conn.execute(
            f"SELECT {QTY} FROM geo_samples WHERE registered_by=? AND status='done'", (uid,)).fetchone()[0]
        geo_jobs  = conn.execute(
            "SELECT COUNT(*) FROM geo_samples WHERE registered_by=?", (uid,)).fetchone()[0]
        geo_active = geo_total - geo_done
        geo_by_type = conn.execute(f"""
            SELECT sample_type, {QTY} as cnt FROM geo_samples
            WHERE registered_by=? GROUP BY sample_type ORDER BY cnt DESC
        """, (uid,)).fetchall()
        geo_recent = conn.execute("""
            SELECT g.*, sr.lab_number FROM geo_samples g
            LEFT JOIN sample_receipt sr ON sr.geo_sample_id=g.id
            WHERE g.registered_by=? ORDER BY g.created_at DESC LIMIT 30
        """, (uid,)).fetchall()
        geo_monthly = conn.execute(f"""
            SELECT strftime('%Y-%m', created_at) as period, {QTY} as cnt
            FROM geo_samples WHERE registered_by=? AND created_at IS NOT NULL
            GROUP BY period ORDER BY period
        """, (uid,)).fetchall()
        # Долоо хоногоор (ISO: %Y-W%W) — сүүлийн 26 долоо хоног
        geo_weekly = conn.execute(f"""
            SELECT strftime('%Y-W%W', created_at) as period, {QTY} as cnt,
                   MIN(date(created_at)) as first_day
            FROM geo_samples
            WHERE registered_by=? AND created_at IS NOT NULL
              AND date(created_at) >= date('now','-182 days')
            GROUP BY period ORDER BY period
        """, (uid,)).fetchall()
        # Үр дүн харсан лог (геологи + баяжуулах)
        view_total = conn.execute("SELECT COUNT(*) FROM result_view_log WHERE user_id=?", (uid,)).fetchone()[0]
        view_log = conn.execute("""
            SELECT vl.viewed_at, sr.lab_number, g.sample_name
            FROM result_view_log vl
            JOIN sample_receipt sr ON sr.id=vl.receipt_id
            JOIN geo_samples g ON g.id=sr.geo_sample_id
            WHERE vl.user_id=? ORDER BY vl.viewed_at DESC LIMIT 20
        """, (uid,)).fetchall()
        conn.close()
        return render_template('staff/detail_geologist.html', target=target, lang=lang,
                               geo_total=geo_total, geo_done=geo_done, geo_active=geo_active,
                               geo_by_type=geo_by_type, geo_recent=geo_recent,
                               geo_monthly=geo_monthly, geo_weekly=geo_weekly, geo_jobs=geo_jobs,
                               view_total=view_total, view_log=view_log)

    logs    = conn.execute("""
        SELECT ul.*, d.name as dname FROM usage_logs ul
        LEFT JOIN devices d ON d.id=ul.device_id
        WHERE ul.user_id=? ORDER BY ul.start_time DESC LIMIT 50
    """, (uid,)).fetchall()
    my_devs = conn.execute("""
        SELECT d.* FROM devices d
        JOIN staff_device_permissions p ON p.device_id=d.id
        WHERE p.user_id=?
    """, (uid,)).fetchall()
    device_usage = conn.execute("""
        SELECT d.id as device_id, d.name as device_name,
               COUNT(*) as sessions,
               COALESCE(SUM(CAST((julianday(ul.end_time)-julianday(ul.start_time))*1440 AS INTEGER)),0) as total_min,
               MAX(ul.start_time) as last_used
        FROM usage_logs ul
        JOIN devices d ON d.id=ul.device_id
        WHERE ul.user_id=? AND ul.end_time IS NOT NULL
        GROUP BY d.id ORDER BY total_min DESC
    """, (uid,)).fetchall()
    # Stats
    # ── Шинжилсэн дээж — ШИНЖИЛГЭЭ ХИЙСНЭЭР тоолно ────────────────────
    # Урьд нь зөвхөн "Дууслаа" ✓ товч дарсан мөрийг тоолдог байсан тул
    # хэмжилт хийсэн ч ✓ дараагүй ажил тоологдохгүй байв.
    _op_any = ' OR '.join(f'{op}=?' for op, _l, _f in ANALYSIS_OPS)
    _op_args = tuple(uid for _ in ANALYSIS_OPS)
    # Нийт оролцсон дээж — нэг дээж дээр хэдэн ч шинжилгээ хийсэн нэг л удаа
    total_analyzed = conn.execute(
        f"""SELECT COUNT(DISTINCT receipt_id || '-' || row_num)
            FROM sample_entries WHERE {_op_any}""", _op_args).fetchone()[0]
    # Шинжилгээний төрөл тус бүрээр
    by_analysis = []
    for op, lbl, _f in ANALYSIS_OPS:
        n = conn.execute(
            f"SELECT COUNT(DISTINCT receipt_id || '-' || row_num) "
            f"FROM sample_entries WHERE {op}=?", (uid,)).fetchone()[0]
        if n:
            by_analysis.append({'label': lbl, 'count': n})
    by_analysis.sort(key=lambda x: -x['count'])
    # "Дуусгасан" ба "Баталгаажуулсан" тоог мөн ДЭЭЖЭЭР тоолно.
    # Урьд нь COUNT(*) — өөрөөр хэлбэл зэрэгцээ, давталтын мөр тус бүрийг
    # тоолдог байсан тул "Шинжилсэн дээж 287 / дуусгасан 430" гэх мэт
    # дээрхээсээ ИХ тоо гарч, ойлгомжгүй болдог байв. Одоо гурвуулаа
    # "хэдэн дээж" гэсэн нэг суурьтай.
    total_done     = conn.execute(
        "SELECT COUNT(DISTINCT receipt_id || '-' || row_num) FROM sample_entries "
        "WHERE done_by=? AND row_status IN ('done','approved')", (uid,)).fetchone()[0]
    total_approved = conn.execute(
        "SELECT COUNT(DISTINCT receipt_id || '-' || row_num) FROM sample_entries "
        "WHERE approved_by=? AND row_status='approved'", (uid,)).fetchone()[0]
    total_hours    = conn.execute("SELECT COALESCE(SUM(duration_hours),0) FROM usage_logs WHERE user_id=? AND end_time IS NOT NULL", (uid,)).fetchone()[0]
    # QC radar
    qc_rows = conn.execute("""
        SELECT qs.parameter,
               SUM(CASE WHEN ABS(e.mad - e2.mad) <= qs.tolerance THEN 1 ELSE 0 END)*1.0/COUNT(*)*100 as pct
        FROM sample_entries e
        JOIN sample_entries e2 ON e2.receipt_id=e.receipt_id AND e2.row_num=e.row_num AND e2.is_duplicate=1
        JOIN qc_settings qs ON qs.parameter='Mad'
        WHERE e.is_duplicate=0 AND e.done_by=? AND e.mad IS NOT NULL AND e2.mad IS NOT NULL
        LIMIT 1
    """, (uid,)).fetchall()
    qc_radar = {}
    # ── Бэлтгэсэн дээж (ажлаар биш дээжээр) ──────────────
    # prep_by шинэ багана. Хуучин бичлэгт байхгүй тул нэрээр нөхөж тооцно.
    PREP_WHERE = """
        FROM sample_receipt sr JOIN geo_samples g ON g.id=sr.geo_sample_id
        WHERE sr.prep_done_at IS NOT NULL
          AND (sr.prep_by=? OR (sr.prep_by IS NULL AND sr.prep_operator=?))
    """
    prep_args = (uid, target['name'])
    total_prepared = conn.execute(
        f"SELECT COALESCE(SUM(COALESCE(g.quantity,1)),0) {PREP_WHERE}", prep_args).fetchone()[0]

    # Monthly / weekly counts — шинжилсэн ба бэлтгэсэн дээжийн тоо
    def series(fmt, days=None, kind='done'):
        if kind == 'done':
            # ДЭЭЖЭЭР тоолно — дээрх "Шинжилсэн дээж" картуудтай нэг суурьтай.
            # Урьд нь COUNT(*) байсан тул зэрэгцээ, давталтын мөр тус бүр
            # тоологдож, график нь картаас (287) их тоо (430) харуулдаг байв.
            sql = f"""
                SELECT strftime('{fmt}', se.done_at) as period,
                       COUNT(DISTINCT se.receipt_id || '-' || se.row_num) as cnt
                FROM sample_entries se
                WHERE se.done_by=? AND se.row_status IN ('done','approved')
                  AND se.done_at IS NOT NULL
                  {"AND date(se.done_at) >= date('now', ?)" if days else ""}
                GROUP BY period ORDER BY period
            """
            args = (uid, f'-{days} days') if days else (uid,)
        else:
            sql = f"""
                SELECT strftime('{fmt}', sr.prep_done_at) as period,
                       COALESCE(SUM(COALESCE(g.quantity,1)),0) as cnt
                {PREP_WHERE}
                  {"AND date(sr.prep_done_at) >= date('now', ?)" if days else ""}
                GROUP BY period ORDER BY period
            """
            args = prep_args + ((f'-{days} days',) if days else ())
        return conn.execute(sql, args).fetchall()

    monthly_all  = series('%Y-%m', kind='done')
    monthly_6    = series('%Y-%m', days=183, kind='done')
    monthly_12   = series('%Y-%m', days=366, kind='done')
    weekly_all   = series('%Y-W%W', days=182, kind='done')
    prep_monthly = series('%Y-%m', kind='prep')
    prep_weekly  = series('%Y-W%W', days=182, kind='prep')
    conn.close()
    return render_template('staff/detail.html', target=target, logs=logs, my_devices=my_devs,
                           device_usage=device_usage, lang=lang,
                           total_done=total_done, total_approved=total_approved, total_hours=total_hours,
                           total_analyzed=total_analyzed, by_analysis=by_analysis,
                           total_prepared=total_prepared,
                           qc_radar=qc_radar, monthly_6=monthly_6, monthly_12=monthly_12,
                           monthly_all=monthly_all, weekly_all=weekly_all,
                           prep_monthly=prep_monthly, prep_weekly=prep_weekly)

# ── INTERNAL QC ─────────────────────────────────────────
@app.route('/internal-qc')
@lab_required
def internal_qc_list():
    conn = get_db()
    rows = conn.execute("""
        SELECT iq.*,
            u.name as assigned_name,
            sr1.lab_number as lab1, g1.sample_name as sname1,
            sr2.lab_number as lab2, g2.sample_name as sname2
        FROM internal_qc iq
        LEFT JOIN users u ON u.id=iq.assigned_to
        LEFT JOIN sample_receipt sr1 ON sr1.id=iq.receipt_id_1
        LEFT JOIN geo_samples g1 ON g1.id=sr1.geo_sample_id
        LEFT JOIN sample_receipt sr2 ON sr2.id=iq.receipt_id_2
        LEFT JOIN geo_samples g2 ON g2.id=sr2.geo_sample_id
        ORDER BY iq.created_at DESC
    """).fetchall()
    conn.close()
    return render_template('analysis/internal_qc.html', rows=rows, lang=session.get('lang','mn'), role=session.get('role'))

@app.route('/internal-qc/create', methods=['POST'])
@senior_required
def internal_qc_create():
    import random
    conn = get_db()
    from datetime import datetime, timedelta
    d_to = datetime.now().strftime('%Y-%m-%d')
    d_from = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    candidates = conn.execute("""
        SELECT DISTINCT sr.id FROM sample_receipt sr
        JOIN geo_samples g ON g.id=sr.geo_sample_id
        JOIN sample_entries se ON se.receipt_id=sr.id
        WHERE g.sample_type IN ('PIT','STOCKPILE','EXPORT','CONTROL')
        AND se.is_duplicate=0
        AND se.row_status IN ('done','approved')
        AND sr.received_date BETWEEN ? AND ?
    """, (d_from, d_to)).fetchall()
    if len(candidates) < 1:
        conn.close()
        flash('Өмнөх 7 хоногт давтан шинжлэх хангалттай дээж байхгүй байна', 'error')
        return redirect(url_for('analysis'))
    picked = random.sample([r['id'] for r in candidates], 1)
    def pick_row(rid):
        rows = conn.execute(
            "SELECT row_num FROM sample_entries WHERE receipt_id=? AND is_duplicate=0 AND row_status IN ('done','approved')",
            (rid,)).fetchall()
        if not rows: return 1
        return random.choice([r['row_num'] for r in rows])
    rn1 = pick_row(picked[0])
    assigned = request.form.get('assigned_to') or session['user_id']
    last = conn.execute("SELECT qc_number FROM internal_qc WHERE qc_number LIKE 'QC%' ORDER BY id DESC LIMIT 1").fetchone()
    if last and last['qc_number'] and last['qc_number'][2:].isdigit():
        seq = int(last['qc_number'][2:]) + 1
    else:
        seq = 1
    qc_number = f'QC{seq:03d}'
    cur = conn.execute("""INSERT INTO internal_qc (qc_number, triggered_date, receipt_id_1, receipt_id_2, row_num_1, row_num_2, assigned_to, created_by)
        VALUES (?,?,?,NULL,?,NULL,?,?)""", (qc_number, d_to, picked[0], rn1, assigned, session['user_id']))
    qc_id = cur.lastrowid
    conn.commit()
    conn.close()
    flash('Дотоод QC үүслээ — доорх жагсаалтаас сонгож шинжилгээ хийнэ үү', 'success')
    return redirect(url_for('analysis'))

@app.route('/internal-qc/<int:qc_id>/measure')
@lab_required
def internal_qc_measure(qc_id):
    conn = get_db()
    qc = conn.execute("""
        SELECT iq.*,
            sr1.lab_number as lab1, g1.sample_name as sname1,
            sr2.lab_number as lab2, g2.sample_name as sname2
        FROM internal_qc iq
        LEFT JOIN sample_receipt sr1 ON sr1.id=iq.receipt_id_1
        LEFT JOIN geo_samples g1 ON g1.id=sr1.geo_sample_id
        LEFT JOIN sample_receipt sr2 ON sr2.id=iq.receipt_id_2
        LEFT JOIN geo_samples g2 ON g2.id=sr2.geo_sample_id
        WHERE iq.id=?
    """, (qc_id,)).fetchone()
    if not qc:
        conn.close()
        flash('Бүртгэл олдсонгүй', 'error')
        return redirect(url_for('internal_qc_list'))
    PARAMS = ['adb','vdb','fcdb','sdb','qdb','fsi','g_index']
    DB_LABELS = {'adb':'Adb','vdb':'Vdb','fcdb':'FCdb','sdb':'Sdb','qdb':'Qgr,db','fsi':'CSN','g_index':'G-index'}
    DB_UNITS  = {'adb':'%','vdb':'%','fcdb':'%','sdb':'%','qdb':'ккал/кг','fsi':'','g_index':''}
    def ad_to_db(ed):
        mad = ed.get('mad') or 0
        f = 100 / (100 - mad) if mad < 100 else 1
        result = {}
        try:
            if ed.get('aad') is not None: result['adb'] = round(ed['aad'] * f, 2)
            if ed.get('vad') is not None: result['vdb'] = round(ed['vad'] * f, 2)
            if ed.get('fc')  is not None: result['fcdb']= round(ed['fc']  * f, 2)
            if ed.get('sulfur') is not None: result['sdb'] = round(ed['sulfur'] * f, 2)
            if ed.get('cal_value') is not None: result['qdb'] = round(ed['cal_value'] / 4.1868 * f, 0)
            if ed.get('fsi') is not None: result['fsi'] = ed['fsi']
            if ed.get('g_val') is not None: result['g_index'] = ed['g_val']
        except: pass
        return result
    def get_orig(rid, row_num=None):
        if not rid: return {}
        if row_num:
            e = conn.execute("SELECT * FROM sample_entries WHERE receipt_id=? AND row_num=? AND is_duplicate=0", (rid, row_num)).fetchone()
        else:
            e = conn.execute("SELECT * FROM sample_entries WHERE receipt_id=? AND is_duplicate=0 AND row_status IN ('done','approved') ORDER BY row_num LIMIT 1", (rid,)).fetchone()
        if not e: return {}
        ed = dict(e)
        try:
            if ed.get('mad') is None and ed.get('dc_sample') and ed['dc_sample'] > 0:
                ed['mad'] = (ed['dc_tare'] + ed['dc_sample'] - ed['dc_dried']) / ed['dc_sample'] * 100
            if ed.get('aad') is None and ed.get('ash_sample') and ed['ash_sample'] > 0:
                ed['aad'] = (ed['ash_burned'] - ed['ash_tare']) / ed['ash_sample'] * 100
            if ed.get('vad') is None and ed.get('vol_sample') and ed['vol_sample'] > 0 and ed.get('mad') is not None:
                ed['vad'] = (ed['vol_tare'] + ed['vol_sample'] - ed['vol_burned']) / ed['vol_sample'] * 100 - ed['mad']
            if ed.get('fc') is None and all(ed.get(x) is not None for x in ['mad','aad','vad']):
                ed['fc'] = 100 - ed['mad'] - ed['aad'] - ed['vad']
        except: pass
        return ad_to_db(ed)
    orig1 = get_orig(qc['receipt_id_1'], qc['row_num_1'])
    orig2 = get_orig(qc['receipt_id_2'], qc['row_num_2'])
    results = conn.execute("SELECT * FROM internal_qc_results WHERE qc_id=?", (qc_id,)).fetchall()
    qc_tol = {r['parameter']: r['tolerance'] for r in conn.execute("SELECT parameter, tolerance FROM qc_settings").fetchall()}
    conn.close()
    db_to_tol = {'adb':'Aad','vdb':'Vad','fcdb':'Fc','sdb':'Stad','qdb':'Qb_ad','fsi':'FSI','g_index':'G_index'}
    return render_template('analysis/internal_qc_measure.html',
        qc=qc, orig1=orig1, orig2=orig2, results=results,
        qc_tol=qc_tol, params=PARAMS, db_labels=DB_LABELS, db_units=DB_UNITS,
        db_to_tol=db_to_tol,
        lang=session.get('lang','mn'), role=session.get('role'))

@app.route('/internal-qc/<int:qc_id>/approve')
@senior_required
def internal_qc_approve(qc_id):
    conn = get_db()
    conn.execute("UPDATE internal_qc SET status='done' WHERE id=? AND status='pending'", (qc_id,))
    conn.commit()
    conn.close()
    flash('QC баталгаажлаа!', 'success')
    return redirect(url_for('archive'))

@app.route('/internal-qc/<int:qc_id>/done', methods=['POST'])
@lab_required
def internal_qc_done(qc_id):
    conn = get_db()
    notes = request.form.get('notes','')
    conn.execute("DELETE FROM internal_qc_results WHERE qc_id=?", (qc_id,))
    params = ['adb','vdb','fcdb','sdb','qdb','fsi','g_index']
    db_to_tol = {'adb':'Aad','vdb':'Vad','fcdb':'Fc','sdb':'Stad','qdb':'Qb_ad','fsi':'FSI','g_index':'G_index'}
    for receipt_id in [request.form.get('rid1'), request.form.get('rid2')]:
        if not receipt_id: continue
        for p in params:
            rep = request.form.get(f'rep_{receipt_id}_{p}')
            orig = request.form.get(f'orig_{receipt_id}_{p}')
            if rep and orig:
                try:
                    o,r = float(orig), float(rep)
                    diff = abs(o-r)
                    tol_key = db_to_tol.get(p, p.capitalize())
                    tol = conn.execute("SELECT tolerance FROM qc_settings WHERE parameter=?", (tol_key,)).fetchone()
                    # round(): хоёртын бутархайн алдаанаас сэргийлнэ
                    # (0.53−0.50 = 0.030000000000000027 нь хүлцэл 0.03-ыг "давдаг")
                    within = 1 if (tol and round(diff, 6) <= float(tol['tolerance']) + 1e-9) else 0
                    conn.execute("""INSERT INTO internal_qc_results (qc_id,receipt_id,parameter,original_value,repeat_value,difference,within_tolerance)
                        VALUES (?,?,?,?,?,?,?)""", (qc_id, int(receipt_id), p, o, r, diff, within))
                except: pass
    conn.execute("UPDATE internal_qc SET status='done', notes=? WHERE id=?", (notes, qc_id))
    conn.commit()
    conn.close()
    flash('Дотоод QC бүртгэгдлээ', 'success')
    return redirect(url_for('internal_qc_list'))

# ── ARCHIVE ─────────────────────────────────────────────
@app.route('/archive')
@login_required
def archive():
    lang = session.get('lang','mn')
    conn = get_db()
    archived_devices = conn.execute("""
        SELECT d.*, dm.manufacturer, dm.model FROM devices d
        LEFT JOIN device_marks dm ON d.mark_id=dm.id
        WHERE d.status IN ('archived','replaced','decommissioned')
        ORDER BY d.name
    """).fetchall()
    archived_staff = conn.execute(
        # Архивт зөвхөн ажлаас ГАРСАН хүн (амарсан хүн ажилтны хуудсанд үлдэнэ)
        "SELECT * FROM users WHERE is_dismissed=1 AND is_deleted=0 AND role != 'admin' ORDER BY name"
    ).fetchall()
    completed_repairs = conn.execute("""
        SELECT r.*, d.name as dname FROM repairs r
        LEFT JOIN devices d ON d.id=r.device_id
        WHERE r.status IN ('done','replaced')
        ORDER BY r.reported_date DESC
    """).fetchall()
    completed_samples = conn.execute("""
        SELECT g.*, sr.lab_number, sr.lab_serial, sr.id as receipt_id,
               u.name as reg_name
        FROM geo_samples g
        JOIN sample_receipt sr ON sr.geo_sample_id=g.id
        LEFT JOIN users u ON u.id=g.registered_by
        WHERE g.status='done'
        ORDER BY sr.lab_serial DESC LIMIT 200
    """).fetchall()
    if session.get('role') == 'bayjuulach':
        u = conn.execute("SELECT can_view_result FROM users WHERE id=?", (session.get('user_id'),)).fetchone()
        if not u or not u['can_view_result']:
            completed_samples = []
        else:
            completed_samples = [s for s in completed_samples if 6000 <= (s['lab_serial'] or 0) <= 6999]
    elif session.get('role') == 'geologist':
        # Геологч зөвхөн зөвшөөрсөн мужийн дууссан дээжийг архиваас харна (view_ranges)
        u = conn.execute("SELECT view_ranges FROM users WHERE id=?", (session.get('user_id'),)).fetchone()
        vr = u['view_ranges'] if u else None
        if vr is not None:
            thousands = [int(x) for x in vr.split(',') if x.strip().isdigit()]
            completed_samples = [s for s in completed_samples if ((s['lab_serial'] or 0)//1000) in thousands]
    done_qc = conn.execute("""
        SELECT iq.*, sr1.lab_number as lab1, g1.sample_name as sname1,
               u.name as assigned_name
        FROM internal_qc iq
        LEFT JOIN sample_receipt sr1 ON sr1.id=iq.receipt_id_1
        LEFT JOIN geo_samples g1 ON g1.id=sr1.geo_sample_id
        LEFT JOIN users u ON u.id=iq.assigned_to
        WHERE iq.status='done'
        ORDER BY iq.triggered_date DESC LIMIT 200
    """).fetchall()
    # Харилцагч (геологч, баяжуулагч) нар зөвхөн шинжилгээ харна — бусдыг хоослоно
    if session.get('role') in ('geologist', 'bayjuulach'):
        archived_devices = []
        archived_staff = []
        completed_repairs = []
        done_qc = []
    conn.close()
    return render_template('admin/archive.html',
        archived_devices=archived_devices,
        archived_staff=archived_staff,
        completed_repairs=completed_repairs,
        done_samples=completed_samples,
        done_qc=done_qc,
        lang=lang)

@app.route('/archive/result/<int:receipt_id>')
@login_required
def archive_result(receipt_id):
    lang = session.get('lang','mn')
    role = session.get('role','')
    conn = get_db()
    _deny = result_access_denied(conn, receipt_id)
    if _deny:
        conn.close()
        flash(_deny[0], 'error')
        return redirect(_deny[1])
    if role in ('bayjuulach', 'geologist'):
        conn.execute('INSERT INTO result_view_log(user_id, receipt_id) VALUES(?,?)',
                     (session['user_id'], receipt_id))
        conn.commit()
    receipt = conn.execute("""
        SELECT sr.*, g.sample_name, g.sample_type, g.location,
               g.collected_date, g.quantity,
               ug.name as geo_name, up.name as prep_name
        FROM sample_receipt sr
        JOIN geo_samples g ON g.id=sr.geo_sample_id
        LEFT JOIN users ug ON ug.id=g.registered_by
        LEFT JOIN users up ON up.id=sr.received_by
        WHERE sr.id=?
    """, (receipt_id,)).fetchone()
    if not receipt:
        conn.close()
        flash('Бүртгэл олдсонгүй', 'error')
        return redirect(url_for('archive'))
    entries_raw = conn.execute("""
        SELECT se.*, u1.name as done_name, u2.name as approved_name
        FROM sample_entries se
        LEFT JOIN users u1 ON u1.id=se.done_by
        LEFT JOIN users u2 ON u2.id=se.approved_by
        WHERE se.receipt_id=?
        ORDER BY se.row_num, se.is_duplicate
    """, (receipt_id,)).fetchall()
    # mt_result (Нийт чийг) тооцоолол — template энэ талбарыг ашигладаг
    row_names, _name_of, _sname = sample_names_for(receipt)
    entries = []
    for e in entries_raw:
        ed = dict(e)
        ed['mt_result'] = total_moisture(ed)
        ed['g_val'] = lab_g_index(ed)   # гараар оруулаагүй бол жингээс бодно
        _en = ed.get('sample_name') or ''
        ed['display_name'] = _en if (_en and _en != _sname) else _name_of(ed.get('row_num', 1))
        entries.append(ed)
    apply_final_results(entries, qc_tolerances(conn))   # давталтаас эцсийн үр дүнг сонгоно
    qc = {r['parameter']: r for r in conn.execute("SELECT * FROM qc_settings").fetchall()}
    # Архивласан CRM дээж дээр ч сертификатын харьцуулалт гарна
    crm_cert = None
    if receipt and receipt['sample_type'] == 'CRM':
        geo = conn.execute("""SELECT crm_mad, crm_aad, crm_vad, crm_sulfur, crm_cal, crm_g,
                  crm_mad_unc, crm_aad_unc, crm_vad_unc, crm_sulfur_unc,
                  crm_cal_unc, crm_g_unc
                              FROM geo_samples WHERE id=?""",
                           (receipt['geo_sample_id'],)).fetchone()
        crm_cert = dict(geo) if geo else None
    conn.close()
    return render_template('analysis/archive_result.html',
        receipt=receipt, entries=entries, qc=qc, lang=lang, role=role, crm_cert=crm_cert,
        row_names=row_names)

@app.route('/archive/measure/<int:receipt_id>')
@login_required
def archive_measure(receipt_id):
    lang = session.get('lang','mn')
    conn = get_db()
    # Геологич/бэлтгэгчийн нэр — үр дүнгийн хуудастай ижил холболт.
    # Урьд нь эдгээр JOIN байхгүй тул архивын хэмжилтэд "—" гардаг байв.
    receipt = conn.execute("""
        SELECT sr.*, g.sample_name, g.sample_type, g.location, g.quantity,
               g.collected_date,
               g.notes as geo_notes,
               ug.name as geo_name,
               COALESCE(upb.name, up.name, sr.prep_operator) as prep_name
        FROM sample_receipt sr
        JOIN geo_samples g ON g.id=sr.geo_sample_id
        LEFT JOIN users ug  ON ug.id=g.registered_by
        LEFT JOIN users up  ON up.id=sr.received_by
        LEFT JOIN users upb ON upb.id=sr.prep_by
        WHERE sr.id=?
    """, (receipt_id,)).fetchone()
    if not receipt:
        conn.close()
        flash('Бүртгэл олдсонгүй', 'error')
        return redirect(url_for('archive'))
    entries = conn.execute("""
        SELECT * FROM sample_entries WHERE receipt_id=?
        ORDER BY row_num, is_duplicate
    """, (receipt_id,)).fetchall()
    calc = archive_calc(entries, qc_tolerances(conn))
    conn.close()
    row_names, _n, _sn = sample_names_for(receipt)
    return render_template('analysis/archive_measure.html',
        receipt=receipt, entries=entries, lang=lang, calc=calc,
        data_groups=measured_groups(entries), row_names=row_names)


# ── Архивын ХЭМЖИЛТИЙН хуудсыг Excel болгож татах ────────────────────────
# Албан тайлан (analysis_export) нь зөвхөн эцсийн үр дүнг гаргадаг. Түүхий
# жингүүд буюу хэмжилтийн хуудсыг бүтнээр нь татах арга байгаагүй тул
# лабораторийн анхдагч бүртгэлийг гадагш гаргах боломжгүй байв.
# Багана, гарчиг, дараалал нь дэлгэц дээрхтэй ЯГ ИЖИЛ.
MEASURE_EXPORT_GROUPS = [
    ('ff',  'Чөлөөт чийг',     [('ff_sample', 'Дээжний масс [гр]'),
                                ('ff_dried',  'Хатаасан масс [гр]')]),
    ('mt',  'Нийт чийг',       [('mt_bux',    'Бюкс №'),
                                ('mt_tare',   'Хоосон бюкс [гр]'),
                                ('mt_sample', 'Дээж масс [гр]'),
                                ('mt_dried',  'Хатаалтын дараах масс [гр]')]),
    ('dc',  'Дотоод чийг',     [('dc_bux',    'Бюкс №'),
                                ('dc_tare',   'Хоосон бюкс [гр]'),
                                ('dc_sample', 'Дээж масс [гр]'),
                                ('dc_dried',  'Хатаалтын дараах масс [гр]')]),
    ('un',  'Үнс',             [('ash_tav',    'Тэвш №'),
                                ('ash_tare',   'Хоосон бюкс [гр]'),
                                ('ash_sample', 'Дээж масс [гр]'),
                                ('ash_burned', 'Шатаалтын дараах масс [гр]')]),
    ('db',  'Дэгдэмхий бодис', [('vol_tig',    'Тигель №'),
                                ('vol_tare',   'Хоосон тигель [гр]'),
                                ('vol_sample', 'Дээж масс [гр]'),
                                ('vol_burned', 'Шатаалтын дараах масс [гр]')]),
    ('gi',  'G-Индекс',        [('g_tig',    'Тигель №'),
                                ('g_tare',   'Хоосон тигель [гр]'),
                                ('g_coke',   'Нийт кокс [гр]'),
                                ('g_sieve1', '1-дэх шигшүүрийн масс [гр]'),
                                ('g_sieve2', '2-дох шигшүүрийн масс [гр]')]),
    ('st',  'Нийт хүхэр',      [('sulfur', 'Sad [%]')]),
    ('il',  'Илчлэгийн утга',  [('cal_value', 'Qb,ad [J/g]'),
                                ('cal_temp',  'Өрөө ⁰C')]),
    ('fsi', 'Чөлөөт хөөлт',    [('fsi', 'CSN')]),
]


@app.route('/archive/export/measure/<int:receipt_id>')
@login_required
def archive_measure_export(receipt_id):
    """Хэмжилтийн хуудсыг (түүхий жин + тооцоо) Excel болгож татна."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_db()
    receipt = conn.execute("""
        SELECT sr.*, g.sample_name, g.sample_type, g.quantity, g.collected_date,
               ug.name AS geo_name,
               COALESCE(upb.name, up.name, sr.prep_operator) AS prep_name
          FROM sample_receipt sr
          JOIN geo_samples g ON g.id = sr.geo_sample_id
          LEFT JOIN users ug  ON ug.id = g.registered_by
          LEFT JOIN users up  ON up.id = sr.received_by
          LEFT JOIN users upb ON upb.id = sr.prep_by
         WHERE sr.id = ?""", (receipt_id,)).fetchone()
    if not receipt:
        conn.close()
        flash('Бүртгэл олдсонгүй', 'error')
        return redirect(url_for('archive'))
    entries = conn.execute("""SELECT * FROM sample_entries WHERE receipt_id=?
                              ORDER BY row_num, is_duplicate""", (receipt_id,)).fetchall()
    calc = archive_calc(entries, qc_tolerances(conn))
    conn.close()
    row_names, _n, _sn = sample_names_for(receipt)
    shown = measured_groups(entries)          # дэлгэцтэй ижил — хэмжсэн бүлэг л
    groups = [g for g in MEASURE_EXPORT_GROUPS if g[0] in shown] or MEASURE_EXPORT_GROUPS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Хэмжилт'
    THIN = Side(style='thin', color='B0B0B0')
    BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HFILL = PatternFill('solid', fgColor='3C3489')
    SFILL = PatternFill('solid', fgColor='4F46A3')
    CFILL = PatternFill('solid', fgColor='E8FAF4')
    MFILL = PatternFill('solid', fgColor='CFEEE2')
    DFILL = PatternFill('solid', fgColor='EEF5FF')
    CEN = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Толгойн мэдээлэл ──
    info = [('Ажлын дугаар', receipt['lab_number']),
            ('Дээжний нэр',  receipt['sample_name']),
            ('Төрөл',        receipt['sample_type']),
            ('Тоо',          receipt['quantity']),
            ('Огноо',        receipt['collected_date'] or receipt['received_date']),
            ('Геологи',      receipt['geo_name'] or '—'),
            ('Дээж бэлтгэгч', receipt['prep_name'] or '—')]
    for i, (k, v) in enumerate(info):
        ws.cell(1, 1 + i * 2, k).font = Font(size=9, color='888888')
        c = ws.cell(2, 1 + i * 2, v)
        c.font = Font(size=11, bold=True)

    # ── Хүснэгтийн гарчиг (2 мөр) ──
    FIXED = ['No.', 'Төрөл', 'Лаб.дугаар', 'Дээжний дугаар', 'Масс [кг]']
    HR1, HR2 = 4, 5
    col = 1
    for lbl in FIXED:
        ws.merge_cells(start_row=HR1, start_column=col, end_row=HR2, end_column=col)
        c = ws.cell(HR1, col, lbl)
        c.fill, c.font, c.alignment, c.border = HFILL, Font(bold=True, color='FFFFFF', size=10), CEN, BORD
        ws.cell(HR2, col).border = BORD
        col += 1
    for _key, title, fields in groups:
        ws.merge_cells(start_row=HR1, start_column=col, end_row=HR1, end_column=col + len(fields) - 1)
        c = ws.cell(HR1, col, title)
        c.fill, c.font, c.alignment, c.border = HFILL, Font(bold=True, color='FFFFFF', size=10), CEN, BORD
        for _f, flbl in fields:
            c2 = ws.cell(HR2, col, flbl)
            c2.fill, c2.font, c2.alignment, c2.border = SFILL, Font(color='FFFFFF', size=9), CEN, BORD
            ws.column_dimensions[get_column_letter(col)].width = 13
            col += 1
    calc_start = col
    ws.merge_cells(start_row=HR1, start_column=col, end_row=HR1,
                   end_column=col + len(ARCHIVE_CALC_COLS) * 2 - 1)
    c = ws.cell(HR1, col, 'Тооцоо')
    c.fill, c.font, c.alignment, c.border = HFILL, Font(bold=True, color='FFFFFF', size=10), CEN, BORD
    for lbl, _k, _p, _dp in ARCHIVE_CALC_COLS:
        for sub in (lbl, 'Дундаж'):
            c2 = ws.cell(HR2, col, sub)
            c2.fill, c2.font, c2.alignment, c2.border = SFILL, Font(color='FFFFFF', size=9), CEN, BORD
            ws.column_dimensions[get_column_letter(col)].width = 10
            col += 1
    for i, w in enumerate([5, 11, 15, 16, 9], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HR1].height = 20
    ws.row_dimensions[HR2].height = 34

    # ── Мөрүүд: үндсэн → зэрэгцээ → давталт ──
    by_row = {}
    for e in entries:
        by_row.setdefault(e['row_num'], {})[e['is_duplicate']] = e
    r = HR2 + 1
    for ri in range(1, (receipt['quantity'] or 1) + 1):
        for dup in sorted(by_row.get(ri, {}).keys()) or [0]:
            e = by_row.get(ri, {}).get(dup)
            is_main = (dup == 0)
            name = ((e['sample_name'] if e and e['sample_name'] else None)
                    or (row_names[ri - 1] if len(row_names) >= ri else '—')) if is_main else (
                    'зэрэгцээ' if dup == 1 else f'давталт {dup}')
            fixed = [ri if is_main else '', receipt['sample_type'] if is_main else '',
                     receipt['lab_number'], name,
                     (e['mass_kg'] if e and is_main else None)]
            for i, v in enumerate(fixed, start=1):
                cc = ws.cell(r, i, v)
                cc.border, cc.alignment = BORD, CEN
                if not is_main:
                    cc.fill = DFILL
            cix = len(FIXED) + 1
            for _key, _title, fields in groups:
                for f, _lbl in fields:
                    v = None
                    try:
                        v = e[f] if e else None
                    except (KeyError, IndexError):
                        v = None
                    cc = ws.cell(r, cix, v)
                    cc.border, cc.alignment = BORD, CEN
                    if isinstance(v, float):
                        cc.number_format = '0.0000'
                    if not is_main:
                        cc.fill = DFILL
                    cix += 1
            cv = calc.get((ri, dup)) or {}
            mean = cv.get('mean') or {}
            diff = cv.get('diff') or {}
            for _lbl, k, _p, dp in ARCHIVE_CALC_COLS:
                c1 = ws.cell(r, cix, cv.get(k))
                c1.border, c1.alignment, c1.fill = BORD, CEN, CFILL
                c1.number_format = '0.' + '0' * dp if dp else '0'
                c1.font = Font(bold=True, color='0A6E3F', size=10)
                cix += 1
                mv = mean.get(k)
                dv = diff.get(k)
                c2 = ws.cell(r, cix, mv if mv is not None else
                             (f'Δ {dv:.{dp}f}' if dv is not None else None))
                c2.border, c2.alignment, c2.fill = BORD, CEN, MFILL
                if mv is not None:
                    c2.number_format = '0.' + '0' * dp if dp else '0'
                c2.font = Font(color='0A4F3C', size=10)
                cix += 1
            r += 1

    ws.freeze_panes = ws.cell(HR2 + 1, len(FIXED) + 1)
    ws.print_area = f'A1:{get_column_letter(cix - 1)}{r - 1}'
    ws.page_setup.orientation = 'landscape'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fn = f"measure_{(receipt['lab_number'] or receipt_id)}.xlsx".replace(' ', '_')
    return send_file(out, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/analysis/find-receipt')
@admin_required
def analysis_find_receipt():
    lab_number = request.args.get('lab_number','').strip()
    conn = get_db()
    row = conn.execute("""
        SELECT sr.id, sr.lab_number, sr.received_date, g.quantity, g.sample_name
        FROM sample_receipt sr
        JOIN geo_samples g ON g.id=sr.geo_sample_id
        WHERE sr.lab_number=?
    """, (lab_number,)).fetchone()
    conn.close()
    if row:
        return jsonify({
            'id': row['id'],
            'lab_number': row['lab_number'],
            'received_date': row['received_date'] or '',
            'quantity': row['quantity'] or 0,
            'sample_name': row['sample_name'] or ''
        })
    return jsonify({})

@app.route('/analysis/delete/<int:receipt_id>', methods=['POST'])
@admin_required
def analysis_delete(receipt_id):
    conn = get_db()
    receipt = conn.execute('SELECT geo_sample_id FROM sample_receipt WHERE id=?', (receipt_id,)).fetchone()
    if receipt:
        conn.execute('DELETE FROM sample_entries WHERE receipt_id=?', (receipt_id,))
        conn.execute('DELETE FROM sample_receipt WHERE id=?', (receipt_id,))
        conn.execute("UPDATE geo_samples SET status='pending' WHERE id=?", (receipt['geo_sample_id'],))
        conn.commit()
        flash('Шинжилгээний бүртгэл устгагдлаа', 'success')
    conn.close()
    return redirect(url_for('analysis'))

@app.route('/archive/reopen/<int:receipt_id>', methods=['POST'])
@perm_required('can_reopen')
def archive_reopen(receipt_id):
    conn = get_db()
    receipt = conn.execute('SELECT geo_sample_id FROM sample_receipt WHERE id=?', (receipt_id,)).fetchone()
    if receipt:
        conn.execute("UPDATE geo_samples SET status='analysing' WHERE id=?", (receipt['geo_sample_id'],))
        conn.execute("UPDATE sample_receipt SET prep_status='ready' WHERE id=?", (receipt_id,))
        conn.execute("UPDATE sample_entries SET row_status='done', approved_by=NULL, approved_at=NULL WHERE receipt_id=? AND row_status='approved'", (receipt_id,))
        conn.commit()
        flash('Шинжилгээ дахин нээгдлээ', 'success')
    conn.close()
    return redirect(url_for('analysis_result', receipt_id=receipt_id))

@app.route('/archive/delete/<int:receipt_id>', methods=['POST'])
@admin_required
def archive_delete(receipt_id):
    conn = get_db()
    try:
        receipt = conn.execute('SELECT geo_sample_id FROM sample_receipt WHERE id=?', (receipt_id,)).fetchone()
        if receipt:
            geo_id = receipt['geo_sample_id']
            # QC байгаа эсэх шалгах — байвал устгахгүй анхааруулна
            qc_linked = conn.execute(
                "SELECT COUNT(*) FROM internal_qc WHERE receipt_id_1=? OR receipt_id_2=?",
                (receipt_id, receipt_id)
            ).fetchone()[0]
            if qc_linked:
                conn.close()
                flash(f'Энэ дээжтэй {qc_linked} QC бүртгэл холбоотой байна. Эхлээд QC-г устгана уу.', 'error')
                return redirect(url_for('archive'))
            # Хамааралтай бүх бичлэгийг эхлээд устгана (child → parent дараалал)
            conn.execute("DELETE FROM result_view_log WHERE receipt_id=?", (receipt_id,))
            conn.execute("DELETE FROM device_usage_log WHERE receipt_id=?", (receipt_id,))
            conn.execute("DELETE FROM sample_entries WHERE receipt_id=?", (receipt_id,))
            conn.execute("DELETE FROM sample_receipt WHERE id=?", (receipt_id,))
            conn.execute("DELETE FROM geo_samples WHERE id=?", (geo_id,))
        conn.commit()
        flash('Бүртгэл бүрмөсөн устгагдлаа.', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Устгахад алдаа гарлаа: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('archive'))

@app.route('/archive/qc-delete/<int:qc_id>', methods=['POST'])
@admin_required
def archive_qc_delete(qc_id):
    conn = get_db()
    conn.execute("DELETE FROM internal_qc_results WHERE qc_id=?", (qc_id,))
    conn.execute("DELETE FROM internal_qc WHERE id=?", (qc_id,))
    conn.commit()
    conn.close()
    flash('QC бүртгэл бүрмөсөн устгагдлаа.', 'success')
    return redirect(url_for('archive'))

@app.route('/archive/export/<int:receipt_id>')
@login_required
def analysis_export_report(receipt_id):
    return redirect(url_for('analysis_export', receipt_id=receipt_id))

# ── STAFF EDIT (admin/senior) ───────────────────────────
@app.route('/staff/<int:uid>/edit', methods=['GET','POST'])
@senior_required
def staff_edit(uid):
    lang = session.get('lang','mn')
    conn = get_db()
    target = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    devices = conn.execute("SELECT * FROM devices ORDER BY name").fetchall()
    perms = [r['device_id'] for r in conn.execute("SELECT device_id FROM staff_device_permissions WHERE user_id=?", (uid,)).fetchall()]
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'update_info':
            photo = save_file(request.files.get('photo'), 'staff')
            # Senior зөвхөн staff эрх өгч болно
            role_to_set = request.form.get('role', target['role'])
            if session.get('role') == 'senior' and role_to_set in ('admin','senior'):
                role_to_set = target['role']
            is_client = role_to_set in ('bayjuulach', 'geologist')
            is_lab    = role_to_set in ('staff', 'preparer')
            can_reg    = 1 if (is_client and request.form.get('can_register'))    else 0
            can_view   = 1 if (is_client and request.form.get('can_view_result')) else 0
            can_export = 1 if (is_client and request.form.get('can_export'))      else 0
            can_approve = 1 if (is_lab and request.form.get('can_approve')) else 0
            can_report  = 1 if (is_lab and request.form.get('can_report'))  else 0
            can_reopen  = 1 if (is_lab and request.form.get('can_reopen'))  else 0
            is_active  = 1 if request.form.get('is_active') else 0
            # Геологчийн харах мужууд (checkbox-оор). Зөвхөн геологчид хадгална.
            if role_to_set == 'geologist':
                _ranges = request.form.getlist('view_ranges')  # ['1','2',...]
                _ranges = [r for r in _ranges if r.isdigit()]
                view_ranges = ','.join(_ranges) if _ranges else ''
            else:
                view_ranges = None
            conn.execute("""
                UPDATE users SET name=?,position=?,phone=?,email=?,role=?,joined_date=?,shift=?,
                                 can_register=?,can_view_result=?,can_export=?,is_active=?,
                                 can_approve=?,can_report=?,can_reopen=?,view_ranges=?
                WHERE id=?
            """, (
                request.form.get('name', target['name']),
                request.form.get('position'),
                request.form.get('phone'),
                request.form.get('email'),
                role_to_set,
                request.form.get('joined_date') or None,
                request.form.get('shift') or None,
                can_reg, can_view, can_export, is_active,
                can_approve, can_report, can_reopen, view_ranges,
                uid
            ))
            if photo:
                conn.execute("UPDATE users SET photo=? WHERE id=?", (photo, uid))
            # Эрх шинэчлэх — геологи/баяжуулах тоног ашиглах эрх олгохгүй
            conn.execute("DELETE FROM staff_device_permissions WHERE user_id=?", (uid,))
            if role_to_set not in ('geologist', 'bayjuulach'):
                for did in request.form.getlist('device_permissions'):
                    conn.execute("INSERT OR IGNORE INTO staff_device_permissions(user_id,device_id) VALUES(?,?)", (uid, did))
            conn.commit()
            flash('Мэдээлэл шинэчлэгдлээ!' if lang=='mn' else 'Updated!', 'success')
            # Хадгалаад жагсаалт руу шилжих
            if request.form.get('next') == 'list':
                conn.close()
                return redirect(url_for('staff_list'))
        elif action == 'reset_password':
            new_pw = request.form.get('new_password','')
            if len(new_pw) < 6:
                flash('Нууц үг хамгийн багадаа 6 тэмдэгт!' if lang=='mn' else 'Min 6 characters!', 'error')
            else:
                conn.execute("UPDATE users SET password_hash=? WHERE id=?", (hash_password(new_pw), uid))
                conn.commit()
                _check_weak_admin()
                flash('Нууц үг шинэчлэгдлээ!' if lang=='mn' else 'Password reset!', 'success')
        conn.close()
        return redirect(url_for('staff_edit', uid=uid))
    conn.close()
    return render_template('admin/staff_edit.html', target=target, devices=devices, perms=perms, lang=lang)

# ── PROFILE EDIT ─────────────────────────────────────────
@app.route('/profile', methods=['GET','POST'])
@login_required
def profile():
    lang = session.get('lang','mn')
    uid  = session.get('user_id', 0)
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'update_info':
            photo = save_file(request.files.get('photo'), 'staff')
            conn.execute("""
                UPDATE users SET position=?, phone=?, email=?
                WHERE id=?
            """, (
                request.form.get('position'),
                request.form.get('phone'),
                request.form.get('email'),
                uid
            ))
            if photo:
                conn.execute("UPDATE users SET photo=? WHERE id=?", (photo, uid))
            conn.commit()
            flash('Мэдээлэл шинэчлэгдлээ!' if lang=='mn' else 'Profile updated!', 'success')
        elif action == 'change_password':
            old_pw = request.form.get('old_password','')
            new_pw = request.form.get('new_password','')
            confirm_pw = request.form.get('confirm_password','')
            if not check_password(user['password_hash'], old_pw):
                flash('Хуучин нууц үг буруу!' if lang=='mn' else 'Wrong current password!', 'error')
            elif new_pw != confirm_pw:
                flash('Шинэ нууц үг таарахгүй байна!' if lang=='mn' else 'Passwords do not match!', 'error')
            elif len(new_pw) < 6:
                flash('Нууц үг хамгийн багадаа 6 тэмдэгт байх ёстой!' if lang=='mn' else 'Password must be at least 6 characters!', 'error')
            else:
                conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                           (hash_password(new_pw), uid))
                conn.commit()
                _check_weak_admin()
                flash('Нууц үг амжилттай солигдлоо!' if lang=='mn' else 'Password changed!', 'success')
        conn.close()
        return redirect(url_for('profile'))
    conn.close()
    return render_template('staff/profile.html', user=user, lang=lang)

# ── STAFF DEACTIVATE ────────────────────────────────────
@app.route('/staff/<int:uid>/deactivate', methods=['POST'])
@senior_required
def staff_deactivate(uid):
    lang = session.get('lang','mn')
    conn = get_db()
    # Өөрийгөө идэвхгүй болгохоос хамгаалах
    if uid == session.get('user_id', 0):
        conn.close()
        flash('Өөрийгөө идэвхгүй болгох боломжгүй!' if lang=='mn' else 'Cannot deactivate yourself!', 'error')
        return redirect(url_for('staff_list'))
    # "Гаргах" = ажлаас гарсан → архивт шилжинэ, ажилтны хуудсанд харагдахгүй
    conn.execute("UPDATE users SET is_active=0, is_dismissed=1 WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    flash('Ажилтан ажлаас гарсан — архивт шилжлээ.' if lang=='mn'
          else 'Staff dismissed — moved to archive.', 'success')
    return redirect(url_for('staff_list'))

@app.route('/staff/shift/<shift>/<action>', methods=['POST'])
@senior_required
def staff_shift_bulk(shift, action):
    """Ээлжээр бөөнөөр идэвхгүй (амраах) / идэвхтэй (ажиллуулах) болгоно.

    Бүлэг тус бүрээр ажиллана: лабораторийн ээлжийг амраахад геологич,
    баяжуулагч нар хөндөгдөхгүй. Урьд нь ээлж бүхэлдээ хамрагддаг байсан
    тул нэг табаас "Амраах" дархад бусад бүлгийн тэр ээлжийнхэн ч
    амардаг байв.
    """
    if shift not in ('A', 'B') or action not in ('rest', 'work'):
        flash('Буруу хүсэлт.', 'error')
        return redirect(url_for('staff_list'))

    group = (request.form.get('group') or 'all').lower()
    GROUPS = {
        'lab': ("role NOT IN ('geologist','bayjuulach')", 'лабораторийн'),
        'geo': ("role = 'geologist'", 'геологийн'),
        'dp':  ("role = 'bayjuulach'", 'баяжуулахын'),
        'all': ('1=1', ''),
    }
    if group not in GROUPS:
        group = 'all'
    where_role, label = GROUPS[group]

    new_active = 0 if action == 'rest' else 1
    conn = get_db()
    # Өөрийгөө, админ болон ажлаас гарсан хүнийг хөндөхгүй.
    # Амраах нь is_dismissed-д хүрэхгүй тул архивт ОРОХГҮЙ — зөвхөн
    # ажилтны хуудсанд идэвхгүй болж харагдана.
    me = session.get('user_id', 0)
    cur = conn.execute(
        f"""UPDATE users SET is_active=? WHERE shift=? AND role != 'admin'
            AND id != ? AND is_dismissed=0 AND {where_role}""",
        (new_active, shift, me))
    n = cur.rowcount
    conn.commit()
    conn.close()
    who = f'{shift} ээлжийн {label}'.strip()
    if action == 'rest':
        flash(f'{who} {n} хүн амрав (идэвхгүй болголоо).', 'success')
    else:
        flash(f'{who} {n} хүн ажилд орлоо (идэвхжүүлэв).', 'success')
    return redirect(url_for('staff_list'))

@app.route('/staff/<int:uid>/activate', methods=['POST'])
@senior_required
def staff_activate(uid):
    lang = session.get('lang','mn')
    conn = get_db()
    # "Буцаах" = ажилдаа эргэж орлоо (амарсан ч бай, ажлаас гарсан ч бай)
    conn.execute("UPDATE users SET is_active=1, is_dismissed=0 WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    flash('Ажилтан идэвхжүүлэгдлээ.' if lang=='mn' else 'Staff activated.', 'success')
    return redirect(url_for('staff_list'))

# ── ЭЦСИЙН ҮР ДҮН СОНГОХ (давталттай үед) ───────────────
# Нэг дээжийн үндсэн (is_duplicate=0), зэрэгцээ (1) болон давталт (2,3…)
# хэмжилтээс албан ёсны үр дүнг сонгоно: ХАМГИЙН ОЙРХОН ХОЁРЫН ДУНДАЖ.
# Үзүүлэлт бүр дээр тусад нь тооцогддог тул зөвхөн зөрсөн шинжилгээг
# давтахад хангалттай — бусад үзүүлэлт хөндөгдөхгүй.
FINAL_RESULT_FIELDS = ('mt_result', 'mad', 'aad', 'vad', 'sulfur',
                       'cal_value', 'g_val', 'fsi')


def total_moisture(d):
    """Нийт чийг Mt = чөлөөт чийг + үлдэгдэл чийг (нийлмэл).

    Хэмжилтийн хуудас болон Excel тайлантай ижил томьёо. Чөлөөт чийг
    бичигдээгүй бол зөвхөн үлдэгдэл чийг буцаана.
    """
    def num(k):
        v = d.get(k)
        try:
            return float(v) if v is not None else None
        except (TypeError, ValueError):
            return None
    fs, fd = num('ff_sample'), num('ff_dried')
    chch = ((fs - fd) / fs * 100) if (fs and fs > 0 and fd is not None) else 0
    mtt, mts, mtd = num('mt_tare'), num('mt_sample'), num('mt_dried')
    if mtt is not None and mts and mts > 0 and mtd is not None:
        tm_raw = (mtt + mts - mtd) / mts * 100
        return (chch + tm_raw * (1 - chch / 100)) if chch else tm_raw
    return None


# Үзүүлэлт бүрийн хүлцэл qc_settings-д ямар нэрээр хадгалагддаг вэ
FINAL_TOL_PARAM = {'mad': 'Mad', 'aad': 'Aad', 'vad': 'Vad', 'sulfur': 'Stad',
                   'cal_value': 'Qb_ad', 'g_val': 'G_index', 'fsi': 'FSI'}


# Хэмжилтийн хуудасны загварт бичигдсэнтэй ИЖИЛ анхдагч хүлцэл. qc_settings-ээс
# устгагдсан үзүүлэлт дээр хөтөч эдгээрийг ашигладаг тул сервер мөн адил байх ёстой
# (эс бөгөөс Тооцоо ба Үр дүнгийн хуудас өөр дундаж гаргана).
QC_DEFAULTS = {'Mad': 0.20, 'Aad': 0.30, 'Vad': 0.30, 'Stad': 0.03,
               'Qb_ad': 120, 'G_index': 3.0, 'FSI': 0.5}


def _mat(row, col):
    """crm_materials-ийн багана байхгүй хуучин DB дээр ч аюулгүй унших"""
    try:
        return row[col] if col in row.keys() else None
    except Exception:
        return None


def qc_tolerances(conn):
    """qc_settings → {parameter: tolerance} (тохируулаагүйд анхдагч утга)"""
    tols = dict(QC_DEFAULTS)
    for r in conn.execute('SELECT parameter, tolerance FROM qc_settings'):
        if r['tolerance'] is not None:
            tols[r['parameter']] = r['tolerance']
    return tols


def best_set_mean(values, tol=None):
    """Хүлцэлд багтах ХАМГИЙН ОЛОН хэмжилтийн дундаж.

    Гурав (ба түүнээс дээш) хэмжилт бүгд хүлцэлд багтвал бүгдийн дундаж,
    эс бөгөөс хамгийн ойрхон хоёрын дундаж. Хүлцэл өгөөгүй үзүүлэлт
    (Mt, FC) дээр хамгийн ойрхон хоёроор бодно.

    Буцаах утга: (дундаж, сонгогдсоны зөрүү, сонгогдсон утгын тоо)
    """
    vals = sorted(v for v in values if v is not None)
    if not vals:
        return None, None, 0
    if len(vals) == 1:
        return vals[0], 0.0, 1
    best = None
    if tol is not None:
        t = float(tol) + 1e-9
        for i in range(len(vals)):
            # i-ээс эхлэх хамгийн УРТ бүлгийг олно (эрэмбэлсэн тул j-ээс буцаж хайна)
            for j in range(len(vals) - 1, i, -1):
                if vals[j] - vals[i] <= t:
                    cand = (j - i + 1, vals[j] - vals[i], i, j)
                    if best is None or cand[0] > best[0] or \
                       (cand[0] == best[0] and cand[1] < best[1]):
                        best = cand
                    break
    if best is None:
        # Хүлцэлд багтах хос алга — хамгийн ойрхон хоёроор (үр дүн нь QC-д унана)
        i = min(range(len(vals) - 1), key=lambda k: vals[k + 1] - vals[k])
        best = (2, vals[i + 1] - vals[i], i, i + 1)
    sel = vals[best[2]:best[3] + 1]
    return sum(sel) / len(sel), best[1], len(sel)


def closest_pair_mean(values):
    """Хамгийн ойрхон хоёр утгын дундаж (нэг л утга байвал өөрийг нь)"""
    return best_set_mean(values)[0]


# Архивын хэмжилтийн "Тооцоо" хэсэг — хэмжилтийн хуудастай ИЖИЛ 9 үзүүлэлт.
# (нэр, тооцооны түлхүүр, хүлцлийн параметр, бутархайн орон)
ARCHIVE_CALC_COLS = [
    ('TM%',    'tm',  None,       1),
    ('Mad%',   'mad', 'Mad',      2),
    ('Aad%',   'aad', 'Aad',      2),
    ('Vad%',   'vad', 'Vad',      2),
    ('FCad%',  'fc',  None,       2),
    ('G',      'g',   'G_index',  0),
    ('Sad%',   'sad', 'Stad',     2),
    ('Qb,ad',  'qb',  'Qb_ad',    1),
    ('CSN',    'fsi', 'FSI',      1),
]


def archive_calc(entries, tols):
    """Архивын хуудсанд харуулах тооцоо: мөр бүрийн утга + дундаж/зөрүү.

    Архив нь зөвхөн харах хуудас тул JavaScript-ээр бодогддоггүй. Урьд нь
    зөвхөн DB-д хадгалагдсан Mad/Aad/Vad/FC дөрөв гардаг байсан тул нийт
    чийг, G, хүхэр, илчлэг, ЧХЗ болон дундаж нь огт харагдахгүй, хэмжилтийн
    хуудаснаас өөр байдалтай байв.

    Буцаах: {(row_num, is_duplicate): {'tm': ..., 'mad': ..., 'mean': {...},
                                       'diff': {...}}}
    Дундаж нь ҮНДСЭН мөрөнд, зөрүү (Δ) нь ЗЭРЭГЦЭЭ мөрөнд гарна —
    хэмжилтийн хуудастай ижил.
    """
    def num(e, k):
        try:
            v = e[k]
            return float(v) if v is not None and v != '' else None
        except (KeyError, IndexError, TypeError, ValueError):
            return None

    vals, rows = {}, {}
    for e in entries:
        key = (e['row_num'], e['is_duplicate'])
        d = dict(e)
        v = {
            'tm':  total_moisture(d),
            'mad': num(e, 'mad'),
            'aad': num(e, 'aad'),
            'vad': num(e, 'vad'),
            'fc':  num(e, 'fc'),
            'g':   num(e, 'g_val'),
            'sad': num(e, 'sulfur'),
            'qb':  num(e, 'cal_value'),
            'fsi': num(e, 'fsi'),
        }
        if v['g'] is None:
            gt, gc = num(e, 'g_tare'), num(e, 'g_coke')
            g1, g2 = num(e, 'g_sieve1'), num(e, 'g_sieve2')
            if None not in (gt, gc, g1, g2) and (gc - gt) > 0:
                v['g'] = 10 + (30 * (g1 - gt) + 70 * (g2 - gt)) / (gc - gt)
        if v['fc'] is None and None not in (v['mad'], v['aad'], v['vad']):
            v['fc'] = 100 - v['mad'] - v['aad'] - v['vad']
        vals[key] = v
        rows.setdefault(e['row_num'], []).append(e['is_duplicate'])

    out = {k: dict(v) for k, v in vals.items()}
    for rn, dups in rows.items():
        if len(dups) < 2:
            continue
        mean, diff = {}, {}
        for _, key, param, _dp in ARCHIVE_CALC_COLS:
            series = [vals[(rn, d)].get(key) for d in dups]
            if len([x for x in series if x is not None]) < 2:
                continue
            m, df, _n = best_set_mean(series, tols.get(param) if param else None)
            mean[key], diff[key] = m, df
        out.setdefault((rn, 0), {})['mean'] = mean
        if 1 in dups:
            out.setdefault((rn, 1), {})['diff'] = diff
    return out


def apply_final_results(entries, tol_map=None):
    """Мөр бүрийн эцсийн үр дүнг үндсэн мөрөнд (is_duplicate=0) бичнэ.

    entries — dict-үүдийн жагсаалт. Үндсэн мөрийн утгууд албан ёсны үр дүн
    болж тайлан болон үр дүнгийн хуудсанд гардаг тул энд бичигдэнэ.
    Зэрэгцээ ба давталтын мөрүүд хэвээр үлдэж QC зөрүү харуулахад орно.

    tol_map — qc_settings-ийн хүлцэл. Өгвөл хүлцэлд багтах бүх хэмжилтийг
    дундажилна (хэмжилтийн хуудасны Тооцоо хэсэгтэй ижил дүрэм).
    """
    by_row = {}
    for e in entries:
        by_row.setdefault(e.get('row_num'), []).append(e)
    for rows in by_row.values():
        if len(rows) < 2:
            continue                       # давталтгүй — сонгох зүйлгүй
        primary = next((r for r in rows if r.get('is_duplicate') == 0), None)
        if not primary:
            continue
        for f in FINAL_RESULT_FIELDS:
            tol = (tol_map or {}).get(FINAL_TOL_PARAM.get(f))
            v, _, _ = best_set_mean([r.get(f) for r in rows], tol)
            if v is not None:
                primary[f] = v
        # FC-г эцсийн Mad/Aad/Vad-аас гаргана (хоорондоо нийцтэй байхын тулд)
        if all(primary.get(x) is not None for x in ('mad', 'aad', 'vad')):
            primary['fc'] = 100 - primary['mad'] - primary['aad'] - primary['vad']
    return entries


def _staff_ref_columns(conn):
    """users(id) руу заасан бүх хүснэгт.багана — схемээс шууд уншина.

    Гараар жагсаалт хөтөлбөл шинэ хүснэгт нэмэгдэхэд мартагдаж, ажилтныг
    'түүхгүй' гэж андуурч устгах гэж оролдоод FOREIGN KEY алдаа өгдөг.
    """
    refs = []
    for (table,) in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"):
        if table in ('users', 'staff_device_permissions'):
            continue  # өөрөө / эрхийн бичлэг тусдаа боловсруулагдана
        try:
            for fk in conn.execute(f'PRAGMA foreign_key_list("{table}")'):
                if fk[2] == 'users' and fk[3]:
                    refs.append((table, fk[3]))
        except Exception:
            pass
    return refs


def _staff_history_count(conn, uid):
    """Ажилтны нэртэй холбоотой бичлэгийн тоо (схемээс автоматаар)"""
    total = 0
    for table, col in _staff_ref_columns(conn):
        try:
            total += conn.execute(
                f'SELECT COUNT(*) FROM "{table}" WHERE "{col}"=?', (uid,)).fetchone()[0]
        except Exception:
            pass
    return total


@app.route('/staff/<int:uid>/delete', methods=['POST'])
@admin_required
def staff_delete(uid):
    lang = session.get('lang','mn')
    if uid == session.get('user_id'):
        flash('Өөрийгөө устгах боломжгүй!', 'error')
        return redirect(url_for('archive'))
    conn = get_db()
    # Устгахын өмнө заавал "Гаргах" хийж архивт орсон байх ёстой
    target = conn.execute(
        "SELECT * FROM users WHERE id=? AND is_dismissed=1 AND is_deleted=0", (uid,)).fetchone()
    if not target:
        conn.close()
        flash('Ажилтан архивт олдсонгүй. Эхлээд "Гаргах" товчийг дарна уу.', 'error')
        return redirect(url_for('archive'))

    name = target['name']
    linked = _staff_history_count(conn, uid)
    # Эрхийн бичлэг бол ажлын түүх биш — үргэлж устгана
    try: conn.execute("DELETE FROM staff_device_permissions WHERE user_id=?", (uid,))
    except Exception: pass

    hard_deleted = False
    if linked == 0:
        # Ажлын түүхгүй → бүрмөсөн устгахыг оролдоно.
        # Хэрэв тооллогод ороогүй холбоос үлдсэн бол алдаа өгөхийн оронд
        # доорх нэр хадгалах горим руу шилжинэ (500 хуудас хэзээ ч гарахгүй).
        try:
            # SQLite-д алдаатай мөр зөвхөн өөрөө буцдаг тул гүйлгээ бүтэн үлдэнэ
            conn.execute("DELETE FROM users WHERE id=?", (uid,))
            hard_deleted = True
            msg = (f'{name} бүрмөсөн устгагдлаа.' if lang=='mn'
                   else f'{name} permanently deleted.')
        except Exception:
            linked = -1   # тооллогод ороогүй холбоос үлдсэн → нэр хадгална
    if not hard_deleted:
        # Ажлын түүхтэй → системээс алга болно, гэхдээ хуучин шинжилгээ,
        # тайлангийн гарын үсэгт нэр нь хадгалагдана
        conn.execute("""
            UPDATE users SET is_deleted=1, is_active=0, deleted_at=?,
                   password_hash='', employee_id=?, phone=NULL, email=NULL,
                   photo=NULL, shift=NULL
            WHERE id=?
        """, (datetime.now().isoformat(), f'DELETED-{uid}', uid))
        cnt = f'{linked} ' if linked > 0 else ''
        msg = (f'{name} устгагдлаа. Холбоотой {cnt}бичлэгт нэр нь '
               f'мөшгих зорилгоор хадгалагдав.' if lang=='mn'
               else f'{name} deleted; name retained in linked records.')
    conn.commit()
    conn.close()
    flash(msg, 'success')
    return redirect(url_for('archive'))

# ── DEVICE ARCHIVE / RESTORE ────────────────────────────
@app.route('/devices/<int:did>/archive', methods=['POST'])
@senior_required
def device_archive(did):
    lang = session.get('lang','mn')
    conn = get_db()
    reason = request.form.get('reason', 'archived')
    conn.execute("UPDATE devices SET status=? WHERE id=?", (reason, did))
    conn.commit(); conn.close()
    msgs = {
        'standby':        'Төхөөрөмж нөөцөд шилжлээ.',
        'repair':         'Төхөөрөмж засварт шилжлээ. Дэлгэрэнгүйг бүртгэнэ үү.',
        'replaced':       'Төхөөрөмж солигдсон гэж тэмдэглэгдлээ.',
        'decommissioned': 'Төхөөрөмж акталагдлаа.',
        'archived':       'Төхөөрөмж архивлагдлаа.',
    }
    flash(msgs.get(reason, 'Төхөөрөмжийн төлөв шинэчлэгдлээ.') if lang=='mn' else 'Device status updated.', 'success')
    # Засварт шилжүүлсэн бол detail хуудас руу (дэлгэрэнгүй бүртгэхэд)
    if reason == 'repair':
        return redirect(url_for('device_detail', did=did) + '#tab-repair')
    return redirect(url_for('devices'))

@app.route('/devices/<int:did>/restore', methods=['POST'])
@senior_required
def device_restore(did):
    lang = session.get('lang','mn')
    conn = get_db()
    conn.execute("UPDATE devices SET status='active' WHERE id=?", (did,))
    conn.commit(); conn.close()
    flash('Төхөөрөмж сэргээгдлээ.' if lang=='mn' else 'Device restored.', 'success')
    return redirect(url_for('devices'))

@app.route('/devices/<int:did>/delete', methods=['POST'])
@admin_required
def device_delete(did):
    """Төхөөрөмжийг бүрмөсөн устгана (зөвхөн архивлагдсаныг, зөвхөн Админ)"""
    lang = session.get('lang','mn')
    conn = get_db()
    dev = conn.execute("SELECT status FROM devices WHERE id=?", (did,)).fetchone()
    if not dev:
        conn.close()
        flash('Төхөөрөмж олдсонгүй.', 'error')
        return redirect(url_for('archive'))
    if dev['status'] not in ('archived','replaced','decommissioned','standby'):
        conn.close()
        flash('Зөвхөн архивлагдсан төхөөрөмжийг бүрмөсөн устгана.', 'error')
        return redirect(url_for('archive'))
    # Холбоотой бичлэгүүдийг устгана
    conn.execute("DELETE FROM staff_device_permissions WHERE device_id=?", (did,))
    conn.execute("DELETE FROM calibrations WHERE device_id=?", (did,))
    conn.execute("DELETE FROM repairs WHERE device_id=?", (did,))
    conn.execute("DELETE FROM usage_logs WHERE device_id=?", (did,))
    conn.execute("DELETE FROM devices WHERE id=?", (did,))
    conn.commit(); conn.close()
    flash('Төхөөрөмж бүрмөсөн устгагдлаа.' if lang=='mn' else 'Device permanently deleted.', 'success')
    return redirect(url_for('archive'))

# ── MARKS ───────────────────────────────────────────────
@app.route('/marks/add', methods=['POST'])
@admin_required
def mark_add():
    conn = get_db()
    conn.execute("INSERT INTO device_marks(manufacturer,model,category) VALUES(?,?,?)",
                 (request.form['manufacturer'], request.form['model'], request.form.get('category')))
    mid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    name = f"{request.form['manufacturer']} {request.form['model']}"
    conn.commit(); conn.close()
    return jsonify({'success': True, 'id': mid, 'name': name})

# Хэмжилтийн хуудасны багана бүлэг → тухайн бүлгийн талбарууд.
# Архивт ХЭМЖСЭН шинжилгээг л харуулахад ашиглагдана.
COLUMN_GROUPS = {
    'ff':  ['ff_sample', 'ff_dried'],
    'mt':  ['mt_bux', 'mt_tare', 'mt_sample', 'mt_dried'],
    'dc':  ['dc_bux', 'dc_tare', 'dc_sample', 'dc_dried'],
    'un':  ['ash_tav', 'ash_tare', 'ash_sample', 'ash_burned'],
    'db':  ['vol_tig', 'vol_tare', 'vol_sample', 'vol_burned'],
    'gi':  ['g_tig', 'g_tare', 'g_coke', 'g_sieve1', 'g_sieve2', 'g_val'],
    'st':  ['sulfur'],
    'il':  ['cal_value', 'cal_temp'],
    'fsi': ['fsi'],
}



def conn2_analysed_date(entries):
    """Шинжилгээ дууссан хамгийн сүүлийн огноо (албан тайлангийн H13).

    Дээж бэлтгэлийн огноо БИШ — мөр бүрийн done_at/approved_at-аас сонгоно.
    """
    best = None
    for e in entries:
        for f in ('approved_at', 'done_at'):
            try:
                v = e[f]
            except (KeyError, IndexError):
                v = None
            if v and (best is None or str(v) > best):
                best = str(v)
    return best

def sample_names_for(receipt):
    """Ажлын дээж бүрийн нэрийг гаргана — (нэрсийн жагсаалт, нэр авах функц).

    sample_entries.sample_name нь autosave-аар л хадгалагддаг тул химич
    нэрийн нүдийг хөндөөгүй бол NULL үлддэг. Тэр үед geo_samples-ийн
    "A;B;C" жагсаалт эсвэл "ROCK-1 - ROCK-30" мужаас нэрийг сэргээнэ.
    Урьд нь үр дүнгийн хуудсанд л ажилладаг, архивын хуудсанд байхгүй тул
    архивлагдсан ажлын дээжийн нэр "—" болж алга болдог байв.
    """
    import re as _re
    sname = (receipt['sample_name'] or '') if receipt else ''
    parts = [x.strip() for x in sname.split(';')] if ';' in sname else None
    m1 = _re.match(r'^(.*?)(\d+)\s*[-–]\s*(\d+)\s*$', sname)
    m2 = _re.match(r'^(.*?)(\d+)$', sname) if not m1 else None

    def name_of(rn):
        if parts and len(parts) >= rn:
            return parts[rn - 1]
        if m1:
            return m1.group(1) + str(int(m1.group(2)) + rn - 1)
        if m2:
            return m2.group(1) + str(int(m2.group(2)) + rn - 1)
        return sname

    qty = (receipt['quantity'] or 1) if receipt else 1
    return [name_of(i) for i in range(1, qty + 1)], name_of, sname


def measured_groups(entries):
    """Аль шинжилгээнд утга орсоныг тогтооно (хоосныг нь харуулахгүй)"""
    out = []
    keys = set(entries[0].keys()) if entries else set()
    for g, fields in COLUMN_GROUPS.items():
        cols = [f for f in fields if f in keys]
        if any(e[f] is not None and e[f] != '' for e in entries for f in cols):
            out.append(g)
    return out


# ── ШИНЖИЛГЭЭ ТУС БҮРИЙН ГҮЙЦЭТГЭГЧ ──────────────────────
# Урьд нь sample_entries-д зөвхөн updated_by (хамгийн сүүлд бичсэн хүн) байсан
# тул нэг дээж дээр хоёр химич өөр өөр шинжилгээ хийвэл зөвхөн сүүлчийнх нь
# тоологддог байв. Одоо шинжилгээний төрөл бүрд гүйцэтгэгчийг тусад нь бичнэ.
ANALYSIS_OPS = [
    ('op_mt',  'Нийт чийг',   ['ff_sample', 'ff_dried',
                               'mt_bux', 'mt_tare', 'mt_sample', 'mt_dried']),
    ('op_mad', 'Дотоод чийг', ['dc_bux', 'dc_tare', 'dc_sample', 'dc_dried']),
    ('op_aad', 'Үнслэг',      ['ash_tav', 'ash_tare', 'ash_sample', 'ash_burned']),
    ('op_vad', 'Дэгдэмхий',   ['vol_tig', 'vol_tare', 'vol_sample', 'vol_burned']),
    ('op_g',   'G индекс',    ['g_tig', 'g_tare', 'g_coke', 'g_sieve1', 'g_sieve2']),
    ('op_st',  'Нийт хүхэр',  ['sulfur']),
    ('op_q',   'Илчлэг',      ['cal_value', 'cal_temp']),
    ('op_fsi', 'Чөлөөт хөөлт', ['fsi']),
]
# {талбарын нэр: гүйцэтгэгчийн багана}
FIELD_OP = {f: op for op, _lbl, fields in ANALYSIS_OPS for f in fields}
# Тухайн шинжилгээ хийгдсэн эсэхийг илтгэх нөхцөл (жин нь орсон эсэх)
OP_HAS_VALUE = {op: ' OR '.join(f'{f} IS NOT NULL' for f in fields)
                for op, _lbl, fields in ANALYSIS_OPS}


# ── DB MIGRATION (called once at startup) ───────────────
# ── Шилжилтийн алдааны бүртгэл ──────────────────────────────────────────
# ensure_tables доторх ALTER TABLE / CREATE TABLE нь хоёр дахь удаагаа
# ажиллахад "duplicate column name" гэж унах нь ХЭВИЙН — тэр нь багана
# аль хэдийн нэмэгдсэн гэсэн үг. Харин бусад алдаа нь шинэчлэлт дутуу
# хэрэгжсэн гэсэн үг: програм асна, гэхдээ хэдэн өдрийн дараа тэр багана
# хэрэгтэй болоход л алдаа гарна. Тиймээс чимээгүй өнгөрөх ёсгүй.
MIGRATION_WARNINGS = []
MIGRATIONS_COMPLETED = False
_MIG_EXPECTED = ('duplicate column name', 'already exists')


def _mig_warn(exc):
    msg = str(exc)
    if any(t in msg.lower() for t in _MIG_EXPECTED):
        return                      # хэвийн — багана/хүснэгт аль хэдийн байна
    MIGRATION_WARNINGS.append(msg)
    app.logger.warning('Шилжилт амжилтгүй: %s', msg)


def ensure_tables():
    conn = get_db()
    # Багана нь models.py-ийн тодорхойлолттой ижил байх ёстой — DB-г models.py
    # эхэлж үүсгэдэг тул зөрвөл энэ CREATE хүчингүй болж, INSERT нь уначихна.
    conn.execute("""CREATE TABLE IF NOT EXISTS lab_report_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_type TEXT NOT NULL,
        year INTEGER NOT NULL,
        period_value INTEGER,
        period_label TEXT NOT NULL,
        file_path TEXT,
        generated_by INTEGER REFERENCES users(id),
        generated_at TEXT DEFAULT CURRENT_TIMESTAMP,
        status TEXT DEFAULT 'active',
        archived_by INTEGER REFERENCES users(id),
        archived_at TEXT
    )""")
    # Химичийн нэгтгэсэн ажлын багц. Урьд нь нэгтгэл зөвхөн хаяган дээр
    # (?ids=2004,1001,5004) байсан тул хуудсаас гармагц алга болж, буцаж
    # ирэх бүрд дахин сонгож эрэмбэлэх шаардлагатай болдог байв.
    # receipt_ids — таслалаар тусгаарлагдсан, ДАРААЛАЛ нь утгатай.
    conn.execute("""CREATE TABLE IF NOT EXISTS work_batch (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER REFERENCES users(id),
        receipt_ids TEXT NOT NULL,
        qc_rows TEXT,
        qc_id INTEGER,
        status TEXT DEFAULT 'open',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        closed_at TEXT
    )""")
    # Хэмжилтийн утга ӨӨРЧЛӨГДӨХ/АРИЛАХ бүрийн бүртгэл. "Хариу нь байгаа
    # хэрнээ масс нь алга болчиж" гэдэг гомдол давтагдаж байсан ч шалтгааныг
    # нь мөшгих ул мөр байгаагүй. Одоо хуучин утга нь хадгалагдана —
    # tools/check_audit.py-аар харна, шаардлагатай бол сэргээнэ.
    conn.execute("""CREATE TABLE IF NOT EXISTS value_audit (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        receipt_id INTEGER,
        row_num INTEGER,
        is_duplicate INTEGER,
        field TEXT,
        old_value TEXT,
        new_value TEXT,
        user_id INTEGER,
        at TEXT,
        source TEXT
    )""")
    conn.execute("""CREATE INDEX IF NOT EXISTS idx_value_audit_at
                    ON value_audit(at)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS crm_materials (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        crm_name TEXT NOT NULL,
        mad_cert REAL, mad_unc REAL, aad_cert REAL, aad_unc REAL,
        vad_cert REAL, vad_unc REAL, sulfur_cert REAL, sulfur_unc REAL,
        cal_cert REAL, cal_unc REAL, notes TEXT,
        manufacture_date TEXT, expiry_date TEXT, open_date TEXT,
        is_active INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS guest_tokens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        token TEXT UNIQUE NOT NULL,
        label TEXT,
        created_by INTEGER REFERENCES users(id),
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        expires_at TEXT NOT NULL
    )""")
    for col in ['manufacture_date TEXT', 'expiry_date TEXT', 'open_date TEXT', 'g_cert REAL', 'g_unc REAL', 'standard TEXT',
                'mad_cert REAL', 'mad_unc REAL',
                'aad_cert REAL', 'aad_unc REAL', 'vad_cert REAL', 'vad_unc REAL',
                'sulfur_cert REAL', 'sulfur_unc REAL', 'cal_cert REAL', 'cal_unc REAL']:
        try: conn.execute(f"ALTER TABLE crm_materials ADD COLUMN {col}")
        except Exception as _e: _mig_warn(_e)
    # crm_g — G индексийн батлагдсан утга. Урьд нь crm_materials-д g_cert байсан
    # ч дээж бүртгэх үед хуулагддаггүй тул зөвхөн G хэмждэг CRM (ж: GBW12023c)
    # дээр харьцуулах зүйлгүй болж, хүснэгт хоосон гардаг байв.
    for col in ['crm_name TEXT','crm_mad REAL','crm_aad REAL','crm_vad REAL',
                'crm_sulfur REAL','crm_cal REAL','crm_g REAL','crm_g_unc REAL',
                'crm_mad_unc REAL','crm_aad_unc REAL',
                'crm_vad_unc REAL','crm_sulfur_unc REAL','crm_cal_unc REAL','sample_range TEXT']:
        try: conn.execute(f"ALTER TABLE geo_samples ADD COLUMN {col}")
        except Exception as _e: _mig_warn(_e)
    # Өмнө нь бүртгэгдсэн CRM дээжид G (ба Mad) батлагдсан утгыг нөхөж дүүргэнэ.
    # Зөвхөн ХООСОН талбарыг дүүргэх тул гараар засварласан утга хөндөгдөхгүй.
    try:
        conn.execute("""
            UPDATE geo_samples SET
              crm_g          = COALESCE(crm_g,          (SELECT g_cert      FROM crm_materials m WHERE m.crm_name=geo_samples.crm_name)),
              crm_g_unc      = COALESCE(crm_g_unc,      (SELECT g_unc       FROM crm_materials m WHERE m.crm_name=geo_samples.crm_name)),
              crm_mad        = COALESCE(crm_mad,        (SELECT mad_cert    FROM crm_materials m WHERE m.crm_name=geo_samples.crm_name)),
              crm_mad_unc    = COALESCE(crm_mad_unc,    (SELECT mad_unc     FROM crm_materials m WHERE m.crm_name=geo_samples.crm_name)),
              crm_aad_unc    = COALESCE(crm_aad_unc,    (SELECT aad_unc     FROM crm_materials m WHERE m.crm_name=geo_samples.crm_name)),
              crm_vad_unc    = COALESCE(crm_vad_unc,    (SELECT vad_unc     FROM crm_materials m WHERE m.crm_name=geo_samples.crm_name)),
              crm_sulfur_unc = COALESCE(crm_sulfur_unc, (SELECT sulfur_unc  FROM crm_materials m WHERE m.crm_name=geo_samples.crm_name)),
              crm_cal_unc    = COALESCE(crm_cal_unc,    (SELECT cal_unc     FROM crm_materials m WHERE m.crm_name=geo_samples.crm_name))
            WHERE sample_type='CRM' AND crm_name IS NOT NULL""")
    except Exception:
        pass
    # Шинжилгээ тус бүрийн гүйцэтгэгч
    for op, _lbl, _f in ANALYSIS_OPS:
        try: conn.execute(f'ALTER TABLE sample_entries ADD COLUMN {op} INTEGER REFERENCES users(id)')
        except Exception as _e: _mig_warn(_e)
    # Хуучин бичлэгт гүйцэтгэгч тэмдэглэгдээгүй тул updated_by/done_by-гаар
    # ойролцоогоор нөхнө. Тухайн шинжилгээний жин орсон мөрөнд л бичигдэнэ.
    for op, _lbl, _f in ANALYSIS_OPS:
        try:
            conn.execute(f"""UPDATE sample_entries
                SET {op} = COALESCE(updated_by, done_by)
                WHERE {op} IS NULL AND COALESCE(updated_by, done_by) IS NOT NULL
                  AND ({OP_HAS_VALUE[op]})""")
        except Exception:
            pass
    # ── ОРЧНЫ ХЯНАЛТ: өрөө + өдөр бүрийн чийг/дулааны бүртгэл ──
    conn.execute("""CREATE TABLE IF NOT EXISTS env_rooms (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        sort_order INTEGER DEFAULT 0,
        is_active INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS env_readings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        room_id INTEGER NOT NULL REFERENCES env_rooms(id),
        reading_date TEXT NOT NULL,
        slot TEXT NOT NULL,               -- 'start' = шинжилгээ эхлэхэд, 'end' = төгсгөлд
        temperature REAL,
        humidity REAL,
        recorded_by INTEGER REFERENCES users(id),
        recorded_at TEXT,
        notes TEXT,
        UNIQUE(room_id, reading_date, slot)
    )""")
    # Анхны өрөөнүүд (зөвхөн хүснэгт хоосон үед)
    if conn.execute("SELECT COUNT(*) FROM env_rooms").fetchone()[0] == 0:
        for i, nm in enumerate(['Шинжилгээний өрөө №115', 'Шинжилгээний өрөө №116',
                                'Шинжилгээний өрөө №117', 'Дээж бэлтгэлийн байр']):
            conn.execute("INSERT INTO env_rooms(name, sort_order) VALUES(?,?)", (nm, i))

    # Бэлтгэгчийг хэрэглэгчийн ID-гаар бүртгэнэ (нэрээр тоолох найдваргүй)
    try: conn.execute("ALTER TABLE sample_receipt ADD COLUMN prep_by INTEGER REFERENCES users(id)")
    except Exception as _e: _mig_warn(_e)
    for col in ['prep_operator TEXT', 'prep_position TEXT', 'prep_devices TEXT']:
        try: conn.execute(f"ALTER TABLE sample_receipt ADD COLUMN {col}")
        except Exception as _e: _mig_warn(_e)
    # Төхөөрөмжийн дэлгэрэнгүй паспорт талбарууд
    for col in ['web_link TEXT','method TEXT','max_temp TEXT','particular TEXT',
                'measuring_time TEXT','measuring_limit TEXT','dimension TEXT','capacity TEXT',
                'weight_kg TEXT','other_spec TEXT','power TEXT','frequency TEXT','voltage TEXT',
                'specification TEXT','operating_state TEXT','received_date TEXT','lab_id TEXT',
                'check_standard TEXT','check_tolerance TEXT','check_unit TEXT',
                'check_enabled INTEGER DEFAULT 1','stage TEXT DEFAULT \'both\'',
                'check_freq TEXT DEFAULT \'daily\'']:
        try: conn.execute(f"ALTER TABLE devices ADD COLUMN {col}")
        except Exception as _e: _mig_warn(_e)
    # Дотоод өдөр тутмын шалгалт (жин г.м.)
    conn.execute("""CREATE TABLE IF NOT EXISTS device_checks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        device_id INTEGER NOT NULL REFERENCES devices(id),
        checked_by INTEGER REFERENCES users(id),
        check_date TEXT NOT NULL,
        standard_value TEXT,
        measured_value TEXT,
        tolerance TEXT,
        result TEXT DEFAULT 'pass',
        calibration_adjusted INTEGER DEFAULT 0,
        notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    for col in ['calibration_adjusted INTEGER DEFAULT 0',
                'measured_value2 TEXT', 'standard_value2 TEXT', 'tolerance2 TEXT']:
        try: conn.execute(f"ALTER TABLE device_checks ADD COLUMN {col}")
        except Exception as _e: _mig_warn(_e)
    for col in ['check_standard2 TEXT', 'check_tolerance2 TEXT',
                'check_param1 TEXT', 'check_param2 TEXT',
                'check_param3 TEXT', 'check_standard3 TEXT', 'check_tolerance3 TEXT',
                'check_param4 TEXT', 'check_standard4 TEXT', 'check_tolerance4 TEXT',
                'check_param5 TEXT', 'check_standard5 TEXT', 'check_tolerance5 TEXT']:
        try: conn.execute(f"ALTER TABLE devices ADD COLUMN {col}")
        except Exception as _e: _mig_warn(_e)
    for col in ['check_group1 TEXT', 'check_group2 TEXT', 'check_group3 TEXT',
                'check_group1_cols INTEGER', 'check_group2_cols INTEGER', 'check_group3_cols INTEGER',
                'check_photo1 TEXT', 'check_photo2 TEXT', 'check_photo3 TEXT', 'check_photo4 TEXT', 'check_photo5 TEXT']:
        try: conn.execute(f"ALTER TABLE devices ADD COLUMN {col}")
        except Exception as _e: _mig_warn(_e)
    for col in ['measured_value3 TEXT', 'standard_value3 TEXT', 'tolerance3 TEXT',
                'measured_value4 TEXT', 'standard_value4 TEXT', 'tolerance4 TEXT',
                'measured_value5 TEXT', 'standard_value5 TEXT', 'tolerance5 TEXT']:
        try: conn.execute(f"ALTER TABLE device_checks ADD COLUMN {col}")
        except Exception as _e: _mig_warn(_e)
    try: conn.execute("ALTER TABLE device_checks ADD COLUMN photo TEXT")
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("ALTER TABLE users ADD COLUMN shift TEXT")
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("ALTER TABLE users ADD COLUMN can_register INTEGER DEFAULT 0")
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("ALTER TABLE users ADD COLUMN can_view_result INTEGER DEFAULT 0")
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("ALTER TABLE users ADD COLUMN can_export INTEGER DEFAULT 0")
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("ALTER TABLE users ADD COLUMN can_approve INTEGER DEFAULT 0")
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("ALTER TABLE users ADD COLUMN can_report INTEGER DEFAULT 0")
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("ALTER TABLE users ADD COLUMN can_reopen INTEGER DEFAULT 0")
    except Exception as _e: _mig_warn(_e)
    # Устгагдсан ажилтан: бүх жагсаалтаас алга болно, гэхдээ нэр нь хуучин
    # шинжилгээ/тайланд хадгалагдаж үлдэнэ (мөшгих чадвар алдагдахгүй)
    try: conn.execute("ALTER TABLE users ADD COLUMN is_deleted INTEGER DEFAULT 0")
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("ALTER TABLE users ADD COLUMN deleted_at TEXT")
    except Exception as _e: _mig_warn(_e)
    # Ажлаас гарсан (Гаргах) ба түр амарсан (Амраах) хоёрыг ялгана.
    # Өмнө нь хоёул is_active=0 тавьдаг байсан тул амарсан ажилтан архивт
    # ажлаас гарсантай хамт орж, ажилтны хуудаснаас алга болдог байсан.
    _new_dismissed = False
    try:
        conn.execute("ALTER TABLE users ADD COLUMN is_dismissed INTEGER DEFAULT 0")
        _new_dismissed = True
    except Exception as _e: _mig_warn(_e)
    if _new_dismissed:
        # Хуучин өгөгдлийг ялгах: ээлжгүй + идэвхгүй = ажлаас гарсан,
        # ээлжтэй + идэвхгүй = зүгээр амарсан (ажилтны хуудсанд үлдэнэ)
        conn.execute("""UPDATE users SET is_dismissed=1
                        WHERE is_active=0 AND (shift IS NULL OR shift='')""")
    # Геологчийн харах эрхтэй ажлын дугаарын муж (мянгатаар: "1,2,6"). NULL = бүгдийг харах
    try: conn.execute("ALTER TABLE users ADD COLUMN view_ranges TEXT")
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("""
        CREATE TABLE IF NOT EXISTS result_view_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER REFERENCES users(id),
            receipt_id INTEGER REFERENCES sample_receipt(id),
            viewed_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("""
        CREATE TABLE IF NOT EXISTS check_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            param1 TEXT, standard1 TEXT, tolerance1 TEXT,
            param2 TEXT, standard2 TEXT, tolerance2 TEXT,
            param3 TEXT, standard3 TEXT, tolerance3 TEXT,
            param4 TEXT, standard4 TEXT, tolerance4 TEXT,
            created_by INTEGER REFERENCES users(id),
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("""
        CREATE TABLE IF NOT EXISTS device_calib_checks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_id INTEGER NOT NULL REFERENCES devices(id),
            checked_by INTEGER REFERENCES users(id),
            check_date TEXT NOT NULL,
            x1 REAL, y1 REAL,
            x2 REAL, y2 REAL,
            x3 REAL, y3 REAL,
            x4 REAL, y4 REAL,
            x5 REAL, y5 REAL,
            slope_b REAL,
            intercept_a REAL,
            r_squared REAL,
            result TEXT DEFAULT 'pass',
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)
    except Exception as _e: _mig_warn(_e)
    # ── Төхөөрөмжийн автомат тохиргоо — ЗӨВХӨН НЭГ УДАА ажиллана ──
    # (PRAGMA user_version=1 болсон бол дахин ажиллахгүй — user/гар тохиргоог хадгална)
    if conn.execute("PRAGMA user_version").fetchone()[0] < 1:
      # Лаб 01-05: жингийн босоо загвар тэмдэглэх (check_group1_cols=0)
      conn.execute("""
        UPDATE devices SET check_group1_cols=0
        WHERE CAST(lab_id AS TEXT) IN ('01','02','03','04','05','001','002','003','004','005')
      """)
      # Лаб 02-05: лаб 01-ийн жингийн тохиргоог хуулна (photo-оос бусад)
      conn.execute("""
        UPDATE devices SET
            check_param1   = (SELECT check_param1   FROM devices WHERE lab_id='01' LIMIT 1),
            check_standard = (SELECT check_standard FROM devices WHERE lab_id='01' LIMIT 1),
            check_tolerance= (SELECT check_tolerance FROM devices WHERE lab_id='01' LIMIT 1),
            check_group1   = (SELECT check_group1   FROM devices WHERE lab_id='01' LIMIT 1),
            check_freq     = (SELECT check_freq     FROM devices WHERE lab_id='01' LIMIT 1),
            check_enabled  = (SELECT check_enabled  FROM devices WHERE lab_id='01' LIMIT 1),
            check_group1_cols = 0
        WHERE CAST(lab_id AS TEXT) IN ('02','03','04','05','002','003','004','005')
      """)
      # Лаб 11-14: лаб 06-тай ижил (зуухны температур, daily)
      conn.execute("""
        UPDATE devices SET
            check_group1='Зуухны температур', check_group2=NULL, check_group3=NULL,
            check_param1='Температур 1', check_standard='850±10 °C', check_tolerance='6 мин - 850 °C',
            check_param2='Температур 2', check_standard2='900±20 °C', check_tolerance2='3 мин - 900 °C',
            check_param3=NULL, check_standard3=NULL, check_tolerance3=NULL,
            check_param4=NULL, check_standard4=NULL, check_tolerance4=NULL,
            check_freq=COALESCE(check_freq,'daily'), check_enabled=1, check_group1_cols=4
        WHERE CAST(lab_id AS TEXT) IN ('11','12','13','14','011','012','013','014')
           OR LOWER(name) LIKE '%барабан%'
      """)
      # Зуух (муфель зуух) — нэрээр тааруулна
      conn.execute("""
        UPDATE devices SET
            check_group1='Зуухны температур', check_group2=NULL, check_group3=NULL,
            check_param1='Температур 1', check_standard='850±10 °C', check_tolerance='6 мин - 850 °C',
            check_param2='Температур 2', check_standard2='900±20 °C', check_tolerance2='3 мин - 900 °C',
            check_param3=NULL, check_standard3=NULL, check_tolerance3=NULL,
            check_param4=NULL, check_standard4=NULL, check_tolerance4=NULL,
            check_param5=NULL, check_standard5=NULL, check_tolerance5=NULL,
            check_freq=COALESCE(check_freq,'daily'), check_enabled=1, check_group1_cols=4
        WHERE CAST(lab_id AS TEXT) IN ('06','07','08','006','007','008')
           OR LOWER(name) LIKE '%муфель%' OR LOWER(name) LIKE '%muffle%'
           OR LOWER(name) LIKE '%зуух%'
      """)
      # Хатаах шүүгээ — нэрээр тааруулна
      conn.execute("""
        UPDATE devices SET
            check_group1='Зуухны температур', check_group2=NULL, check_group3=NULL,
            check_param1='Температур 1', check_standard='105±2 °C', check_tolerance='±2 °C',
            check_param2='Температур 2', check_standard2='105±2 °C', check_tolerance2='±2 °C',
            check_param3=NULL, check_standard3=NULL, check_tolerance3=NULL,
            check_param4=NULL, check_standard4=NULL, check_tolerance4=NULL,
            check_param5=NULL, check_standard5=NULL, check_tolerance5=NULL,
            check_freq=COALESCE(check_freq,'daily'), check_enabled=1, check_group1_cols=4
        WHERE CAST(lab_id AS TEXT) IN ('19','20','21','019','020','021')
           OR LOWER(name) LIKE '%хатаах%' OR LOWER(name) LIKE '%шүүгээ%'
           OR LOWER(name) LIKE '%drying%'
      """)
      # Холигч — нэрээр тааруулна
      conn.execute("""
        UPDATE devices SET
            check_group1='Зуухны температур', check_group2=NULL, check_group3=NULL,
            check_param1='Температур 1', check_standard='850±10 °C', check_tolerance='6 мин - 850 °C',
            check_param2='Температур 2', check_standard2='900±20 °C', check_tolerance2='3 мин - 900 °C',
            check_param3=NULL, check_standard3=NULL, check_tolerance3=NULL,
            check_param4=NULL, check_standard4=NULL, check_tolerance4=NULL,
            check_param5=NULL, check_standard5=NULL, check_tolerance5=NULL,
            check_freq=COALESCE(check_freq,'daily'), check_enabled=1, check_group1_cols=4
        WHERE CAST(lab_id AS TEXT) IN ('15','16','015','016')
           OR LOWER(name) LIKE '%холигч%' OR LOWER(name) LIKE '%mixer%'
      """)
      # Хөөлтийн зэрэг тодорхойлох багаж — зуухтай ижил
      conn.execute("""
        UPDATE devices SET
            check_group1='Зуухны температур', check_group2=NULL, check_group3=NULL,
            check_param1='Температур 1', check_standard='850±10 °C', check_tolerance='6 мин - 850 °C',
            check_param2='Температур 2', check_standard2='900±20 °C', check_tolerance2='3 мин - 900 °C',
            check_param3=NULL, check_standard3=NULL, check_tolerance3=NULL,
            check_param4=NULL, check_standard4=NULL, check_tolerance4=NULL,
            check_param5=NULL, check_standard5=NULL, check_tolerance5=NULL,
            check_freq=COALESCE(check_freq,'daily'), check_enabled=1, check_group1_cols=4
        WHERE name LIKE '%Хөөлт%' OR name LIKE '%хөөлт%'
           OR name LIKE '%Алхан%' OR name LIKE '%алхан%'
           OR name LIKE '%Аяган%' OR name LIKE '%аяган%'
           OR name LIKE '%Роторт%' OR name LIKE '%роторт%'
      """)
      # Калориметр — сарын шалгалт
      conn.execute("""
        UPDATE devices SET
            check_freq=COALESCE(NULLIF(check_freq,'daily'),'monthly'),
            check_enabled=1,
            check_param1=NULL, check_standard=NULL, check_tolerance=NULL,
            check_param2=NULL, check_standard2=NULL, check_tolerance2=NULL,
            check_group1='Калориметрийн шалгалт'
        WHERE LOWER(name) LIKE '%калориметр%'
      """)
      # Хүхрийн багаж — сарын калибровкийн шалгалт
      conn.execute("""
        UPDATE devices SET
            check_freq=COALESCE(NULLIF(check_freq,'daily'),'monthly'),
            check_enabled=1,
            check_param1=NULL, check_standard=NULL, check_tolerance=NULL,
            check_param2=NULL, check_standard2=NULL, check_tolerance2=NULL,
            check_group1='Шугман регрессийн калибровк'
        WHERE LOWER(name) LIKE '%хүхэр%' OR LOWER(name) LIKE '%sulfur%'
      """)
      # Буруу орсон check_param5='5' утгыг цэвэрлэнэ
      conn.execute("UPDATE devices SET check_param5=NULL, check_standard5=NULL, check_tolerance5=NULL WHERE check_param5='5'")
      # Энэ блок дахин ажиллахгүй болгоно
      conn.execute("PRAGMA user_version=1")
    # ЗААВАЛ: CREATE эхэлж, ALTER дараа нь. Урвуу дараалалтай байхад шинэ
    # мэдээллийн санд ALTER амжилтгүй болж (хүснэгт хараахан байхгүй), улмаас
    # хүснэгт дутуу баганатай үүсээд /analysis хуудас 500 алдаа өгдөг байсан.
    conn.execute("""CREATE TABLE IF NOT EXISTS internal_qc (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        qc_number TEXT,
        triggered_date TEXT NOT NULL,
        receipt_id_1 INTEGER REFERENCES sample_receipt(id),
        receipt_id_2 INTEGER REFERENCES sample_receipt(id),
        row_num_1 INTEGER,
        row_num_2 INTEGER,
        assigned_to INTEGER REFERENCES users(id),
        status TEXT DEFAULT 'pending',
        notes TEXT,
        created_by INTEGER REFERENCES users(id),
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    # Хуучин мэдээллийн санг шинэчлэх (аль хэдийн байвал алдаа өгөхгүй)
    try: conn.execute("ALTER TABLE internal_qc ADD COLUMN qc_number TEXT")
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("ALTER TABLE internal_qc ADD COLUMN row_num_1 INTEGER")
    except Exception as _e: _mig_warn(_e)
    try: conn.execute("ALTER TABLE internal_qc ADD COLUMN row_num_2 INTEGER")
    except Exception as _e: _mig_warn(_e)
    conn.execute("""CREATE TABLE IF NOT EXISTS internal_qc_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        qc_id INTEGER REFERENCES internal_qc(id),
        receipt_id INTEGER REFERENCES sample_receipt(id),
        parameter TEXT,
        original_value REAL,
        repeat_value REAL,
        difference REAL,
        within_tolerance INTEGER DEFAULT 0,
        notes TEXT
    )""")
    conn.commit()
    conn.close()
    # Бүх шилжилт эцэс хүртэл ажилласны тэмдэг. ensure_tables доторх зарим
    # үйлдэл try/except-гүй тул дунд нь унавал үлдсэн нь огт ажиллахгүй —
    # healthcheck энэ тугаар дутуу шинэчлэлтийг илрүүлнэ.
    global MIGRATIONS_COMPLETED
    MIGRATIONS_COMPLETED = True

# ── REPORTS ─────────────────────────────────────────────
# Дээжийн болон шинжилгээний төрлүүд — тайлангийн бүх хуудас нэг эх сурвалжаас
# уншина (график ба Excel хоёр зөрөхөөс сэргийлнэ).
SAMPLE_TYPES_MAP = [
    ('PIT','Уурхай'),('STOCKPILE','Овоолго'),('EXPORT','Ачилт'),
    ('CONTROL','Хяналт'),('DP','Баяжуулах'),('EQ_CONTROL','Гадаад хяналт'),
]
ANALYSIS_FIELDS = [
    ('mt_dried','Нийт чийг'),('mad','Дотоод чийг'),('aad','Үнслэг'),('vad','Дэгдэмхий'),
    ('sulfur','Хүхэр'),('cal_value','Илчлэг'),('g_coke','G индекс'),('fsi','Чөлөөт хөөлтийн зэрэг'),
]
# G индексийг гараар (g_val) эсвэл жингээр (g_coke) оруулж болно — хоёуланг тооцно
ANALYSIS_COUNT_COL = {'g_coke': 'COALESCE(se.g_val, se.g_coke)'}
# Шинжилгээг ХЭМЖИЛТ бүрээр тоолно — зэрэгцээ, давталт нь тус тусдаа хийгдсэн
# ажил тул тухайн үзүүлэлтийн тоонд нэмэгдэнэ.
#
# Огноог мөрийнхөө ҮНДСЭН хэмжилтээс (is_duplicate=0) авна. Учир нь зэрэгцээ,
# давталтын мөрөнд "Дууслаа" товч байдаггүй тул тэдгээрийн done_at ҮРГЭЛЖ
# хоосон байдаг. Урьд нь шууд se.done_at-аар шүүдэг байсан тул давхар хэмжилт
# бүр тоонд ОРОЛГҮЙ унаж, 100 дээж дуусахад үзүүлэлт 120-130 байх ёстой атал
# жишиг шугамаас доогуур гардаг байв.
# row_status нөхцөл нь жишиг шугамтай ЯГ ижил дээжийн багцыг сонгоно.
ANALYSIS_COUNT_SQL = '''SELECT COUNT(*) FROM sample_entries se
    JOIN sample_entries p ON p.receipt_id    = se.receipt_id
                         AND p.row_num       = se.row_num
                         AND p.is_duplicate  = 0
    WHERE {col} IS NOT NULL
      AND p.row_status IN ('done','approved')
      AND substr(p.done_at,1,10) BETWEEN ? AND ?'''
# Анхаар: CRM-ийн хэмжилт ЭНД тоологдоно. CRM бол чанарын хяналтын ажил
# бөгөөд лаборатори үнэхээр хэмжилт хийсэн тул үзүүлэлтийн тоонд орно.
# Харин ДЭЭЖИЙН тоонд (жишиг шугам) ордоггүй — доорх SAMPLE_DONE_SQL үзнэ үү.


def analysis_count_sql(field):
    """Тухайн үзүүлэлтийн хэмжилтийн тоог гаргах SQL"""
    return ANALYSIS_COUNT_SQL.format(col=ANALYSIS_COUNT_COL.get(field, f'se.{field}'))

# ── Графикийн "Дээж" шугам — ШИНЖИЛГЭЭ БҮРЭН ДУУССАН дээжийн тоо ─────────
# Тулгуур огноо нь бэлтгэлийн огноо БИШ, мөр бүрийн шинжилгээ дууссан огноо
# (se.done_at). Ж: 100 дээжтэй ажлын бэлтгэл өмнөх долоо хоногт хийгдсэн ч
# тэр долоо хоногт зөвхөн 40 дээжийн шинжилгээ бүрэн дууссан бол 40 гэж
# тоологдоно; үлдсэн 60 нь дуусах долоо хоногтоо очно. Ингэснээр үзүүлэлт
# бүрийн тоо энэ шугамтай нэг сууриар харьцуулагдана.
# is_duplicate=0 — зэрэгцээ, давталт нь тусдаа дээж биш, нэг мөрийн давтан
# хэмжилт тул энд тоологдохгүй (үзүүлэлтийн баганад л нэмэгдэнэ).
#
# CRM (баталгаажсан стандарт дээж) нь ҮЙЛДВЭРЛЭЛИЙН дээж биш, чанарын
# хяналтын материал бөгөөд бүх үзүүлэлтийг үзүүлдэггүй (ж: нийт чийг).
# Дээжийн тоонд оруулбал жишиг шугам хиймлээр өсч, багана нь түүнээс
# доогуур унасан мэт харагддаг тул ЭНД хасагдана.
# Харин CRM дээр хийсэн ХЭМЖИЛТ нь үзүүлэлтийн тоонд хэвээр орно —
# лаборатори тэр ажлыг үнэхээр хийсэн (ANALYSIS_COUNT_SQL үзнэ үү).
SAMPLE_DONE_SQL = '''SELECT COUNT(*) FROM sample_entries se
                     JOIN sample_receipt sr ON sr.id = se.receipt_id
                     JOIN geo_samples   g  ON g.id  = sr.geo_sample_id
                     WHERE se.is_duplicate = 0
                       AND g.sample_type <> 'CRM'
                       AND se.row_status IN ('done', 'approved')
                       AND substr(se.done_at, 1, 10) BETWEEN ? AND ?'''


def lab_period_range(rtype, year, month=1, week=1, half=1):
    """Тайлангийн хугацааг (эхлэх огноо, дуусах огноо, гарчиг) болгож буцаана.

    График ба Excel тайлан хоёр ижил хугацаа авахын тулд энэ функцээр дамжина.
    """
    import calendar as cal_mod
    import datetime as dt_mod
    if rtype == 'week':
        d0 = dt_mod.datetime.fromisocalendar(year, week, 1).date()
        d1 = dt_mod.datetime.fromisocalendar(year, week, 7).date()
        return d0.isoformat(), d1.isoformat(), f"{year} оны {week}-р долоо хоног ({d0} – {d1})"
    if rtype == 'month':
        _, ld = cal_mod.monthrange(year, month)
        return (f"{year}-{month:02d}-01", f"{year}-{month:02d}-{ld:02d}",
                f"{year} оны {month}-р сар")
    if rtype == 'half':
        if half == 1:
            return f"{year}-01-01", f"{year}-06-30", f"{year} оны эхний хагас жил"
        return f"{year}-07-01", f"{year}-12-31", f"{year} оны хоёрдугаар хагас жил"
    return f"{year}-01-01", f"{year}-12-31", f"{year} он"


def _record_dl_query(rec):
    """Хадгалсан тайлангийн бичлэгээс дахин татах query үүсгэнэ"""
    t = rec['period_type'] or 'month'
    q = f"type={t}&year={rec['year'] or datetime.now().year}"
    v = rec['period_value']
    if v and t in ('month', 'week', 'half'):
        q += f"&{t}={v}"
    return q


@app.route('/reports')
@perm_required('can_report')
def reports():
    conn = get_db()
    try:
        lab_records = conn.execute(
            '''SELECT r.*, u.name as gen_name FROM lab_report_records r
               LEFT JOIN users u ON u.id=r.generated_by
               WHERE r.status='active' ORDER BY r.generated_at DESC LIMIT 30'''
        ).fetchall()
    except Exception:
        lab_records = []
    # Бичлэг бүрийг дахин татах холбоос
    lab_records = [dict(r, dl_query=_record_dl_query(r)) for r in lab_records]
    iqc_rows = conn.execute("""
        SELECT iq.id, iq.triggered_date, iq.status,
            sr1.lab_number as lab1, g1.sample_name as sname1,
            sr2.lab_number as lab2, g2.sample_name as sname2,
            COUNT(r.id) as total,
            SUM(r.within_tolerance) as passed
        FROM internal_qc iq
        LEFT JOIN sample_receipt sr1 ON sr1.id=iq.receipt_id_1
        LEFT JOIN geo_samples g1 ON g1.id=sr1.geo_sample_id
        LEFT JOIN sample_receipt sr2 ON sr2.id=iq.receipt_id_2
        LEFT JOIN geo_samples g2 ON g2.id=sr2.geo_sample_id
        LEFT JOIN internal_qc_results r ON r.qc_id=iq.id
        GROUP BY iq.id ORDER BY iq.created_at DESC
    """).fetchall()
    conn.close()
    return render_template('admin/reports.html', lang=session.get('lang','mn'),
        lab_records=lab_records,
        sample_types=[name for _, name in SAMPLE_TYPES_MAP],
        analysis_types=[name for _, name in ANALYSIS_FIELDS],
        iqc_rows=iqc_rows,
        # Орчны хяналтын нэгтгэл (Тайлангийн 4 дэх таб)
        **env_stats_data(request.args.get('env_period', 'month')))

@app.route('/reports/chart-data')
@perm_required('can_report')
def reports_chart_data():
    rtype  = request.args.get('type', 'month')
    year   = int(request.args.get('year', datetime.now().year))
    month  = int(request.args.get('month', datetime.now().month))
    week   = int(request.args.get('week', 1))
    half   = int(request.args.get('half', 1))
    d0s, d1s, _ = lab_period_range(rtype, year, month, week, half)
    conn = get_db()
    # Дээжийн тоо — АЖЛААР биш ДЭЭЖЭЭР (нэг ажилд quantity дээж байна)
    sample_totals = []
    for code, name in SAMPLE_TYPES_MAP:
        v = conn.execute(
            '''SELECT COALESCE(SUM(COALESCE(g.quantity,1)),0) FROM geo_samples g
               JOIN sample_receipt sr ON sr.geo_sample_id=g.id
               WHERE g.sample_type=? AND sr.received_date BETWEEN ? AND ?''',
            (code, d0s, d1s)).fetchone()[0]
        sample_totals.append(v)
    # done_at нь isoformat ('2026-07-31T14:30') тул зайтай хязгаартай
    # харьцуулбал 'T' > ' ' болж СҮҮЛИЙН ӨДӨР бүхэлдээ хасагддаг.
    # Огноогоор шүүх нь тусгаарлагчаас хамаарахгүй.
    analysis_totals = []
    # Багана нь жишиг шугамаас доогуур байвал ЯАГААД болохыг тайлбарлана:
    # тухайн үзүүлэлт хэдэн дууссан дээж дээр огт хэмжигдээгүй, тэдгээр нь
    # ямар төрлийн дээж вэ. Урьд нь консол дээр tools/report_check.py
    # ажиллуулж байж мэддэг байсныг графикийн зөвлөмжинд шууд гаргана.
    analysis_missing = []
    for field, name in ANALYSIS_FIELDS:
        v = conn.execute(analysis_count_sql(field), (d0s, d1s)).fetchone()[0]
        analysis_totals.append(v)
        col = ANALYSIS_COUNT_COL.get(field, f'se.{field}')
        types = conn.execute(
            f"""SELECT g.sample_type t, COUNT(*) n
                  FROM sample_entries se
                  JOIN sample_receipt sr ON sr.id = se.receipt_id
                  JOIN geo_samples g ON g.id = sr.geo_sample_id
                 WHERE se.is_duplicate = 0
                   AND g.sample_type <> 'CRM'
                   AND se.row_status IN ('done','approved')
                   AND substr(se.done_at,1,10) BETWEEN ? AND ?
                   AND {col} IS NULL
                 GROUP BY g.sample_type ORDER BY n DESC""", (d0s, d1s)).fetchall()
        analysis_missing.append({'n': sum(r['n'] for r in types),
                                 'by_type': [[r['t'] or '—', r['n']] for r in types[:4]]})
    # Жишиг шугам — энэ хугацаанд ШИНЖИЛГЭЭ НЬ БҮРЭН ДУУССАН дээжийн тоо.
    # Урьд нь хүлээн авсан, дараа нь бэлтгэл дууссан огноогоор тоолдог байсан
    # тул үзүүлэлтийн тоотой өөр сууриар харьцуулагдаж зөрдөг байв.
    prep_total = conn.execute(SAMPLE_DONE_SQL, (d0s, d1s)).fetchone()[0]
    conn.close()
    return jsonify({
        'sample_labels': [n for _, n in SAMPLE_TYPES_MAP],
        'sample_data': sample_totals,
        'analysis_labels': [n for _, n in ANALYSIS_FIELDS],
        'analysis_data': analysis_totals,
        'analysis_missing': analysis_missing,
        'prep_total': prep_total,
    })

@app.route('/reports/export')
@perm_required('can_report')
def report_export():
    if session.get('role') == 'guest':
        flash('Зочин горимд Excel татах боломжгүй.', 'error')
        return redirect(url_for('reports'))
    import datetime as dt_mod
    rtype   = request.args.get('type', 'month')
    year    = int(request.args.get('year', datetime.now().year))
    month   = int(request.args.get('month', datetime.now().month))
    quarter = int(request.args.get('quarter', 1))
    half    = int(request.args.get('half', 1))
    week    = int(request.args.get('week', datetime.now().isocalendar()[1]))

    date_filter_mode = 'ym'  # 'ym' or 'range'

    if rtype == 'week':
        # ISO week → эхлэх/дуусах огноо
        week_start = dt_mod.datetime.fromisocalendar(year, week, 1).date()
        week_end   = dt_mod.datetime.fromisocalendar(year, week, 7).date()
        period_label = f"{year} оны {week}-р 7 хоног ({week_start} – {week_end})"
        date_filter_mode = 'range'
        date_start = str(week_start)
        date_end   = str(week_end)
    elif rtype == 'month':
        months = [month]
        period_label = f"{year} оны {month}-р сар"
    elif rtype == 'quarter':
        qm = {1:[1,2,3], 2:[4,5,6], 3:[7,8,9], 4:[10,11,12]}
        months = qm[quarter]
        period_label = f"{year} оны {quarter}-р улирал"
    elif rtype == 'half':
        months = list(range(1,7)) if half==1 else list(range(7,13))
        period_label = f"{year} оны {'эхний' if half==1 else 'хоёрдугаар'} хагас жил"
    else:
        months = list(range(1,13))
        period_label = f"{year} он"

    if date_filter_mode == 'range':
        wu  = f"date(ul.start_time) BETWEEN '{date_start}' AND '{date_end}'"
        wc  = f"c.calibration_date BETWEEN '{date_start}' AND '{date_end}'"
        wr  = f"r.reported_date BETWEEN '{date_start}' AND '{date_end}'"
        wd  = f"date(start_time) BETWEEN '{date_start}' AND '{date_end}'"
        ws_filter = f"sr.received_date BETWEEN '{date_start}' AND '{date_end}'"
    else:
        ym_list = ",".join([f"'{year}-{m:02d}'" for m in months])
        wu  = f"strftime('%Y-%m',ul.start_time) IN ({ym_list})"
        wc  = f"strftime('%Y-%m',c.calibration_date) IN ({ym_list})"
        wr  = f"strftime('%Y-%m',r.reported_date) IN ({ym_list})"
        wd  = f"strftime('%Y-%m',start_time) IN ({ym_list})"
        ws_filter = f"strftime('%Y-%m', sr.received_date) IN ({ym_list})"

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    conn = get_db()
    NAVY='1A2744'; TEAL='0F6E56'; CORAL='993C1D'; PURPLE='3C3489'
    WHITE='FFFFFF'; GRAY='F7F7F5'

    def th():
        s=Side(style='thin',color='CCCCCC')
        return Border(left=s,right=s,top=s,bottom=s)
    def hdr(ws,r,c,v,bg=NAVY):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name='Arial',bold=True,color=WHITE,size=10)
        cell.fill=PatternFill('solid',fgColor=bg)
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border=th()
    def dat(ws,r,c,v,fmt=None,bold=False,bg=None,left=False):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name='Arial',size=10,bold=bold)
        cell.alignment=Alignment(horizontal='left' if left else 'center',vertical='center',wrap_text=True)
        cell.border=th()
        if fmt: cell.number_format=fmt
        if bg: cell.fill=PatternFill('solid',fgColor=bg)
    def title(ws,text,cols):
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=cols)
        c=ws.cell(row=1,column=1,value=f"{text} — {period_label}")
        c.font=Font(name='Arial',bold=True,size=13,color=WHITE)
        c.fill=PatternFill('solid',fgColor=NAVY)
        c.alignment=Alignment(horizontal='center',vertical='center')
        ws.row_dimensions[1].height=32
        ws.row_dimensions[2].height=28

    wb=Workbook()
    ws1=wb.active; ws1.title='Нийт дүгнэлт'
    ws1.sheet_view.showGridLines=False
    title(ws1,'НИЙТ ДҮГНЭЛТ',6)
    for ci,h in enumerate(['№','Төхөөрөмжийн нэр','Mark','Нийт цаг','Calibration','Засвар'],1):
        hdr(ws1,2,ci,h)
    devices=conn.execute('SELECT d.*,dm.manufacturer,dm.model FROM devices d LEFT JOIN device_marks dm ON d.mark_id=dm.id ORDER BY d.name').fetchall()
    for ri,d in enumerate(devices,3):
        bg=WHITE if ri%2==0 else GRAY
        hrs=conn.execute(f'SELECT COALESCE(SUM(duration_hours),0) as t FROM usage_logs WHERE device_id=? AND {wd}',(d['id'],)).fetchone()['t']
        cal=conn.execute('SELECT calibration_date FROM calibrations WHERE device_id=? ORDER BY calibration_date DESC LIMIT 1',(d['id'],)).fetchone()
        orep=conn.execute('SELECT COUNT(*) as c FROM repairs WHERE device_id=? AND status=?',(d['id'],'new')).fetchone()['c']
        dat(ws1,ri,1,ri-2,bg=bg); dat(ws1,ri,2,d['name'],bg=bg,left=True)
        dat(ws1,ri,3,f"{d['manufacturer'] or ''} {d['model'] or ''}".strip(),bg=bg,left=True)
        dat(ws1,ri,4,round(hrs,2),fmt='0.00',bg=bg)
        dat(ws1,ri,5,cal['calibration_date'] if cal else '—',bg=bg)
        dat(ws1,ri,6,'Шинэ' if orep else '—',bg=bg)
    for ci,w in enumerate([5,30,22,14,16,14],1):
        ws1.column_dimensions[get_column_letter(ci)].width=w

    ws2=wb.create_sheet('Ашиглалтын бүртгэл')
    ws2.sheet_view.showGridLines=False
    title(ws2,'АШИГЛАЛТЫН БҮРТГЭЛ',7)
    for ci,h in enumerate(['№','Огноо','Төхөөрөмж','Ажилтан','Эхэлсэн','Гүйцэтгэсэн','Цаг'],1):
        hdr(ws2,2,ci,h,bg=TEAL)
    logs=conn.execute(f'''SELECT ul.*,d.name as dname,u.name as uname FROM usage_logs ul LEFT JOIN devices d ON d.id=ul.device_id LEFT JOIN users u ON u.id=ul.user_id WHERE {wu} ORDER BY ul.start_time''').fetchall()
    for ri,l in enumerate(logs,3):
        bg=WHITE if ri%2==0 else GRAY
        dat(ws2,ri,1,ri-2,bg=bg); dat(ws2,ri,2,l['start_time'][:10] if l['start_time'] else '',bg=bg)
        dat(ws2,ri,3,l['dname'] or '',bg=bg,left=True); dat(ws2,ri,4,l['uname'] or '',bg=bg,left=True)
        dat(ws2,ri,5,l['start_time'][11:16] if l['start_time'] else '',bg=bg)
        dat(ws2,ri,6,l['end_time'][11:16] if l['end_time'] else '—',bg=bg)
        dat(ws2,ri,7,round(l['duration_hours'],2) if l['duration_hours'] else 0,fmt='0.00',bg=bg)
    lr2=3+len(logs)
    ws2.merge_cells(start_row=lr2,start_column=1,end_row=lr2,end_column=6)
    dat(ws2,lr2,1,'НИЙТ ЦАГ',bold=True,bg='D6DCF0',left=True)
    dat(ws2,lr2,7,round(sum(l['duration_hours'] for l in logs if l['duration_hours']),2),fmt='0.00',bold=True,bg='D6DCF0')
    for ci,w in enumerate([5,14,28,20,12,12,12],1):
        ws2.column_dimensions[get_column_letter(ci)].width=w

    ws3=wb.create_sheet('Calibration бүртгэл')
    ws3.sheet_view.showGridLines=False
    title(ws3,'CALIBRATION БҮРТГЭЛ',6)
    for ci,h in enumerate(['№','Төхөөрөмж','Хийсэн огноо','Дараагийн огноо','Хийсэн ажилтан','Үр дүн'],1):
        hdr(ws3,2,ci,h,bg=PURPLE)
    cals=conn.execute(f'''SELECT c.*,d.name as dname,u.name as uname FROM calibrations c LEFT JOIN devices d ON d.id=c.device_id LEFT JOIN users u ON u.id=c.performed_by WHERE {wc} ORDER BY c.calibration_date''').fetchall()
    for ri,c in enumerate(cals,3):
        bg=WHITE if ri%2==0 else GRAY
        dat(ws3,ri,1,ri-2,bg=bg); dat(ws3,ri,2,c['dname'] or '',bg=bg,left=True)
        dat(ws3,ri,3,c['calibration_date'] or '',bg=bg); dat(ws3,ri,4,c['next_date'] or '—',bg=bg)
        dat(ws3,ri,5,c['uname'] or '—',bg=bg)
        cell=ws3.cell(row=ri,column=6,value='Тэнцсэн' if c['result']=='passed' else 'Тэнцээгүй')
        cell.font=Font(name='Arial',size=10,color='0F6E56' if c['result']=='passed' else '993C1D',bold=True)
        cell.fill=PatternFill('solid',fgColor=bg); cell.border=th()
        cell.alignment=Alignment(horizontal='center',vertical='center')
    for ci,w in enumerate([5,28,14,14,20,14],1):
        ws3.column_dimensions[get_column_letter(ci)].width=w

    ws4=wb.create_sheet('Засварын бүртгэл')
    ws4.sheet_view.showGridLines=False
    title(ws4,'ЗАСВАРЫН БҮРТГЭЛ',7)
    for ci,h in enumerate(['№','Төхөөрөмж','Огноо','Тайлбар','Компани','Зардал (₮)','Статус'],1):
        hdr(ws4,2,ci,h,bg=CORAL)
    reps=conn.execute(f'''SELECT r.*,d.name as dname FROM repairs r LEFT JOIN devices d ON d.id=r.device_id WHERE {wr} ORDER BY r.reported_date''').fetchall()
    for ri,r in enumerate(reps,3):
        bg=WHITE if ri%2==0 else GRAY
        dat(ws4,ri,1,ri-2,bg=bg); dat(ws4,ri,2,r['dname'] or '',bg=bg,left=True)
        dat(ws4,ri,3,r['reported_date'] or '',bg=bg); dat(ws4,ri,4,r['description'] or '',bg=bg,left=True)
        dat(ws4,ri,5,r['company'] or '—',bg=bg); dat(ws4,ri,6,r['cost'] or 0,fmt='#,##0',bg=bg)
        cell=ws4.cell(row=ri,column=7,value='Шинэ' if r['status']=='new' else 'Гүйцэтгэсэн')
        cell.font=Font(name='Arial',size=10,color='993C1D' if r['status']=='new' else '0F6E56',bold=True)
        cell.fill=PatternFill('solid',fgColor=bg); cell.border=th()
        cell.alignment=Alignment(horizontal='center',vertical='center')
    lr4=3+len(reps)
    ws4.merge_cells(start_row=lr4,start_column=1,end_row=lr4,end_column=5)
    dat(ws4,lr4,1,'НИЙТ ЗАРДАЛ',bold=True,bg='FAECE7',left=True)
    dat(ws4,lr4,6,sum(r['cost'] for r in reps if r['cost']),fmt='#,##0',bold=True,bg='FAECE7')
    dat(ws4,lr4,7,'',bg='FAECE7')
    for ci,w in enumerate([5,26,12,32,18,14,12],1):
        ws4.column_dimensions[get_column_letter(ci)].width=w

    ws5=wb.create_sheet('Дотоод шалгалт')
    ws5.sheet_view.showGridLines=False
    title(ws5,'ДОТООД ШАЛГАЛТЫН БҮРТГЭЛ',8)
    for ci,h in enumerate(['№','Огноо','Төхөөрөмж','Стандарт','Хэмжсэн','Зөрүү зөвшөөрөл','Үр дүн','Тохируулга'],1):
        hdr(ws5,2,ci,h,bg='2D6A4F')
    if date_filter_mode == 'range':
        wch = f"ch.check_date BETWEEN '{date_start}' AND '{date_end}'"
    else:
        wch = f"strftime('%Y-%m',ch.check_date) IN ({ym_list})"
    chks=conn.execute(f'''SELECT ch.*,d.name as dname,u.name as uname
        FROM device_checks ch
        LEFT JOIN devices d ON d.id=ch.device_id
        LEFT JOIN users u ON u.id=ch.checked_by
        WHERE {wch} ORDER BY ch.check_date,d.name''').fetchall()
    pass_c=fail_c=adj_c=0
    for ri,ch in enumerate(chks,3):
        bg=WHITE if ri%2==0 else GRAY
        is_pass=ch['result']=='pass'
        is_adj=ch['calibration_adjusted'] if ch['calibration_adjusted'] else 0
        if is_pass: pass_c+=1
        else: fail_c+=1
        if is_adj: adj_c+=1
        dat(ws5,ri,1,ri-2,bg=bg)
        dat(ws5,ri,2,ch['check_date'] or '',bg=bg)
        dat(ws5,ri,3,ch['dname'] or '',bg=bg,left=True)
        dat(ws5,ri,4,ch['standard_value'] or '—',bg=bg)
        dat(ws5,ri,5,ch['measured_value'] or '—',bg=bg)
        dat(ws5,ri,6,ch['tolerance'] or '—',bg=bg)
        cell=ws5.cell(row=ri,column=7,value='Тэнцсэн' if is_pass else 'Тэнцээгүй')
        cell.font=Font(name='Arial',size=10,color='0F6E56' if is_pass else '993C1D',bold=True)
        cell.fill=PatternFill('solid',fgColor=bg); cell.border=th()
        cell.alignment=Alignment(horizontal='center',vertical='center')
        cell2=ws5.cell(row=ri,column=8,value='Тийм' if is_adj else '')
        cell2.font=Font(name='Arial',size=10,color='D97706' if is_adj else '999999',bold=bool(is_adj))
        cell2.fill=PatternFill('solid',fgColor=bg); cell2.border=th()
        cell2.alignment=Alignment(horizontal='center',vertical='center')
    # Дүгнэлт мөр
    lr5=3+len(chks)
    total=len(chks)
    ws5.merge_cells(start_row=lr5,start_column=1,end_row=lr5,end_column=3)
    dat(ws5,lr5,1,f'Нийт: {total}  |  Тэнцсэн: {pass_c}  |  Тэнцээгүй: {fail_c}  |  Тохируулга: {adj_c}',bold=True,bg='E8F5E9',left=True)
    for ci in range(4,9):
        dat(ws5,lr5,ci,'',bg='E8F5E9')
    for ci,w in enumerate([5,14,28,14,14,16,14,12],1):
        ws5.column_dimensions[get_column_letter(ci)].width=w

    SAMPLE_TYPE_MN = {
        'PIT':'Уурхай (PIT)', 'STOCKPILE':'Овоолго (Stockpile)',
        'EXPORT':'Экспорт (Export)', 'CONTROL':'Гааль (Control)',
        'EQ_CONTROL':'Гадаад хяналт (EQ Control)', 'DP':'Баяжуулах (DP)',
    }
    STATUS_MN = {'pending':'Хүлээгдэж байна','received':'Хүлээн авсан',
                 'prepared':'Бэлтгэсэн','analysing':'Шинжилж байна','done':'Дууссан'}

    ws_samples=conn.execute(f'''
        SELECT sr.lab_number, g.sample_name, g.sample_type, sr.received_date,
               sr.mass_kg, g.quantity, g.status
        FROM sample_receipt sr
        JOIN geo_samples g ON g.id=sr.geo_sample_id
        WHERE {ws_filter}
        ORDER BY sr.received_date, sr.lab_number
    ''').fetchall()

    type_summary=conn.execute(f'''
        SELECT g.sample_type,
               COUNT(*) as cnt,
               COALESCE(SUM(sr.mass_kg),0) as total_kg
        FROM sample_receipt sr
        JOIN geo_samples g ON g.id=sr.geo_sample_id
        WHERE {ws_filter}
        GROUP BY g.sample_type
        ORDER BY cnt DESC
    ''').fetchall()

    # ── Sheet 6: Дээжний төрөл (нэгтгэсэн) ────────────────
    ws6=wb.create_sheet('Дээжний төрөл')
    ws6.sheet_view.showGridLines=False
    title(ws6,'ДЭЭЖНИЙ ТӨРЛӨӨР НЭГТГЭСЭН',3)
    for ci,h in enumerate(['Дээжний төрөл','Шинжилгээний тоо','Нийт жин (кг)'],1):
        hdr(ws6,2,ci,h,bg=TEAL)
    for ri,t in enumerate(type_summary,3):
        bg=WHITE if ri%2==0 else GRAY
        dat(ws6,ri,1,SAMPLE_TYPE_MN.get(t['sample_type'], t['sample_type'] or ''),bg=bg,left=True)
        dat(ws6,ri,2,t['cnt'],bg=bg)
        dat(ws6,ri,3,round(t['total_kg'],2),fmt='0.00',bg=bg)
    tr=3+len(type_summary)
    dat(ws6,tr,1,'НИЙТ',bold=True,bg='D6F0E8',left=True)
    dat(ws6,tr,2,sum(t['cnt'] for t in type_summary),bold=True,bg='D6F0E8')
    dat(ws6,tr,3,round(sum(t['total_kg'] for t in type_summary),2),fmt='0.00',bold=True,bg='D6F0E8')
    for ci,w in enumerate([28,20,18],1):
        ws6.column_dimensions[get_column_letter(ci)].width=w

    # ── Sheet 7: Шинжилгээний үр дүн (төрлөөр нэгтгэсэн) ─
    ws7=wb.create_sheet('Шинжилгээ')
    ws7.sheet_view.showGridLines=False
    title(ws7,'ШИНЖИЛГЭЭНИЙ ҮР ДҮНГИЙН ХҮСНЭГТ',11)
    hdrs7=['№','Дээжний төрөл','Нийт дээж','Mt','Mad','Aad','Vad','Fc','St','Q','G']
    for ci,h in enumerate(hdrs7,1):
        hdr(ws7,2,ci,h,bg=TEAL)
    ws7.row_dimensions[2].height=22

    def has_mt(row):
        try: return row['ff_sample'] is not None and float(row['ff_sample'])>0 and row['ff_dried'] is not None
        except: return False
    def has_g(row):
        if row['g_val'] is not None: return True
        try: return all(row[f] is not None for f in ['g_coke','g_tare','g_sieve1','g_sieve2'])
        except: return False

    type_receipts = conn.execute(f'''
        SELECT g.sample_type, COUNT(DISTINCT sr.id) as cnt
        FROM sample_receipt sr
        JOIN geo_samples g ON g.id=sr.geo_sample_id
        WHERE {ws_filter}
        GROUP BY g.sample_type
        ORDER BY cnt DESC
    ''').fetchall()

    analysis_rows=conn.execute(f'''
        SELECT g.sample_type, se.mad, se.aad, se.vad, se.fc,
               se.sulfur, se.cal_value, se.g_val,
               se.ff_sample, se.ff_dried,
               se.g_coke, se.g_tare, se.g_sieve1, se.g_sieve2
        FROM sample_receipt sr
        JOIN geo_samples g ON g.id=sr.geo_sample_id
        JOIN sample_entries se ON se.receipt_id=sr.id
        WHERE {ws_filter}
          AND se.row_status IN ('done','approved')
    ''').fetchall()

    from collections import defaultdict
    type_cnt = defaultdict(lambda: {k:0 for k in ['mt','mad','aad','vad','fc','sulfur','cal_value','g']})
    for s in analysis_rows:
        t = s['sample_type']
        if has_mt(s): type_cnt[t]['mt'] += 1
        for f in ['mad','aad','vad','fc','sulfur','cal_value']:
            if s[f] is not None: type_cnt[t][f] += 1
        if has_g(s): type_cnt[t]['g'] += 1

    for ri, r in enumerate(type_receipts, 3):
        bg = WHITE if ri%2==0 else GRAY
        t = r['sample_type']
        c = type_cnt[t]
        dat(ws7,ri,1,ri-2,bg=bg)
        dat(ws7,ri,2,SAMPLE_TYPE_MN.get(t,t),bg=bg,left=True)
        dat(ws7,ri,3,r['cnt'],bg=bg)
        dat(ws7,ri,4,c['mt'] or None,bg=bg)
        dat(ws7,ri,5,c['mad'] or None,bg=bg)
        dat(ws7,ri,6,c['aad'] or None,bg=bg)
        dat(ws7,ri,7,c['vad'] or None,bg=bg)
        dat(ws7,ri,8,c['fc'] or None,bg=bg)
        dat(ws7,ri,9,c['sulfur'] or None,bg=bg)
        dat(ws7,ri,10,c['cal_value'] or None,bg=bg)
        dat(ws7,ri,11,c['g'] or None,bg=bg)

    # Нийт нийлбэр мөр
    total_row = 3 + len(type_receipts)
    all_c = defaultdict(int)
    for t_data in type_cnt.values():
        for k,v in t_data.items():
            all_c[k] += v
    dat(ws7,total_row,1,'',bg='D6F0E8')
    dat(ws7,total_row,2,'НИЙТ',bold=True,bg='D6F0E8',left=True)
    dat(ws7,total_row,3,sum(r['cnt'] for r in type_receipts),bold=True,bg='D6F0E8')
    dat(ws7,total_row,4,all_c['mt'] or None,bold=True,bg='D6F0E8')
    dat(ws7,total_row,5,all_c['mad'] or None,bold=True,bg='D6F0E8')
    dat(ws7,total_row,6,all_c['aad'] or None,bold=True,bg='D6F0E8')
    dat(ws7,total_row,7,all_c['vad'] or None,bold=True,bg='D6F0E8')
    dat(ws7,total_row,8,all_c['fc'] or None,bold=True,bg='D6F0E8')
    dat(ws7,total_row,9,all_c['sulfur'] or None,bold=True,bg='D6F0E8')
    dat(ws7,total_row,10,all_c['cal_value'] or None,bold=True,bg='D6F0E8')
    dat(ws7,total_row,11,all_c['g'] or None,bold=True,bg='D6F0E8')

    for ci,w in enumerate([5,24,14,10,10,10,10,10,10,10,10],1):
        ws7.column_dimensions[get_column_letter(ci)].width=w

    for ws in [ws1,ws2,ws3,ws4,ws5,ws6,ws7]:
        ws.page_setup.orientation='landscape'
        ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1

    conn.close()
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname='Тоног_төхөөрөмж_' + period_label.replace(' ','_') + '.xlsx'
    return send_file(buf,as_attachment=True,download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── ЛАБОРАТОРИЙН ТАЙЛАН (Excel) ─────────────────────────
SAMPLE_TYPE_LONG = {
    'PIT': 'Уурхай (PIT)', 'STOCKPILE': 'Овоолго (Stockpile)',
    'EXPORT': 'Ачилт (Export)', 'CONTROL': 'Хяналт (Control)',
    'EQ_CONTROL': 'Гадаад хяналт (EQ Control)', 'DP': 'Баяжуулах (DP)',
    'CRM': 'Стандарт материал (CRM)',
}
STATUS_LONG = {'pending': 'Хүлээгдэж байна', 'received': 'Хүлээн авсан',
               'prepared': 'Бэлтгэсэн', 'analysing': 'Шинжилж байна',
               'done': 'Дууссан'}
ROLE_MN = {'admin': 'Админ', 'senior': 'Ахлах', 'staff': 'Химич',
           'preparer': 'Дээж бэлтгэгч', 'geologist': 'Геологич',
           'bayjuulach': 'Баяжуулах цех', 'guest': 'Зочин'}
# Дэлгэрэнгүй хуудсанд гарах үзүүлэлтүүд: (гарчиг, талбар, формат)
LAB_RESULT_COLS = [
    ('Mt, %',      'mt_result', '0.00'),
    ('Mad, %',     'mad',       '0.00'),
    ('Aad, %',     'aad',       '0.00'),
    ('Vad, %',     'vad',       '0.00'),
    ('FCad, %',    'fc',        '0.00'),
    ('Sad, %',     'sulfur',    '0.00'),
    ('Qad, ккал',  'cal_kcal',  '0'),
    ('GR.I',       'g_index',   '0'),
    ('FSI',        'fsi',       '0.0'),
]


def lab_g_index(e):
    """G индекс — гараар оруулсан утга, эсвэл жингээр бодно"""
    try:
        if e.get('g_val') is not None:
            return float(e['g_val'])
        gc, gt = e.get('g_coke'), e.get('g_tare')
        s1, s2 = e.get('g_sieve1'), e.get('g_sieve2')
        if None not in (gc, gt, s1, s2):
            d = float(gc) - float(gt)
            if d > 0:
                return 10 + (30 * (float(s1) - float(gt)) + 70 * (float(s2) - float(gt))) / d
    except (TypeError, ValueError):
        pass
    return None


def lab_report_rows(conn, d0s, d1s):
    """Хугацаанд хүлээн авсан дээж бүрийн ЭЦСИЙН үр дүн.

    Давталттай мөрүүдээс хамгийн ойрхон хоёрын дундажийг сонгоно — үр дүнгийн
    хуудас болон албан тайлантай ижил дүрэм.
    """
    receipts = conn.execute(
        '''SELECT sr.id, sr.lab_number, sr.received_date, sr.mass_kg,
                  g.sample_name, g.sample_type, g.status, g.quantity
           FROM sample_receipt sr JOIN geo_samples g ON g.id=sr.geo_sample_id
           WHERE sr.received_date BETWEEN ? AND ?
           ORDER BY sr.received_date, sr.lab_serial''', (d0s, d1s)).fetchall()
    tols = qc_tolerances(conn)
    out = []
    for r in receipts:
        entries = [dict(e) for e in conn.execute(
            '''SELECT * FROM sample_entries WHERE receipt_id=?
               ORDER BY row_num, is_duplicate''', (r['id'],)).fetchall()]
        for e in entries:
            e['mt_result'] = total_moisture(e)
        apply_final_results(entries, tols)
        for e in entries:
            if e.get('is_duplicate') != 0:
                continue           # зэрэгцээ/давталт нь QC-д, тайланд үндсэн мөр гарна
            e['g_index'] = lab_g_index(e)
            e['cal_kcal'] = (e['cal_value'] / 4.1868) if e.get('cal_value') else None
            e['_receipt'] = r
            out.append(e)
    return out


@app.route('/reports/export/lab')
@perm_required('can_report')
def lab_report_export():
    """Лабораторийн тайлан — дээж, шинжилгээ, ажилтны гүйцэтгэл (Excel).

    Тоног төхөөрөмжийн тайлан (/reports/export) -аас тусдаа файл.
    """
    if session.get('role') == 'guest':
        flash('Зочин горимд Excel татах боломжгүй.', 'error')
        return redirect(url_for('reports'))
    rtype = request.args.get('type', 'month')
    year  = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    week  = int(request.args.get('week', datetime.now().isocalendar()[1]))
    half  = int(request.args.get('half', 1))
    d0s, d1s, period_label = lab_period_range(rtype, year, month, week, half)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    NAVY='1A2744'; TEAL='0F6E56'; PURPLE='3C3489'; AMBER='8A5A00'
    WHITE='FFFFFF'; GRAY='F7F7F5'; SUM_BG='D6F0E8'

    def th():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)
    def hdr(ws, r, c, v, bg=NAVY):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name='Arial', bold=True, color=WHITE, size=10)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = th()
    def dat(ws, r, c, v, fmt=None, bold=False, bg=None, left=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name='Arial', size=10, bold=bold)
        cell.alignment = Alignment(horizontal='left' if left else 'center', vertical='center')
        cell.border = th()
        if fmt: cell.number_format = fmt
        if bg: cell.fill = PatternFill('solid', fgColor=bg)
    def title(ws, text, cols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        c = ws.cell(row=1, column=1, value=f"{text} — {period_label}")
        c.font = Font(name='Arial', bold=True, size=13, color=WHITE)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 32
        ws.row_dimensions[2].height = 26
    def widths(ws, ws_widths):
        for ci, w in enumerate(ws_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    conn = get_db()
    rows = lab_report_rows(conn, d0s, d1s)
    SW = "sr.received_date BETWEEN ? AND ?"          # дээж хүлээн авсан хугацаа
    AW = "substr(se.done_at,1,10) BETWEEN ? AND ?"   # шинжилгээ хийсэн хугацаа
    P = (d0s, d1s)

    def one(sql, args=P):
        return conn.execute(sql, args).fetchone()[0]

    wb = Workbook()

    # ── Хуудас 1: Нийт дүгнэлт ───────────────────────────
    ws1 = wb.active; ws1.title = 'Нийт дүгнэлт'
    ws1.sheet_view.showGridLines = False
    title(ws1, 'ЛАБОРАТОРИЙН ТАЙЛАН', 3)
    n_jobs = one(f'''SELECT COUNT(*) FROM sample_receipt sr WHERE {SW}''')
    n_samples = one(f'''SELECT COALESCE(SUM(COALESCE(g.quantity,1)),0)
        FROM sample_receipt sr JOIN geo_samples g ON g.id=sr.geo_sample_id WHERE {SW}''')
    n_done = one(f'''SELECT COALESCE(SUM(COALESCE(g.quantity,1)),0)
        FROM sample_receipt sr JOIN geo_samples g ON g.id=sr.geo_sample_id
        WHERE {SW} AND g.status='done' ''')
    n_mass = one(f'''SELECT COALESCE(SUM(sr.mass_kg),0) FROM sample_receipt sr WHERE {SW}''')
    n_analysis = sum(one(analysis_count_sql(f)) for f, _ in ANALYSIS_FIELDS)
    n_rows_done = one(f'''SELECT COUNT(*) FROM sample_entries se
        WHERE se.is_duplicate=0 AND se.row_status IN ('done','approved') AND {AW}''')
    n_rows_appr = one(f'''SELECT COUNT(*) FROM sample_entries se
        WHERE se.is_duplicate=0 AND se.row_status='approved' AND {AW}''')
    # Зэрэгцээ/давталтын мөрүүд өөрсдөө done_at-гүй байж болно тул дээжийн
    # хүлээн авсан хугацаагаар шүүнэ
    n_parallel = one(f'''SELECT COUNT(DISTINCT se.receipt_id||'-'||se.row_num)
        FROM sample_entries se JOIN sample_receipt sr ON sr.id=se.receipt_id
        WHERE se.is_duplicate=1 AND se.row_status!='empty' AND {SW}''')
    n_repeat = one(f'''SELECT COUNT(DISTINCT se.receipt_id||'-'||se.row_num)
        FROM sample_entries se JOIN sample_receipt sr ON sr.id=se.receipt_id
        WHERE se.is_duplicate>=2 AND {SW}''')
    _gen = get_user(session.get('user_id', 0))
    _gen_by_name = (_gen['name'] if _gen else None) or '—'
    summary = [
        ('Хугацаа', period_label, None),
        ('Тайлан гаргасан', datetime.now().strftime('%Y-%m-%d %H:%M'), None),
        ('Гаргасан ажилтан', _gen_by_name, None),
        None,
        ('Хүлээн авсан ажил (бүртгэл)', n_jobs, '0'),
        ('Нийт дээж', n_samples, '0'),
        ('Дууссан дээж', n_done, '0'),
        ('Хүлээгдэж байгаа дээж', max(n_samples - n_done, 0), '0'),
        ('Нийт масс, кг', round(n_mass, 2), '0.00'),
        None,
        ('Хийсэн шинжилгээ (үзүүлэлтээр)', n_analysis, '0'),
        ('Шинжилж дуусгасан мөр', n_rows_done, '0'),
        ('Баталгаажуулсан мөр', n_rows_appr, '0'),
        ('Зэрэгцээ шинжилгээтэй дээж', n_parallel, '0'),
        ('Давталт хийсэн дээж', n_repeat, '0'),
    ]
    hdr(ws1, 2, 1, 'Үзүүлэлт'); hdr(ws1, 2, 2, 'Хэмжигдэхүүн'); hdr(ws1, 2, 3, '')
    r = 3
    for item in summary:
        if item is None:
            r += 1
            continue
        label, value, fmt = item
        bg = WHITE if r % 2 == 0 else GRAY
        dat(ws1, r, 1, label, bg=bg, left=True)
        dat(ws1, r, 2, value, fmt=fmt, bg=bg, bold=True)
        dat(ws1, r, 3, '', bg=bg)
        r += 1
    widths(ws1, [34, 20, 4])

    # ── Хуудас 2: Дээжийн төрлөөр ────────────────────────
    ws2 = wb.create_sheet('Дээжийн төрөл')
    ws2.sheet_view.showGridLines = False
    title(ws2, 'ДЭЭЖИЙН ТӨРЛӨӨР НЭГТГЭЛ', 6)
    for ci, h in enumerate(['№','Дээжийн төрөл','Ажил','Дээж','Дууссан дээж','Масс, кг'], 1):
        hdr(ws2, 2, ci, h, bg=TEAL)
    t_rows = conn.execute(f'''
        SELECT g.sample_type, COUNT(*) as jobs,
               COALESCE(SUM(COALESCE(g.quantity,1)),0) as samples,
               COALESCE(SUM(CASE WHEN g.status='done' THEN COALESCE(g.quantity,1) END),0) as done_n,
               COALESCE(SUM(sr.mass_kg),0) as kg
        FROM sample_receipt sr JOIN geo_samples g ON g.id=sr.geo_sample_id
        WHERE {SW} GROUP BY g.sample_type ORDER BY samples DESC''', P).fetchall()
    for ri, t in enumerate(t_rows, 3):
        bg = WHITE if ri % 2 == 0 else GRAY
        dat(ws2, ri, 1, ri - 2, bg=bg)
        dat(ws2, ri, 2, SAMPLE_TYPE_LONG.get(t['sample_type'], t['sample_type'] or '—'), bg=bg, left=True)
        dat(ws2, ri, 3, t['jobs'], bg=bg)
        dat(ws2, ri, 4, t['samples'], bg=bg)
        dat(ws2, ri, 5, t['done_n'], bg=bg)
        dat(ws2, ri, 6, round(t['kg'], 2), fmt='0.00', bg=bg)
    tr = 3 + len(t_rows)
    dat(ws2, tr, 1, '', bg=SUM_BG); dat(ws2, tr, 2, 'НИЙТ', bold=True, bg=SUM_BG, left=True)
    dat(ws2, tr, 3, sum(t['jobs'] for t in t_rows), bold=True, bg=SUM_BG)
    dat(ws2, tr, 4, sum(t['samples'] for t in t_rows), bold=True, bg=SUM_BG)
    dat(ws2, tr, 5, sum(t['done_n'] for t in t_rows), bold=True, bg=SUM_BG)
    dat(ws2, tr, 6, round(sum(t['kg'] for t in t_rows), 2), fmt='0.00', bold=True, bg=SUM_BG)
    widths(ws2, [5, 30, 12, 12, 16, 14])

    # ── Хуудас 3: Шинжилгээний төрлөөр ───────────────────
    ws3 = wb.create_sheet('Шинжилгээний төрөл')
    ws3.sheet_view.showGridLines = False
    title(ws3, 'ШИНЖИЛГЭЭНИЙ ТӨРЛӨӨР НЭГТГЭЛ', 6)
    for ci, h in enumerate(['№','Үзүүлэлт','Хийсэн тоо','Дундаж','Хамгийн бага','Хамгийн их'], 1):
        hdr(ws3, 2, ci, h, bg=PURPLE)
    # Тоо — done_at-аар (хэдэн шинжилгээ хийсэн), дундаж/муж — эцсийн үр дүнгээр
    STAT_FIELD = {'mt_dried': 'mt_result', 'g_coke': 'g_index'}
    for ri, (field, name) in enumerate(ANALYSIS_FIELDS, 3):
        bg = WHITE if ri % 2 == 0 else GRAY
        cnt = one(analysis_count_sql(field))
        vals = [v for v in (e.get(STAT_FIELD.get(field, field)) for e in rows) if v is not None]
        fmt = '0' if field in ('cal_value',) else '0.00'
        dat(ws3, ri, 1, ri - 2, bg=bg)
        dat(ws3, ri, 2, name, bg=bg, left=True)
        dat(ws3, ri, 3, cnt, bg=bg)
        # Утга бүрэн нарийвчлалтай хадгалагдаж, харагдац нь форматаар зохицуулагдана
        dat(ws3, ri, 4, round(sum(vals) / len(vals), 4) if vals else '—', fmt=fmt if vals else None, bg=bg)
        dat(ws3, ri, 5, round(min(vals), 4) if vals else '—', fmt=fmt if vals else None, bg=bg)
        dat(ws3, ri, 6, round(max(vals), 4) if vals else '—', fmt=fmt if vals else None, bg=bg)
    lr3 = 3 + len(ANALYSIS_FIELDS)
    dat(ws3, lr3, 1, '', bg=SUM_BG); dat(ws3, lr3, 2, 'НИЙТ ШИНЖИЛГЭЭ', bold=True, bg=SUM_BG, left=True)
    dat(ws3, lr3, 3, n_analysis, bold=True, bg=SUM_BG)
    for ci in (4, 5, 6):
        dat(ws3, lr3, ci, '', bg=SUM_BG)
    widths(ws3, [5, 28, 14, 14, 16, 14])

    # ── Хуудас 4: Ажилтны гүйцэтгэл ──────────────────────
    ws4 = wb.create_sheet('Ажилтны гүйцэтгэл')
    ws4.sheet_view.showGridLines = False
    title(ws4, 'АЖИЛТНЫ ГҮЙЦЭТГЭЛ (ДЭЭЖИЙН ТООГООР)', 7)
    for ci, h in enumerate(['№','Ажилтан','Албан тушаал','Бүртгэсэн','Бэлтгэсэн',
                            'Шинжилсэн','Баталсан'], 1):
        hdr(ws4, 2, ci, h, bg=AMBER)
    staff = conn.execute(
        "SELECT id, name, position, role FROM users ORDER BY name").fetchall()
    perf = []
    for u in staff:
        reg = one(f'''SELECT COALESCE(SUM(COALESCE(g.quantity,1)),0)
            FROM sample_receipt sr JOIN geo_samples g ON g.id=sr.geo_sample_id
            WHERE {SW} AND g.registered_by=?''', P + (u['id'],))
        prep = one(f'''SELECT COALESCE(SUM(COALESCE(g.quantity,1)),0)
            FROM sample_receipt sr JOIN geo_samples g ON g.id=sr.geo_sample_id
            WHERE {SW} AND sr.prep_done_at IS NOT NULL
              AND (sr.prep_by=? OR (sr.prep_by IS NULL AND sr.prep_operator=?))''',
            P + (u['id'], u['name']))
        done = one(f'''SELECT COUNT(*) FROM sample_entries se
            WHERE se.done_by=? AND se.row_status IN ('done','approved') AND {AW}''',
            (u['id'],) + P)
        appr = one(f'''SELECT COUNT(*) FROM sample_entries se
            WHERE se.approved_by=? AND se.row_status='approved' AND {AW}''',
            (u['id'],) + P)
        if reg or prep or done or appr:
            perf.append((u, reg, prep, done, appr))
    perf.sort(key=lambda x: -(x[1] + x[2] + x[3]))
    for ri, (u, reg, prep, done, appr) in enumerate(perf, 3):
        bg = WHITE if ri % 2 == 0 else GRAY
        dat(ws4, ri, 1, ri - 2, bg=bg)
        dat(ws4, ri, 2, u['name'] or '—', bg=bg, left=True)
        dat(ws4, ri, 3, u['position'] or ROLE_MN.get(u['role'], u['role'] or '—'), bg=bg, left=True)
        dat(ws4, ri, 4, reg, bg=bg); dat(ws4, ri, 5, prep, bg=bg)
        dat(ws4, ri, 6, done, bg=bg); dat(ws4, ri, 7, appr, bg=bg)
    if not perf:
        ws4.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
        dat(ws4, 3, 1, 'Энэ хугацаанд бүртгэл байхгүй', bg=GRAY, left=True)
    else:
        lr4 = 3 + len(perf)
        dat(ws4, lr4, 1, '', bg=SUM_BG)
        ws4.merge_cells(start_row=lr4, start_column=2, end_row=lr4, end_column=3)
        dat(ws4, lr4, 2, 'НИЙТ', bold=True, bg=SUM_BG, left=True)
        for ci, idx in ((4, 1), (5, 2), (6, 3), (7, 4)):
            dat(ws4, lr4, ci, sum(p[idx] for p in perf), bold=True, bg=SUM_BG)
    widths(ws4, [5, 24, 24, 13, 13, 13, 13])

    # ── Хуудас 5: Хугацааны динамик ──────────────────────
    ws5 = wb.create_sheet('Динамик')
    ws5.sheet_view.showGridLines = False
    by_day = rtype in ('week', 'month')
    title(ws5, 'ӨДӨР ТУТМЫН ДИНАМИК' if by_day else 'САР ТУТМЫН ДИНАМИК', 4)
    for ci, h in enumerate(['№', 'Өдөр' if by_day else 'Сар', 'Хүлээн авсан дээж',
                            'Хийсэн шинжилгээ'], 1):
        hdr(ws5, 2, ci, h, bg=TEAL)
    cut = 10 if by_day else 7        # 'YYYY-MM-DD' эсвэл 'YYYY-MM'
    smp = dict(conn.execute(f'''
        SELECT substr(sr.received_date,1,{cut}) as p,
               COALESCE(SUM(COALESCE(g.quantity,1)),0)
        FROM sample_receipt sr JOIN geo_samples g ON g.id=sr.geo_sample_id
        WHERE {SW} GROUP BY p''', P).fetchall())
    ana = dict(conn.execute(f'''
        SELECT substr(se.done_at,1,{cut}) as p, COUNT(*)
        FROM sample_entries se
        WHERE se.is_duplicate=0 AND se.row_status IN ('done','approved') AND {AW}
        GROUP BY p''', P).fetchall())
    buckets = sorted(set(smp) | set(ana))
    for ri, b in enumerate(buckets, 3):
        bg = WHITE if ri % 2 == 0 else GRAY
        dat(ws5, ri, 1, ri - 2, bg=bg); dat(ws5, ri, 2, b, bg=bg)
        dat(ws5, ri, 3, smp.get(b, 0), bg=bg); dat(ws5, ri, 4, ana.get(b, 0), bg=bg)
    if buckets:
        lr5 = 3 + len(buckets)
        dat(ws5, lr5, 1, '', bg=SUM_BG); dat(ws5, lr5, 2, 'НИЙТ', bold=True, bg=SUM_BG, left=True)
        dat(ws5, lr5, 3, sum(smp.values()), bold=True, bg=SUM_BG)
        dat(ws5, lr5, 4, sum(ana.values()), bold=True, bg=SUM_BG)
    else:
        ws5.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
        dat(ws5, 3, 1, 'Энэ хугацаанд бүртгэл байхгүй', bg=GRAY, left=True)
    widths(ws5, [5, 16, 20, 20])

    # ── Хуудас 6: Дээжийн дэлгэрэнгүй ────────────────────
    ws6 = wb.create_sheet('Дээжийн дэлгэрэнгүй')
    ws6.sheet_view.showGridLines = False
    heads6 = ['№','Лаб дугаар','Дээжийн нэр','Төрөл','Хүлээн авсан','Статус'] + \
             [h for h, _, _ in LAB_RESULT_COLS]
    title(ws6, 'ДЭЭЖ БҮРИЙН ҮР ДҮН', len(heads6))
    for ci, h in enumerate(heads6, 1):
        hdr(ws6, 2, ci, h)
    for ri, e in enumerate(rows, 3):
        bg = WHITE if ri % 2 == 0 else GRAY
        rc = e['_receipt']
        dat(ws6, ri, 1, ri - 2, bg=bg)
        dat(ws6, ri, 2, rc['lab_number'] or '—', bg=bg)
        dat(ws6, ri, 3, e.get('sample_name') or rc['sample_name'] or '—', bg=bg, left=True)
        dat(ws6, ri, 4, SAMPLE_TYPE_LONG.get(rc['sample_type'], rc['sample_type'] or '—'), bg=bg, left=True)
        dat(ws6, ri, 5, (rc['received_date'] or '')[:10], bg=bg)
        dat(ws6, ri, 6, STATUS_LONG.get(rc['status'], rc['status'] or '—'), bg=bg)
        for k, (h, field, fmt) in enumerate(LAB_RESULT_COLS):
            v = e.get(field)
            dat(ws6, ri, 7 + k, round(v, 4) if isinstance(v, (int, float)) else None,
                fmt=fmt, bg=bg)
    if not rows:
        ws6.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(heads6))
        dat(ws6, 3, 1, 'Энэ хугацаанд дээж хүлээн авсан бүртгэл байхгүй', bg=GRAY, left=True)
    widths(ws6, [5, 18, 22, 24, 14, 16] + [10] * len(LAB_RESULT_COLS))

    for ws in wb.worksheets:
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.freeze_panes = 'A3'

    # Гаргасан тайланг бүртгэнэ — Тайлан хуудсын жагсаалтад харагдана
    period_value = {'month': month, 'week': week, 'half': half}.get(rtype)
    try:
        conn.execute('''INSERT INTO lab_report_records
            (period_type, year, period_value, period_label, generated_by)
            VALUES(?,?,?,?,?)''',
            (rtype, year, period_value, period_label, session.get('user_id')))
        conn.commit()
    except Exception:
        app.logger.exception('lab_report_records бүртгэж чадсангүй')
    conn.close()

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = 'Лабораторийн_тайлан_' + period_label.split(' (')[0].replace(' ', '_') + '.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── ANALYSIS MODULE ──────────────────────────────────────

# Лаб дугаар автоматаар үүсгэх
def generate_lab_number(sample_type, date_str):
    """Дээжний төрлөөс хамааран лаб дугаар үүсгэнэ"""
    prefix_map = {
        'PIT': 1000, 'STOCKPILE': 2000, 'EXPORT': 3000,
        'CONTROL': 4000, 'EQ_CONTROL': 5000, 'DP': 6000, 'CRM': 9000
    }
    base = prefix_map.get(sample_type, 1000)
    conn = get_db()
    # Тухайн өдрийн хамгийн сүүлийн дугаар олох
    last = conn.execute("""
        SELECT lab_serial FROM sample_receipt
        WHERE lab_number LIKE ?
        ORDER BY lab_serial DESC LIMIT 1
    """, (f"%-{date_str}",)).fetchone()
    conn.close()
    if last:
        next_serial = last['lab_serial'] + 1
    else:
        next_serial = base + 1
    return f"{next_serial}-{date_str}", next_serial

# Тооцоолол функц
def calculate_results(m):
    """Sheet 2-ын томьёогоор тооцоолно"""
    r = {}
    try:
        # Нийт чийг Mt
        fm = 0
        if m.get('fm_sample_mass') and m.get('fm_dried_mass'):
            fm = (m['fm_sample_mass'] - m['fm_dried_mass']) / m['fm_sample_mass'] * 100

        im1 = im2 = 0
        if m.get('mt_tare1') and m.get('mt_sample1') and m.get('mt_dried1'):
            im1 = (m['mt_tare1'] + m['mt_sample1'] - m['mt_dried1']) / m['mt_sample1'] * 100
        if m.get('mt_tare2') and m.get('mt_sample2') and m.get('mt_dried2'):
            im2 = (m['mt_tare2'] + m['mt_sample2'] - m['mt_dried2']) / m['mt_sample2'] * 100
        im_avg = (im1 + im2) / 2 if im2 else im1
        r['Mt'] = round(fm + im_avg * (1 - fm/100), 4) if fm else round(im_avg, 4)

        # Дотоод чийг Mad
        mad1 = mad2 = 0
        if m.get('im_tare1') and m.get('im_sample1') and m.get('im_dried1'):
            mad1 = (m['im_tare1'] + m['im_sample1'] - m['im_dried1']) / m['im_sample1'] * 100
        if m.get('im_tare2') and m.get('im_sample2') and m.get('im_dried2'):
            mad2 = (m['im_tare2'] + m['im_sample2'] - m['im_dried2']) / m['im_sample2'] * 100
        r['Mad'] = round((mad1 + mad2) / 2 if mad2 else mad1, 4)

        # Үнс Aad
        a1 = a2 = 0
        if m.get('ash_tare1') and m.get('ash_sample1') and m.get('ash_burned1'):
            a1 = (m['ash_burned1'] - m['ash_tare1']) / m['ash_sample1'] * 100
        if m.get('ash_tare2') and m.get('ash_sample2') and m.get('ash_burned2'):
            a2 = (m['ash_burned2'] - m['ash_tare2']) / m['ash_sample2'] * 100
        r['Aad'] = round((a1 + a2) / 2 if a2 else a1, 4)
        r['Adb'] = round(r['Aad'] * 100 / (100 - r['Mad']), 4) if r.get('Mad') else 0

        # Дэгдэмхий Vad
        v1 = v2 = 0
        if m.get('vol_tare1') and m.get('vol_sample1') and m.get('vol_burned1'):
            v1 = (m['vol_tare1'] + m['vol_sample1'] - m['vol_burned1']) / m['vol_sample1'] * 100 - r.get('Mad', 0)
        if m.get('vol_tare2') and m.get('vol_sample2') and m.get('vol_burned2'):
            v2 = (m['vol_tare2'] + m['vol_sample2'] - m['vol_burned2']) / m['vol_sample2'] * 100 - r.get('Mad', 0)
        r['Vad'] = round((v1 + v2) / 2 if v2 else v1, 4)
        r['Vdb'] = round(r['Vad'] * 100 / (100 - r['Mad']), 4) if r.get('Mad') else 0
        r['Vdaf'] = round(r['Vad'] * 100 / (100 - r['Mad'] - r['Aad']), 4) if r.get('Mad') and r.get('Aad') else 0
        r['FCad'] = round(100 - r['Mad'] - r['Aad'] - r['Vad'], 4)

        # Хүхэр Stad
        s1 = m.get('sulfur1', 0) or 0
        s2 = m.get('sulfur2', 0) or 0
        r['Stad'] = round((s1 + s2) / 2 if s2 else s1, 4)
        r['Std'] = round(r['Stad'] * 100 / (100 - r['Mad']), 4) if r.get('Mad') else 0

        # Илчлэг
        q1 = m.get('cal_value1', 0) or 0
        q2 = m.get('cal_value2', 0) or 0
        qb_jg = (q1 + q2) / 2 if q2 else q1
        r['Qb_ad_jg'] = round(qb_jg, 2)
        r['Qb_ad_kcal'] = round(qb_jg / 4.1868, 2)
        # Qgr,ad = (Qb,ad - St*94.1 - Qb*0.0016) / 4.1868
        r['Qgr_ad'] = round((qb_jg - r['Stad'] * 94.1 - qb_jg * 0.0016) / 4.1868, 2)
        # Qgr,ar = Qgr,ad * (100-Mt)/(100-Mad)
        if r.get('Mt') is not None and r.get('Mad'):
            r['Qgr_ar'] = round(r['Qgr_ad'] * (100 - r['Mt']) / (100 - r['Mad']), 2)
        # Had = 2.888 + 0.393*sqrt(Vdaf) - 0.0023*Adb
        import math
        if r.get('Vdaf') and r.get('Adb'):
            r['Had'] = round(2.888 + 0.393 * math.sqrt(r['Vdaf']) - 0.0023 * r['Adb'], 4)
            r['Hdaf'] = round(r['Had'] * 100 / (100 - r['Mad'] - r['Aad']), 4)
        # Qnet,ar = ((Qgr,ad/238.846*1000 - 206*Had)*(100-Mt)/(100-Mad) - 23*Mt)/4.1868
        if r.get('Qgr_ad') and r.get('Had') and r.get('Mt') is not None and r.get('Mad'):
            qbs = r['Qgr_ad'] / 238.846 * 1000
            r['Qnet_ar'] = round(((qbs - 206 * r['Had']) * (100 - r['Mt']) / (100 - r['Mad']) - 23 * r['Mt']) / 4.1868, 2)

        # G индекс
        if m.get('g_tare') and m.get('g_coke') and m.get('g_sieve1') and m.get('g_sieve2'):
            g1 = 10 + (30*(m['g_sieve1']-m['g_tare']) + 70*(m['g_sieve2']-m['g_tare'])) / (m['g_coke']-m['g_tare'])
        else:
            g1 = 0
        if m.get('g_tare2') and m.get('g_coke2') and m.get('g_sieve1b') and m.get('g_sieve2b'):
            g2 = 10 + (30*(m['g_sieve1b']-m['g_tare2']) + 70*(m['g_sieve2b']-m['g_tare2'])) / (m['g_coke2']-m['g_tare2'])
        else:
            g2 = 0
        r['G_index'] = round((g1 + g2) / 2 if g2 else g1, 2)

        # FSI
        f1 = m.get('fsi_value1', 0) or 0
        f2 = m.get('fsi_value2', 0) or 0
        r['FSI'] = round((f1 + f2) / 2 if f2 else f1, 1)

        # Y индекс
        if r.get('G_index'):
            r['Y_index'] = round(4.89 + r['G_index'] ** 2 * 0.00102, 2)

    except Exception as e:
        app.logger.error(f"Calculation error: {e}")
    return r

def check_qc(results, is_duplicate, meas1, meas2=None):
    """QC шалгалт — зэрэгцээ шинжилгээний зөрүү шалгах"""
    if not is_duplicate or not meas2:
        return 'passed', ''
    conn = get_db()
    settings = {r['parameter']: r['tolerance'] for r in conn.execute("SELECT * FROM qc_settings").fetchall()}
    conn.close()
    warnings = []
    checks = [
        ('Mad', 'Mad'), ('Aad', 'Aad'), ('Vad', 'Vad'),
        ('Stad', 'Stad'), ('Qb_ad_kcal', 'Qb_ad'), ('G_index', 'G_index'), ('FSI', 'FSI')
    ]
    status = 'passed'
    for key, setting_key in checks:
        v1 = results.get(key, 0) or 0
        v2 = meas2.get(key, 0) or 0
        tol = settings.get(setting_key, 999)
        diff = abs(v1 - v2)
        if diff > tol:
            warnings.append(f"{key}: зөрүү {diff:.3f} > {tol}")
            status = 'warning'
    return status, '; '.join(warnings)

@app.route('/analysis')
@login_required
def analysis():
    lang = session.get('lang','mn')
    conn = get_db()
    role = session.get('role')
    uid = session.get('user_id', 0)

    if role == 'bayjuulach':
        # Баяжуулагч зөвхөн 6000-6999 серийн ЯВЖ БУЙ ажлыг харна.
        # Урьд нь эсрэгээрээ (status='done') шүүдэг байсан тул дууссан ажил
        # Шинжилгээ хуудсанд үлддэг байв — бусад үүрэгтэй нийцэхгүй.
        # Дууссан ажил "Нэгдсэн архив"-т үр дүнтэйгээ хамт харагдана.
        u = conn.execute("SELECT can_view_result FROM users WHERE id=?", (uid,)).fetchone()
        if not u or not u['can_view_result']:
            conn.close()
            flash('Үр дүн харах эрх байхгүй байна', 'error')
            return redirect(url_for('dashboard'))
        samples = conn.execute("""
            SELECT g.*, u.name as reg_name,
                   sr.lab_number, sr.lab_serial, sr.received_date, sr.mass_kg, sr.id as receipt_id, sr.prep_status
            FROM geo_samples g
            LEFT JOIN users u ON u.id=g.registered_by
            LEFT JOIN sample_receipt sr ON sr.geo_sample_id=g.id
            WHERE g.status != 'done' AND sr.lab_serial BETWEEN 6000 AND 6999
            ORDER BY sr.lab_serial DESC LIMIT 100
        """).fetchall()
        conn.close()
        return render_template('analysis/index.html', samples=samples, lang=session.get('lang','mn'),
            today=datetime.now().strftime('%Y-%m-%d'), prep_devices=[], pending_qc=[],
            batches=[], active_ids=set())
    if role == 'geologist':
        # Геологч зөвшөөрөгдсөн ажлын дугаарын мужийн дээжийг л харна (view_ranges).
        # view_ranges NULL бол бүгдийг харна (default).
        _u = conn.execute("SELECT view_ranges FROM users WHERE id=?", (uid,)).fetchone()
        _vr = _u['view_ranges'] if _u else None
        if _vr is None:
            samples = conn.execute("""
                SELECT g.*, u.name as reg_name,
                       sr.lab_number, sr.lab_serial, sr.received_date, sr.mass_kg, sr.id as receipt_id, sr.prep_status
                FROM geo_samples g
                LEFT JOIN users u ON u.id=g.registered_by
                LEFT JOIN sample_receipt sr ON sr.geo_sample_id=g.id
                WHERE g.status != 'done'
                ORDER BY sr.lab_serial DESC, g.created_at DESC LIMIT 200
            """).fetchall()
        else:
            _thousands = [int(x) for x in _vr.split(',') if x.strip().isdigit()]
            if not _thousands:
                samples = []
            else:
                _ph = ','.join('?' * len(_thousands))
                samples = conn.execute(f"""
                    SELECT g.*, u.name as reg_name,
                           sr.lab_number, sr.lab_serial, sr.received_date, sr.mass_kg, sr.id as receipt_id, sr.prep_status
                    FROM geo_samples g
                    LEFT JOIN users u ON u.id=g.registered_by
                    LEFT JOIN sample_receipt sr ON sr.geo_sample_id=g.id
                    WHERE g.status != 'done' AND (sr.lab_serial/1000) IN ({_ph})
                    ORDER BY sr.lab_serial DESC, g.created_at DESC LIMIT 200
                """, _thousands).fetchall()
    else:
        samples = conn.execute("""
            SELECT g.*, u.name as reg_name,
                   sr.lab_number, sr.lab_serial, sr.received_date, sr.mass_kg, sr.id as receipt_id, sr.prep_status
            FROM geo_samples g
            LEFT JOIN users u ON u.id=g.registered_by
            LEFT JOIN sample_receipt sr ON sr.geo_sample_id=g.id
            WHERE g.status != 'done'
            ORDER BY sr.lab_serial DESC, g.created_at DESC LIMIT 200
        """).fetchall()
    prep_devices = conn.execute("""
        SELECT id, name FROM devices
        WHERE status='active' AND COALESCE(stage,'both') IN ('prep','both')
        ORDER BY name
    """).fetchall()
    pending_qc = conn.execute("""
        SELECT iq.id, iq.triggered_date, iq.qc_number,
            iq.receipt_id_1, iq.receipt_id_2, iq.row_num_1, iq.row_num_2,
            sr1.lab_number as lab1, g1.sample_name as sname1,
            sr2.lab_number as lab2, g2.sample_name as sname2
        FROM internal_qc iq
        LEFT JOIN sample_receipt sr1 ON sr1.id=iq.receipt_id_1
        LEFT JOIN geo_samples g1 ON g1.id=sr1.geo_sample_id
        LEFT JOIN sample_receipt sr2 ON sr2.id=iq.receipt_id_2
        LEFT JOIN geo_samples g2 ON g2.id=sr2.geo_sample_id
        WHERE iq.status='pending'
        ORDER BY iq.created_at DESC
    """).fetchall()
    # ── Идэвхтэй ажлын багц + "хийгдэж байгаа / хүлээгдэж байгаа" ялгалт ──
    batches = open_batches(conn, uid)
    in_batch = {r for b in batches for r in batch_receipt_ids(b)}
    # Утга орсон эсэх: autosave үед л мөр үүсдэг тул мөрийн оршихуй нь дохио.
    # CRM дээж хүлээн авахдаа хоосон мөр үүсгэдэг тул статус/утгыг мөн шалгана.
    started = {r['receipt_id'] for r in conn.execute("""
        SELECT DISTINCT receipt_id FROM sample_entries
        WHERE (row_status IS NOT NULL AND row_status<>'empty')
           OR mass_kg IS NOT NULL OR dc_tare IS NOT NULL OR ash_tare IS NOT NULL
           OR vol_tare IS NOT NULL OR g_tare IS NOT NULL OR mt_tare IS NOT NULL
           OR ff_sample IS NOT NULL OR sulfur IS NOT NULL OR cal_value IS NOT NULL
           OR fsi IS NOT NULL""")}
    active_ids = in_batch | started
    conn.close()
    return render_template('analysis/index.html', samples=samples, lang=lang,
        today=datetime.now().strftime('%Y-%m-%d'), prep_devices=prep_devices,
        pending_qc=pending_qc, batches=batches, active_ids=active_ids)

def parse_quantity(sample_type, sample_name, fallback=1):
    """Дээжийн нэрнээс дээжийн тоог тооцно.

    Бүртгэл ба ЗАСВАР хоёр ИЖИЛ дүрмээр ажиллах ёстой. Урьд нь энэ логик
    зөвхөн бүртгэлд байсан тул засварлаж нэр нэмэхэд тоо нь хэвээр үлдэж,
    нэмсэн дээж хүснэгтэд гардаггүй байв.

      "a1; a2; b3"  → 3   (цэг таслалаар тусгаарласан нэрс)
      "1-100"       → 100 (зөвхөн PIT — мужаар)
      бусад         → fallback
    """
    import re as _re
    name = (sample_name or '').strip()
    if ';' in name:
        parts = [p.strip() for p in name.split(';') if p.strip()]
        if parts:
            return len(parts)
    if sample_type == 'PIT':
        m = _re.match(r'^([0-9]+)-([0-9]+)$', name)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            if b >= a:
                return b - a + 1
    return fallback


# Мөрөнд ямар нэг хэмжилт/төлөв орсон эсэх — дээжийн тоог багасгахаас өмнө
# шалгана (утга орсон мөрийг чимээгүй устгаж болохгүй).
# sample_entries нь "se" алиастай байх ёстой: mass_kg, sample_name зэрэг багана
# sample_receipt/geo_samples дээр ч байдаг тул тодотголгүй бол SQL нь
# "ambiguous column name" гэж унана.
ROW_HAS_DATA = """((se.row_status IS NOT NULL AND se.row_status<>'empty')
    OR se.mass_kg IS NOT NULL OR se.sample_name IS NOT NULL
    OR se.dc_tare IS NOT NULL OR se.ash_tare IS NOT NULL OR se.vol_tare IS NOT NULL
    OR se.g_tare IS NOT NULL OR se.mt_tare IS NOT NULL OR se.ff_sample IS NOT NULL
    OR se.sulfur IS NOT NULL OR se.cal_value IS NOT NULL OR se.fsi IS NOT NULL)"""


@app.route('/analysis/register', methods=['GET','POST'])
@login_required
def analysis_register():
    """Геологи дээж бүртгэнэ"""
    lang = session.get('lang','mn')
    role = session.get('role')
    if role not in ('admin','senior','staff','preparer','geologist','bayjuulach'):
        return redirect(url_for('dashboard'))
    if role == 'bayjuulach':
        conn2 = get_db()
        u2 = conn2.execute("SELECT can_register FROM users WHERE id=?", (session.get('user_id'),)).fetchone()
        conn2.close()
        if not u2 or not u2['can_register']:
            flash('Дээж бүртгэх эрх байхгүй байна.', 'error')
            return redirect(url_for('dashboard'))
    if request.method == 'POST':
        conn = get_db()
        sample_type = request.form['sample_type']
        sample_name = request.form['sample_name']
        quantity    = int(request.form.get('quantity', 1))

        # Нэрнээс тоог тооцно — засварт мөн ижил функц ашиглагдана
        quantity = parse_quantity(sample_type, sample_name, quantity)

        # created_at-ыг гараар бичнэ: SQLite-ийн CURRENT_TIMESTAMP үргэлж UTC
        # буцаадаг тул статистик долоо хоног/сараар буруу хуваагдана
        conn.execute("""
            INSERT INTO geo_samples(sample_name,sample_type,location,collected_date,
            quantity,notes,registered_by,status,created_at)
            VALUES(?,?,?,?,?,?,?,'pending',?)
        """, (
            sample_name,
            sample_type,
            request.form.get('location'),
            request.form.get('collected_date'),
            quantity,
            request.form.get('notes'),
            session['user_id'],
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ))
        conn.commit(); conn.close()
        flash(f'Дээж бүртгэгдлээ! Нийт {quantity} дээж.', 'success')
        return redirect(url_for('analysis'))
    conn = get_db()
    stypes = conn.execute("SELECT * FROM sample_types WHERE is_active=1 ORDER BY sort_order,serial_from").fetchall()
    conn.close()
    return render_template('analysis/register.html', lang=lang, 
        today=datetime.now().strftime('%Y-%m-%d'),
        sample_types=stypes)

@app.route('/analysis/crm/register', methods=['GET', 'POST'])
@lab_required
def analysis_crm_register():
    conn = get_db()
    if request.method == 'POST':
        mat_id = request.form.get('crm_material_id', '').strip()
        collected_date = request.form.get('collected_date') or datetime.now().strftime('%Y-%m-%d')
        notes = request.form.get('notes', '').strip()

        if not mat_id:
            flash('CRM материал сонгоно уу', 'error')
            conn.close()
            return redirect(url_for('analysis_crm_register'))

        mat = conn.execute("SELECT * FROM crm_materials WHERE id=?", (mat_id,)).fetchone()
        if not mat:
            flash('CRM материал олдсонгүй', 'error')
            conn.close()
            return redirect(url_for('analysis_crm_register'))

        crm_name = mat['crm_name']
        try:
            cur = conn.execute("""
                INSERT INTO geo_samples (sample_name, sample_type, location, collected_date, quantity, notes,
                    registered_by, status, crm_name, crm_aad, crm_vad, crm_sulfur, crm_cal,
                    crm_g, crm_g_unc, crm_mad,
                    crm_aad_unc, crm_vad_unc, crm_sulfur_unc, crm_cal_unc, crm_mad_unc)
                VALUES (?, 'CRM', 'CRM', ?, 1, ?, ?, 'received', ?, ?, ?, ?, ?, ?, ?, ?,
                        ?, ?, ?, ?, ?)
            """, (crm_name, collected_date, notes, session['user_id'],
                  crm_name, mat['aad_cert'], mat['vad_cert'], mat['sulfur_cert'], mat['cal_cert'],
                  _mat(mat, 'g_cert'), _mat(mat, 'g_unc'), _mat(mat, 'mad_cert'),
                  _mat(mat, 'aad_unc'), _mat(mat, 'vad_unc'), _mat(mat, 'sulfur_unc'),
                  _mat(mat, 'cal_unc'), _mat(mat, 'mad_unc')))
            geo_id = cur.lastrowid

            crm_lab_num = f"CRM {crm_name} {collected_date.replace('-','')}"
            # ensure uniqueness if same CRM registered multiple times on same day
            _suffix = 0
            while conn.execute("SELECT 1 FROM sample_receipt WHERE lab_number=?", (crm_lab_num + (f'-{_suffix}' if _suffix else ''),)).fetchone():
                _suffix += 1
            if _suffix:
                crm_lab_num = crm_lab_num + f'-{_suffix}'
            cur2 = conn.execute("""
                INSERT INTO sample_receipt (geo_sample_id, lab_number, lab_serial, received_date, received_by, prep_status)
                VALUES (?, ?, ?, ?, ?, 'ready')
            """, (geo_id, crm_lab_num, 0, collected_date, session['user_id']))
            receipt_id = cur2.lastrowid

            conn.execute("""
                INSERT INTO sample_entries (receipt_id, row_num, is_duplicate, sample_name, row_status)
                VALUES (?, 1, 0, ?, 'empty')
            """, (receipt_id, crm_name))

            conn.commit()
            conn.close()
            flash(f'CRM дээж бүртгэгдлээ: {crm_name}', 'success')
            return redirect(url_for('analysis_measure', receipt_id=receipt_id))
        except Exception as e:
            import traceback; traceback.print_exc()
            conn.rollback()
            conn.close()
            flash(f'Алдаа: {e}', 'error')
            return redirect(url_for('analysis_crm_register'))

    crm_materials = conn.execute("SELECT * FROM crm_materials WHERE is_active=1 ORDER BY crm_name").fetchall()
    conn.close()
    return render_template('analysis/crm_register.html', today=datetime.now().strftime('%Y-%m-%d'),
                           crm_materials=crm_materials)

def can_edit_sample(geo, receipt_exists=None):
    """Тухайн хэрэглэгч энэ дээжийн бүртгэлийг засах/устгах эрхтэй эсэх.

    - Админ, ахлах химич: үргэлж
    - Бүртгэсэн хүн өөрөө (геологич, баяжуулах): ЗӨВХӨН лаб хүлээж авахаас
      өмнө. Хүлээн авалт бүртгэгдмэгц ажлын дугаар олгогдож, лаборатори
      ажлаа эхэлсэн байдаг тул тэр цэгээс хойш зөвхөн ахлах засна.

    Ингэснээр геологич өөрийн бичсэн алдаагаа шууд засах боловч
    лабораторийн бүртгэлийг сүүлээр өөрчлөх боломжгүй болно.
    """
    role = session.get('role')
    if role in ('admin', 'senior'):
        return True
    if role == 'guest' or not geo:
        return False
    if geo['registered_by'] != session.get('user_id'):
        return False
    if receipt_exists:
        return False
    return (geo['status'] or 'pending') == 'pending'


@app.context_processor
def inject_sample_perm():
    """Загварт can_edit_sample-ийг ашиглах боломжтой болгоно"""
    return {'can_edit_sample': can_edit_sample}


@app.route('/analysis/sample/<int:geo_id>/edit', methods=['POST'])
@login_required
def sample_edit(geo_id):
    """Бүртгэсэн дээжийн нэр болон ажлын дугаарыг засах.

    Админ/ахлах, эсвэл бүртгэсэн хүн өөрөө (лаб хүлээж авахаас өмнө).
    """
    lang = session.get('lang','mn')
    new_name   = (request.form.get('sample_name') or '').strip()
    new_serial = request.form.get('lab_serial','').strip()
    conn = get_db()
    geo = conn.execute("SELECT * FROM geo_samples WHERE id=?", (geo_id,)).fetchone()
    if not geo:
        conn.close()
        flash('Дээж олдсонгүй.', 'error')
        return redirect(url_for('analysis'))
    has_rec = conn.execute('SELECT 1 FROM sample_receipt WHERE geo_sample_id=?',
                           (geo_id,)).fetchone() is not None
    if not can_edit_sample(geo, has_rec):
        conn.close()
        flash('Лаборатори хүлээн авсны дараа зөвхөн ахлах засна.'
              if has_rec else 'Энэ бүртгэлийг засах эрх байхгүй байна.', 'error')
        return redirect(request.referrer or url_for('analysis'))
    # ── Нэр засах — дээжийн ТООГ мөн дагуулж шинэчилнэ ──
    # Урьд нь зөвхөн нэр шинэчлэгддэг байсан тул "4 дээж бүртгэх байснаа 3
    # бүртгээд" дараа нь дутуу нэрээ нэмэхэд тоо нь 3 хэвээр үлдэж, нэмсэн
    # дээж хүснэгтэд огт гарахгүй байв.
    if new_name:
        cur_qty = geo['quantity'] or 1
        # Нэрнээс тоо гарахгүй бол (ганц нэр дээр олон дээж) маягтад гараар
        # бичсэн тоог ашиглана. Хоёулаа байхгүй бол хуучин тоо хэвээр.
        _q = request.form.get('quantity', '').strip()
        want = int(_q) if _q.isdigit() and int(_q) >= 1 else cur_qty
        new_qty = parse_quantity(geo['sample_type'], new_name, want)
        if new_qty < cur_qty:
            # Багасгах үед хасагдах мөрөнд утга орсон эсэхийг шалгана —
            # хэмжсэн үр дүнг чимээгүй устгаж болохгүй.
            used = conn.execute(f"""
                SELECT COUNT(*) c FROM sample_entries se
                  JOIN sample_receipt sr ON sr.id = se.receipt_id
                 WHERE sr.geo_sample_id = ? AND se.row_num > ? AND {ROW_HAS_DATA}
            """, (geo_id, new_qty)).fetchone()['c']
            if used:
                conn.close()
                flash(f'Дээжийн тоо {cur_qty} → {new_qty} болж багасах ба хасагдах '
                      f'мөрөнд хэмжилтийн утга орсон байна. Эхлээд тэр мөрүүдийн '
                      f'утгыг цэвэрлэнэ үү.', 'error')
                return redirect(request.referrer or url_for('analysis'))
            conn.execute("""DELETE FROM sample_entries WHERE row_num > ? AND receipt_id IN
                            (SELECT id FROM sample_receipt WHERE geo_sample_id=?)""",
                         (new_qty, geo_id))
        conn.execute("UPDATE geo_samples SET sample_name=?, quantity=? WHERE id=?",
                     (new_name, new_qty, geo_id))
    # Ажлын дугаар засах (receipt байвал)
    if new_serial and new_serial.isdigit():
        new_serial = int(new_serial)
        receipt = conn.execute("SELECT * FROM sample_receipt WHERE geo_sample_id=?", (geo_id,)).fetchone()
        if receipt:
            # Давхцал шалгах (өөр дээж дээр ижил дугаар байвал болохгүй)
            dup = conn.execute(
                "SELECT id FROM sample_receipt WHERE lab_serial=? AND id!=?",
                (new_serial, receipt['id'])).fetchone()
            if dup:
                conn.close()
                flash(f'{new_serial} дугаар өөр ажилд бүртгэгдсэн байна.', 'error')
                return redirect(request.referrer or url_for('analysis'))
            date_str = (receipt['received_date'] or '').replace('-','')
            new_labnum = f"{new_serial}-{date_str}" if date_str else str(new_serial)
            conn.execute("UPDATE sample_receipt SET lab_serial=?, lab_number=? WHERE id=?",
                         (new_serial, new_labnum, receipt['id']))
    conn.commit()
    conn.close()
    if new_name and new_qty != cur_qty:
        flash(f'Дээжийн мэдээлэл засагдлаа. Дээжийн тоо {cur_qty} → {new_qty} боллоо.',
              'success')
    else:
        flash('Дээжийн мэдээлэл засагдлаа.', 'success')
    return redirect(request.referrer or url_for('analysis'))

@app.route('/analysis/sample/<int:geo_id>/delete', methods=['POST'])
@login_required
def sample_delete(geo_id):
    """Буруу бүртгэсэн дээжийг жагсаалтаас бүрмөсөн устгана.

    Админ/ахлах, эсвэл бүртгэсэн хүн өөрөө (лаб хүлээж авахаас өмнө).
    Хэмжилтийн утга орсон бол устгахгүй — тэр тохиолдолд Архивын устгалыг
    (админ) ашиглана. Буцаах боломжгүй тул устгахын өмнө нөөцөлнө.
    """
    conn = get_db()
    geo = conn.execute('SELECT * FROM geo_samples WHERE id=?', (geo_id,)).fetchone()
    if not geo:
        conn.close()
        flash('Дээж олдсонгүй.', 'error')
        return redirect(url_for('analysis'))

    rec = conn.execute('SELECT id, lab_number FROM sample_receipt WHERE geo_sample_id=?',
                       (geo_id,)).fetchone()
    rid = rec['id'] if rec else None

    if not can_edit_sample(geo, rid is not None):
        conn.close()
        flash('Лаборатори хүлээн авсны дараа зөвхөн ахлах устгана.'
              if rid else 'Энэ бүртгэлийг устгах эрх байхгүй байна.', 'error')
        return redirect(request.referrer or url_for('analysis'))

    if rid:
        # 1. Хэмжилт орсон эсэх — орсон бол устгахгүй
        used = conn.execute(
            f'SELECT COUNT(*) c FROM sample_entries se WHERE se.receipt_id=? '
            f'AND {ROW_HAS_DATA}', (rid,)).fetchone()['c']
        if used:
            conn.close()
            flash(f'«{geo["sample_name"]}» дээр {used} мөрөнд хэмжилтийн утга орсон '
                  f'байна — устгах боломжгүй. Шаардлагатай бол админ Архиваас устгана.',
                  'error')
            return redirect(request.referrer or url_for('analysis'))
        # 2. Дотоод QC-тэй холбоотой эсэх
        qc = conn.execute('SELECT COUNT(*) c FROM internal_qc '
                          'WHERE receipt_id_1=? OR receipt_id_2=?',
                          (rid, rid)).fetchone()['c']
        if qc:
            conn.close()
            flash(f'Энэ дээжтэй {qc} дотоод QC бүртгэл холбоотой байна. '
                  f'Эхлээд QC-г устгана уу.', 'error')
            return redirect(request.referrer or url_for('analysis'))

    try:
        _p = make_backup()
        bk = os.path.basename(_p) if _p else None
    except Exception:
        app.logger.exception('Дээж устгахын өмнөх нөөцлөлт амжилтгүй')
        bk = None

    try:
        if rid:
            # Нээлттэй багцад орсон бол багцаас нь салгана (үлдвэл "Үргэлжлүүлэх"
            # нь байхгүй ажил руу заана)
            for b in conn.execute("SELECT id, receipt_ids FROM work_batch "
                                  "WHERE status='open'").fetchall():
                ids = [i for i in (b['receipt_ids'] or '').split(',') if i.strip()]
                if str(rid) not in ids:
                    continue
                left = [i for i in ids if i != str(rid)]
                if left:
                    conn.execute('UPDATE work_batch SET receipt_ids=? WHERE id=?',
                                 (','.join(left), b['id']))
                else:
                    conn.execute("UPDATE work_batch SET status='closed', closed_at=? "
                                 "WHERE id=?", (datetime.now().isoformat(), b['id']))
            conn.execute('DELETE FROM result_view_log  WHERE receipt_id=?', (rid,))
            conn.execute('DELETE FROM device_usage_log WHERE receipt_id=?', (rid,))
            conn.execute('DELETE FROM sample_entries   WHERE receipt_id=?', (rid,))
            conn.execute('DELETE FROM sample_receipt   WHERE id=?', (rid,))
        conn.execute('DELETE FROM geo_samples WHERE id=?', (geo_id,))
        conn.commit()
        msg = f'«{geo["sample_name"]}» устгагдлаа.'
        if rec and rec['lab_number']:
            msg = f'«{geo["sample_name"]}» ({rec["lab_number"]}) устгагдлаа.'
        if bk:
            msg += f' Устгахын өмнө нөөц авсан: {bk}'
        flash(msg, 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Устгахад алдаа гарлаа: {e}', 'error')
    finally:
        conn.close()
    return redirect(request.referrer or url_for('analysis'))


@app.route('/analysis/crm/chart')
@login_required
def crm_control_chart():
    conn = get_db()
    materials = conn.execute("SELECT * FROM crm_materials WHERE is_active=1 ORDER BY crm_name").fetchall()
    mat_id = request.args.get('mat_id', type=int)
    selected = None
    history = []
    if not mat_id and materials:
        mat_id = materials[0]['id']
    if mat_id:
        selected = conn.execute("SELECT * FROM crm_materials WHERE id=?", (mat_id,)).fetchone()
        if selected:
            rows = conn.execute("""
                SELECT g.collected_date, sr.received_date, sr.lab_number,
                       se.mad, se.aad, se.vad, se.sulfur, se.cal_value, se.g_val as g_value
                FROM geo_samples g
                JOIN sample_receipt sr ON sr.geo_sample_id=g.id
                JOIN sample_entries se ON se.receipt_id=sr.id AND se.is_duplicate=0 AND se.row_num=1
                WHERE g.crm_name=? AND g.sample_type='CRM'
                  AND se.row_status IN ('done','approved')
                ORDER BY COALESCE(sr.received_date, g.collected_date)
            """, (selected['crm_name'],)).fetchall()
            # ad → db хувиргалт
            history = []
            for r in rows:
                mad = r['mad'] or 0
                factor = 1 - mad / 100 if mad < 100 else 1
                def to_db(v):
                    if v is None or factor == 0: return None
                    return round(v / factor, 4)
                history.append({
                    'collected_date': r['collected_date'] or r['received_date'],
                    'lab_number':     r['lab_number'],
                    'mad':            r['mad'],
                    'aad':            r['aad'],
                    'vad':            r['vad'],
                    'sulfur':         r['sulfur'],
                    'cal_value':      r['cal_value'],
                    'g_value':        r['g_value'],
                    # dry basis
                    'adb':  to_db(r['aad']),
                    'vdb':  to_db(r['vad']),
                    'sdb':  to_db(r['sulfur']),
                    'qbd':  round(r['cal_value'] / factor / 1000, 4) if r['cal_value'] and factor else None,
                })
    conn.close()
    return render_template('analysis/crm_chart.html',
        materials=materials, selected=selected, selected_id=mat_id, history=history,
        today=date.today().isoformat())


@app.route('/analysis/receive/<int:geo_id>', methods=['GET','POST'])
@preparer_required
def analysis_receive(geo_id):
    """Дээж бэлтгэгч хүлээн авна — зөвхөн ажлын дугаар + жин"""
    lang = session.get('lang','mn')
    conn = get_db()
    geo = conn.execute("SELECT * FROM geo_samples WHERE id=?", (geo_id,)).fetchone()
    if not geo:
        conn.close(); return redirect(url_for('analysis'))

    # Санал болгох серийн дугаар
    prefix_map = {'PIT':1000,'STOCKPILE':2000,'EXPORT':3000,
                  'CONTROL':4000,'EQ_CONTROL':5000,'DP':6000}
    base = prefix_map.get(geo['sample_type'], 1000)
    last = conn.execute(
        "SELECT lab_serial FROM sample_receipt WHERE lab_serial>=? AND lab_serial<? ORDER BY lab_serial DESC LIMIT 1",
        (base, base+1000)
    ).fetchone()
    suggested = (last['lab_serial'] + 1) if last else (base + 1)

    if request.method == 'POST':
        serial = int(request.form.get('lab_serial', suggested))
        date_str = request.form.get('received_date','').replace('-','')
        lab_num = f"{serial}-{date_str}"
        conn.execute("""
            INSERT INTO sample_receipt(
                geo_sample_id, lab_number, lab_serial, received_date,
                received_by, notes
            ) VALUES(?,?,?,?,?,?)
        """, (
            geo_id, lab_num, serial,
            request.form.get('received_date', datetime.now().strftime('%Y-%m-%d')),
            session['user_id'],
            request.form.get('notes')
        ))
        conn.execute("UPDATE geo_samples SET status='received' WHERE id=?", (geo_id,))
        conn.commit(); conn.close()
        flash(f'Дээж хүлээн авлаа! Ажлын дугаар: {lab_num}', 'success')
        return redirect(url_for('analysis'))
    conn.close()
    return render_template('analysis/receive.html', geo=geo, lang=lang,
        today=datetime.now().strftime('%Y-%m-%d'),
        suggested_serial=suggested)

@app.route('/analysis/measure/<int:receipt_id>', methods=['GET','POST'])
@lab_required
def analysis_measure(receipt_id):
    """Химич шинжилгээний утгуудыг оруулна — Excel Sheet 2 шиг"""
    lang = session.get('lang','mn')
    conn = get_db()
    receipt = conn.execute("""
        SELECT sr.*, g.sample_name, g.sample_type, g.location,
               g.collected_date, g.quantity,
               g.notes as geo_notes,
               ug.name as geo_name,
               up.name as prep_name
        FROM sample_receipt sr
        JOIN geo_samples g ON g.id=sr.geo_sample_id
        LEFT JOIN users ug ON ug.id=g.registered_by
        LEFT JOIN users up ON up.id=sr.received_by
        WHERE sr.id=?
    """, (receipt_id,)).fetchone()
    if not receipt:
        conn.close(); return redirect(url_for('analysis'))
    if request.method == 'POST':
        def flt(key): return float(request.form.get(key)) if request.form.get(key) else None
        m = {
            'fm_sample_mass': receipt['fm_sample_mass'],
            'fm_dried_mass': receipt['fm_dried_mass'],
            'mt_tare1': receipt['mt_tare1'], 'mt_sample1': receipt['mt_sample1'],
            'mt_dried1': receipt['mt_dried1'], 'mt_tare2': receipt['mt_tare2'],
            'mt_sample2': receipt['mt_sample2'], 'mt_dried2': receipt['mt_dried2'],
            'im_tare1': flt('im_tare1'), 'im_sample1': flt('im_sample1'), 'im_dried1': flt('im_dried1'),
            'im_tare2': flt('im_tare2'), 'im_sample2': flt('im_sample2'), 'im_dried2': flt('im_dried2'),
            'ash_tare1': flt('ash_tare1'), 'ash_sample1': flt('ash_sample1'), 'ash_burned1': flt('ash_burned1'),
            'ash_tare2': flt('ash_tare2'), 'ash_sample2': flt('ash_sample2'), 'ash_burned2': flt('ash_burned2'),
            'vol_tare1': flt('vol_tare1'), 'vol_sample1': flt('vol_sample1'), 'vol_burned1': flt('vol_burned1'),
            'vol_tare2': flt('vol_tare2'), 'vol_sample2': flt('vol_sample2'), 'vol_burned2': flt('vol_burned2'),
            'sulfur1': flt('sulfur1'), 'sulfur2': flt('sulfur2'),
            'cal_value1': flt('cal_value1'), 'cal_value2': flt('cal_value2'),
            'g_tare': flt('g_tare'), 'g_coke': flt('g_coke'),
            'g_sieve1': flt('g_sieve1'), 'g_sieve2': flt('g_sieve2'),
            'g_tare2': flt('g_tare2'), 'g_coke2': flt('g_coke2'),
            'g_sieve1b': flt('g_sieve1b'), 'g_sieve2b': flt('g_sieve2b'),
            'fsi_value1': flt('fsi_value1'), 'fsi_value2': flt('fsi_value2'),
        }
        conn.execute("""
            INSERT INTO analysis_measurements(
                receipt_id,
                im_tare1,im_sample1,im_dried1,im_crucible1,
                im_tare2,im_sample2,im_dried2,im_crucible2,
                im_date,im_shift,im_operator,im_device,
                ash_tare1,ash_sample1,ash_burned1,ash_crucible1,
                ash_tare2,ash_sample2,ash_burned2,ash_crucible2,
                ash_date,ash_shift,ash_operator,ash_device,
                vol_tare1,vol_sample1,vol_burned1,vol_crucible1,
                vol_tare2,vol_sample2,vol_burned2,vol_crucible2,
                vol_date,vol_shift,vol_operator,vol_device,
                sulfur1,sulfur2,sulfur_date,sulfur_shift,sulfur_operator,
                cal_value1,cal_value2,cal_date,cal_shift,cal_operator,
                g_tare,g_coke,g_sieve1,g_sieve2,g_crucible,g_date,g_shift,g_operator,g_device,
                g_tare2,g_coke2,g_sieve1b,g_sieve2b,
                fsi_value1,fsi_value2,fsi_date,fsi_shift,fsi_operator
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            receipt_id,
            m['im_tare1'],m['im_sample1'],m['im_dried1'],request.form.get('im_crucible1'),
            m['im_tare2'],m['im_sample2'],m['im_dried2'],request.form.get('im_crucible2'),
            request.form.get('im_date'),request.form.get('im_shift','A'),session['user_id'],request.form.get('im_device'),
            m['ash_tare1'],m['ash_sample1'],m['ash_burned1'],request.form.get('ash_crucible1'),
            m['ash_tare2'],m['ash_sample2'],m['ash_burned2'],request.form.get('ash_crucible2'),
            request.form.get('ash_date'),request.form.get('ash_shift','A'),session['user_id'],request.form.get('ash_device'),
            m['vol_tare1'],m['vol_sample1'],m['vol_burned1'],request.form.get('vol_crucible1'),
            m['vol_tare2'],m['vol_sample2'],m['vol_burned2'],request.form.get('vol_crucible2'),
            request.form.get('vol_date'),request.form.get('vol_shift','A'),session['user_id'],request.form.get('vol_device'),
            m['sulfur1'],m['sulfur2'],request.form.get('sulfur_date'),request.form.get('sulfur_shift','A'),session['user_id'],
            m['cal_value1'],m['cal_value2'],request.form.get('cal_date'),request.form.get('cal_shift','A'),session['user_id'],
            m['g_tare'],m['g_coke'],m['g_sieve1'],m['g_sieve2'],
            request.form.get('g_crucible'),request.form.get('g_date'),request.form.get('g_shift','A'),session['user_id'],request.form.get('g_device'),
            m['g_tare2'],m['g_coke2'],m['g_sieve1b'],m['g_sieve2b'],
            m['fsi_value1'],m['fsi_value2'],request.form.get('fsi_date'),request.form.get('fsi_shift','A'),session['user_id']
        ))
        # Тооцоолол
        results = calculate_results(m)
        qc_status, qc_notes = 'passed', ''
        # Үр дүн хадгалах
        conn.execute("""
            INSERT OR REPLACE INTO analysis_results(
                receipt_id, Mt, Mad, Aad, Adb, Vad, Vdb, Vdaf, FCad,
                Stad, Std, Qb_ad_jg, Qb_ad_kcal, Qgr_ad, Qgr_ar, Qnet_ar,
                Had, Hdaf, G_index, FSI, Y_index, qc_status, qc_notes
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            receipt_id,
            results.get('Mt'), results.get('Mad'), results.get('Aad'), results.get('Adb'),
            results.get('Vad'), results.get('Vdb'), results.get('Vdaf'), results.get('FCad'),
            results.get('Stad'), results.get('Std'),
            results.get('Qb_ad_jg'), results.get('Qb_ad_kcal'),
            results.get('Qgr_ad'), results.get('Qgr_ar'), results.get('Qnet_ar'),
            results.get('Had'), results.get('Hdaf'),
            results.get('G_index'), results.get('FSI'), results.get('Y_index'),
            qc_status, qc_notes
        ))
        conn.execute("UPDATE geo_samples SET status='done' WHERE id=(SELECT geo_sample_id FROM sample_receipt WHERE id=?)", (receipt_id,))
        conn.commit(); conn.close()
        flash('Шинжилгээ бүртгэгдлээ! Үр дүн тооцоологдлоо.', 'success')
        return redirect(url_for('analysis_result', receipt_id=receipt_id))
    qc_map = {r['parameter']: r['tolerance'] for r in conn.execute("SELECT parameter, tolerance FROM qc_settings").fetchall()}
    conn.close()
    return render_template('analysis/measure.html', receipt=receipt, lang=lang, qc_map=qc_map, today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/analysis/approve/<int:receipt_id>', methods=['POST'])
@senior_required
def analysis_approve(receipt_id):
    conn = get_db()
    conn.execute("""UPDATE analysis_results SET approved_by=?, approved_at=?, qc_status='approved'
                   WHERE receipt_id=?""",
                (session['user_id'], datetime.now().isoformat(), receipt_id))
    conn.commit(); conn.close()
    flash('Баталгаажлаа! Тайлан гаргах боломжтой.', 'success')
    return redirect(url_for('analysis_result', receipt_id=receipt_id))



# ── ANALYSIS AUTO-SAVE ──────────────────────────────────
@app.route('/analysis/autosave', methods=['POST'])
@lab_required
def analysis_autosave():
    """Нүд орхих бүрт автоматаар хадгална"""
    try:
        data = request.get_json()
        rid    = data.get('receipt_id')
        row    = data.get('row_num')
        is_dup = data.get('is_duplicate', 0)
        field  = data.get('field')
        value  = data.get('value')

        if rid is None or row is None or not field:
            return jsonify({'ok': False, 'error': 'Missing params'})

        conn = get_db()
        # SQL injection-оос хамгаалах: field нь sample_entries-ийн жинхэнэ багана байх ёстой
        _cols = {r[1] for r in conn.execute("PRAGMA table_info(sample_entries)")}
        _protected = {'id','receipt_id','row_num','is_duplicate','created_at'}
        if field not in _cols or field in _protected:
            conn.close()
            return jsonify({'ok': False, 'error': 'Invalid field'})
        # Мөр байгаа эсэх шалгах
        existing = conn.execute(
            "SELECT id FROM sample_entries WHERE receipt_id=? AND row_num=? AND is_duplicate=?",
            (rid, row, is_dup)
        ).fetchone()

        _empty = (value is None or value == '')
        # ── Утга арилах/дарагдахыг БҮРТГЭНЭ ──────────────────────────────
        # "Хариу нь байгаа хэрнээ масс нь алга болчиж" гэдэг гомдол давтагдаж
        # байгаа тул урьд нь утгатай байсан талбар өөрчлөгдөх бүрд хуучин
        # утгыг нь хадгална. Ингэснээр (а) яг хэн, хэзээ, хаанаас хийснийг
        # мэдэх, (б) шаардлагатай бол утгыг нь буцааж сэргээх боломжтой.
        if existing:
            try:
                _old = conn.execute(
                    f'SELECT {field} AS v FROM sample_entries WHERE id=?',
                    (existing['id'],)).fetchone()['v']
                _new = None if _empty else value
                if _old is not None and str(_old) != str(_new):
                    conn.execute("""
                        INSERT INTO value_audit(receipt_id, row_num, is_duplicate,
                            field, old_value, new_value, user_id, at, source)
                        VALUES(?,?,?,?,?,?,?,?,?)""",
                        (rid, row, is_dup, field, str(_old),
                         None if _new is None else str(_new),
                         session.get('user_id'), datetime.now().isoformat(),
                         (request.referrer or '')[-120:]))
            except Exception as _e:
                app.logger.warning('value_audit бичигдсэнгүй: %s', _e)
        if existing or not _empty:
            # Атомик UPSERT — тооцооллын хадгалалттай зэрэг ажиллахад UNIQUE
            # зөрчил үүсгэхгүй. Хоосон утгаар шинэ мөр үүсгэхгүй (дээрх нөхцөл).
            # Тухайн талбар аль шинжилгээнийх вэ — гүйцэтгэгчийг нь тэмдэглэнэ.
            # Утгыг ХООСРУУЛАХАД гүйцэтгэгч хэвээр үлдэнэ (буруу шивээд засах
            # нь ажил хийгээгүй гэсэн үг биш).
            _op = FIELD_OP.get(field) if not _empty else None
            _op = _op if (_op and _op in _cols) else None
            _extra_ins = f', {_op}' if _op else ''
            _extra_val = ', ?' if _op else ''
            _extra_upd = f', {_op}=excluded.{_op}' if _op else ''
            # Хоосныг NULL болгож бичнэ. Урьд нь '' гэсэн хоосон мөр
            # хадгалагдаж, "утга байгаа" (IS NOT NULL) мэт тоологдож байв.
            _args = [rid, row, is_dup, None if _empty else value,
                     session['user_id'], datetime.now().isoformat()]
            if _op:
                _args.append(session['user_id'])
            conn.execute(
                f"""INSERT INTO sample_entries(receipt_id, row_num, is_duplicate,
                        {field}, updated_by, updated_at{_extra_ins})
                    VALUES(?,?,?,?,?,?{_extra_val})
                    ON CONFLICT(receipt_id, row_num, is_duplicate) DO UPDATE SET
                        {field}=excluded.{field},
                        updated_by=excluded.updated_by,
                        updated_at=excluded.updated_at{_extra_upd}""",
                _args
            )
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        app.logger.exception('autosave алдаа')
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/analysis/autosave/calc', methods=['POST'])
@lab_required
def analysis_autosave_calc():
    """Тооцоолсон үр дүнг хадгална"""
    try:
        data = request.get_json()
        rid    = data.get('receipt_id')
        row    = data.get('row_num')
        is_dup = data.get('is_duplicate', 0)
        mad    = data.get('mad')
        aad    = data.get('aad')
        vad    = data.get('vad')
        fc     = data.get('fc')
        g_val  = data.get('g_val')

        conn = get_db()
        if not any(v is not None for v in (mad, aad, vad, fc, g_val)):
            # Хэмжигч жингүүдийг бүгдийг хоосруулсан — БАЙГАА мөрийн тооцоог
            # цэвэрлэнэ. Урьд нь юу ч хийхгүй буцдаг тул дэлгэц "—" харуулж
            # байхад DB-д хуучин тооцоо үлдэж, тайланд гардаг байсан.
            conn.execute("""UPDATE sample_entries
                SET mad=NULL, aad=NULL, vad=NULL, fc=NULL, g_val=NULL,
                    updated_by=?, updated_at=?
                WHERE receipt_id=? AND row_num=? AND is_duplicate=?""",
                (session['user_id'], datetime.now().isoformat(), rid, row, is_dup))
            conn.commit()
            conn.close()
            return jsonify({'ok': True})

        # Атомик UPSERT. Урьд нь SELECT→INSERT хийдэг байсан нь нүдний autosave-тай
        # зэрэг ажиллахад UNIQUE зөрчил үүсгэж, тооцоолсон утга чимээгүй алдагддаг
        # байсан (хүсэлт нь 200 буцаадаг тул мэдэгддэггүй).
        conn.execute("""
            INSERT INTO sample_entries(receipt_id, row_num, is_duplicate,
                mad, aad, vad, fc, g_val, updated_by, updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(receipt_id, row_num, is_duplicate) DO UPDATE SET
                mad=excluded.mad, aad=excluded.aad, vad=excluded.vad,
                fc=excluded.fc, g_val=excluded.g_val,
                updated_at=excluded.updated_at
        """, (rid, row, is_dup, mad, aad, vad, fc, g_val,
              session['user_id'], datetime.now().isoformat()))
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        app.logger.exception('autosave/calc алдаа')
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/analysis/load/<int:receipt_id>')
@lab_required
def analysis_load(receipt_id):
    """Хадгалагдсан өгөгдлийг ачаална"""
    conn = get_db()
    entries = conn.execute(
        "SELECT * FROM sample_entries WHERE receipt_id=? ORDER BY row_num, is_duplicate",
        (receipt_id,)
    ).fetchall()
    conn.close()
    result = {}
    for e in entries:
        key = f"{e['row_num']}_{e['is_duplicate']}"
        result[key] = dict(e)
    return jsonify(result)


@app.route('/analysis/prep/start/<int:receipt_id>', methods=['POST'])
@preparer_required
def prep_start(receipt_id):
    conn = get_db()
    conn.execute("""UPDATE sample_receipt SET prep_status='preparing', prep_started_at=?
                   WHERE id=?""", (datetime.now().isoformat(), receipt_id))
    conn.execute("UPDATE geo_samples SET status='received' WHERE id=(SELECT geo_sample_id FROM sample_receipt WHERE id=?)", (receipt_id,))
    conn.commit(); conn.close()
    flash('Дээж бэлтгэж эхэллээ!', 'success')
    return redirect(url_for('analysis'))

@app.route('/analysis/prep/done/<int:receipt_id>', methods=['POST'])
@preparer_required
def prep_done(receipt_id):
    conn = get_db()
    notes = request.form.get('notes','')
    prep_operator = request.form.get('prep_operator','').strip()
    prep_position = request.form.get('prep_position','').strip()
    prep_devices  = ', '.join(filter(None, request.form.getlist('prep_device')))
    conn.execute("""UPDATE sample_receipt
        SET prep_status='ready', prep_done_at=?, prep_notes=?,
            prep_operator=?, prep_position=?, prep_devices=?, prep_by=?
        WHERE id=?""", (datetime.now().isoformat(), notes,
                        prep_operator, prep_position, prep_devices,
                        session.get('user_id'), receipt_id))
    conn.execute("UPDATE geo_samples SET status='prepared' WHERE id=(SELECT geo_sample_id FROM sample_receipt WHERE id=?)", (receipt_id,))
    lab_num = conn.execute('SELECT lab_number FROM sample_receipt WHERE id=?', (receipt_id,)).fetchone()
    lab_str = lab_num[0] if lab_num else ''
    conn.commit(); conn.close()
    flash(f'✅ Дээж бэлтгэж дууслаа! Химичид шилжлээ. Ажлын дугаар: {lab_str}', 'success')
    return redirect(url_for('analysis'))

# ── SAMPLE TYPES ────────────────────────────────────────

@app.route('/sample-types', methods=['GET','POST'])
@admin_required
def sample_types():
    lang = session.get('lang','mn')
    conn = get_db()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            conn.execute("""
                INSERT INTO sample_types(code,name_mn,name_en,icon,color,serial_from,serial_to,is_pit,sort_order)
                VALUES(?,?,?,?,?,?,?,?,?)
            """, (
                request.form['code'].upper().strip(),
                request.form['name_mn'],
                request.form.get('name_en',''),
                request.form.get('icon','🧪'),
                request.form.get('color','#185fa5'),
                int(request.form.get('serial_from') or 7000),
                int(request.form.get('serial_to') or 7999),
                1 if request.form.get('is_pit') else 0,
                int(request.form.get('sort_order') or 99)
            ))
            flash('Дээжний төрөл нэмэгдлээ!', 'success')
        elif action == 'toggle':
            tid = request.form.get('id')
            conn.execute("UPDATE sample_types SET is_active=1-is_active WHERE id=?", (tid,))
            flash('Өөрчлөгдлөө!', 'success')
        elif action == 'delete':
            tid = request.form.get('id')
            conn.execute("DELETE FROM sample_types WHERE id=?", (tid,))
            flash('Устгагдлаа!', 'success')
        conn.commit()
        conn.close()
        return redirect(url_for('lab_settings') + '?tab=types')
    
    types = conn.execute("SELECT * FROM sample_types ORDER BY sort_order, serial_from").fetchall()
    conn.close()
    return jsonify([dict(t) for t in types])


@app.route('/analysis/row/done', methods=['POST'])
@lab_required
def row_done():
    """Химич мөрийг дууслаа гэж тэмдэглэнэ"""
    data = request.get_json()
    rid = data.get('receipt_id')
    row = data.get('row_num')
    dup = data.get('is_duplicate', 0)
    conn = get_db()
    # Мөр байгаа эсэх
    existing = conn.execute(
        "SELECT id FROM sample_entries WHERE receipt_id=? AND row_num=? AND is_duplicate=?",
        (rid, row, dup)
    ).fetchone()
    if existing:
        conn.execute("""UPDATE sample_entries SET row_status='done', done_by=?, done_at=?
                       WHERE receipt_id=? AND row_num=? AND is_duplicate=?""",
                    (session['user_id'], datetime.now().isoformat(), rid, row, dup))
    else:
        conn.execute("""INSERT INTO sample_entries(receipt_id,row_num,is_duplicate,row_status,done_by,done_at)
                       VALUES(?,?,?,'done',?,?)""",
                    (rid, row, dup, session['user_id'], datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ── АЖЛЫН БАГЦ ──────────────────────────────────────────
# Химич хэд хэдэн ажлыг нэг хуудсанд нэгтгэхэд тэр нэгтгэл бүртгэгдэнэ.
# Ингэснээр төхөөрөмж рүү ороод, refresh хийгээд буцаж ирэхэд дахин сонгож
# эрэмбэлэх шаардлагагүй — "Үргэлжлүүлэх" дарахад яг тэр дарааллаараа нээгдэнэ.

def batch_url(b):
    """Багцыг нээх хаяг — үүсгэсэн үеийн дараалал, QC параметр хэвээрээ"""
    url = url_for('analysis_measure_multi') + '?ids=' + b['receipt_ids']
    if b['qc_rows']:
        url += '&qc_rows=' + b['qc_rows']
    if b['qc_id']:
        url += '&qc_id=' + str(b['qc_id'])
    return url


def batch_receipt_ids(b):
    return [int(i) for i in (b['receipt_ids'] or '').split(',') if i.strip().isdigit()]


def open_batches(conn, uid=None):
    """Лабораторийн БҮХ идэвхтэй багц. Бүх мөр баталгаажсаныг өөрөө хаана.

    Урьд нь зөвхөн үүсгэсэн хүнд харагддаг байсан тул химич бүр ижил
    ажлууд дээр өөр өөрийн багц үүсгэдэг байв. Одоо нэг хүн (ихэвчлэн
    ахлах химич) үүсгэхэд бусад нь түүнийг хараад үргэлжлүүлнэ.
    """
    rows = conn.execute("""SELECT wb.*, u.name AS owner_name
                           FROM work_batch wb
                           LEFT JOIN users u ON u.id = wb.user_id
                           WHERE wb.status='open'
                           ORDER BY wb.created_at DESC""").fetchall()
    out = []
    for b in rows:
        rids = batch_receipt_ids(b)
        if not rids:
            continue
        q = ','.join('?' * len(rids))
        # Дуусаагүй мөр үлдсэн эсэх — үлдээгүй бол багц дуусжээ
        left = conn.execute(
            f"""SELECT COUNT(*) c FROM sample_entries
                WHERE receipt_id IN ({q}) AND is_duplicate=0
                  AND (row_status IS NULL OR row_status<>'approved')""", rids).fetchone()['c']
        total = conn.execute(
            f'SELECT COUNT(*) c FROM sample_entries WHERE receipt_id IN ({q})', rids
        ).fetchone()['c']
        if total and not left:
            conn.execute("UPDATE work_batch SET status='closed', closed_at=? WHERE id=?",
                         (datetime.now().isoformat(), b['id']))
            continue
        d = dict(b)
        d['labs'] = [r['lab_number'] for r in conn.execute(
            f'SELECT id, lab_number FROM sample_receipt WHERE id IN ({q})', rids)]
        # Хэрэглэгчийн сонгосон дарааллаар эрэмбэлнэ
        by_id = {r['id']: r['lab_number'] for r in conn.execute(
            f'SELECT id, lab_number FROM sample_receipt WHERE id IN ({q})', rids)}
        d['labs'] = [by_id[i] for i in rids if i in by_id]
        d['n_samples'] = conn.execute(
            f"""SELECT COALESCE(SUM(g.quantity),0) c FROM sample_receipt sr
                JOIN geo_samples g ON g.id=sr.geo_sample_id
                WHERE sr.id IN ({q})""", rids).fetchone()['c']
        d['url'] = batch_url(b)
        # Тэмдэглэл — геологич, хүлээн авагч, дээж бэлтгэгчийн бичсэн.
        # Химич багцаа үргэлжлүүлэхээсээ өмнө шинжилгээний хуудсан дээрээ
        # шууд уншина (урьд нь эдгээр хаана ч харагддаггүй байв).
        nrows = {r['id']: r for r in conn.execute(
            f"""SELECT sr.id, sr.lab_number, sr.notes, sr.prep_notes,
                       g.notes AS geo_notes, ug.name AS geo_name
                  FROM sample_receipt sr
                  JOIN geo_samples g ON g.id = sr.geo_sample_id
                  LEFT JOIN users ug ON ug.id = g.registered_by
                 WHERE sr.id IN ({q})""", rids)}
        notes = []
        for rid in rids:                       # багцын дараалал хэвээр
            r = nrows.get(rid)
            if not r:
                continue
            who_geo = 'Геологич'
            if r['geo_name']:
                who_geo += ' · ' + r['geo_name']
            for col, who in (('geo_notes', who_geo),
                             ('notes', 'Хүлээн авахад'),
                             ('prep_notes', 'Дээж бэлтгэл')):
                if r[col] and r[col].strip():
                    notes.append({'lab': r['lab_number'], 'who': who,
                                  'txt': r[col].strip()})
        d['notes'] = notes
        out.append(d)
    conn.commit()
    return out


@app.route('/analysis/batch/start', methods=['POST'])
@lab_required
def batch_start():
    """Нэгтгэл үүсгээд хэмжилтийн хуудас руу шилжинэ"""
    data = request.get_json() or {}
    ids = [str(int(i)) for i in (data.get('ids') or []) if str(i).strip().isdigit()]
    if not ids:
        return jsonify({'ok': False, 'error': 'Ажил сонгоогүй байна'}), 400
    uid = session.get('user_id')
    conn = get_db()
    # Ижил бүрэлдэхүүнтэй багц аль хэдийн нээлттэй бол давхардуулахгүй.
    # Багц нь бүх хүнд харагддаг тул ӨӨР хүний үүсгэсэн багцыг ч
    # дахин үүсгэхгүй — хоёр химич нэг ажлыг тусад нь эхлүүлэхээс сэргийлнэ.
    same = conn.execute("""SELECT id FROM work_batch
                           WHERE status='open' AND receipt_ids=?""",
                        (','.join(ids),)).fetchone()
    if not same:
        conn.execute("""INSERT INTO work_batch(user_id,receipt_ids,qc_rows,qc_id,created_at)
                        VALUES(?,?,?,?,?)""",
                     (uid, ','.join(ids), data.get('qc_rows') or None,
                      data.get('qc_id') or None, datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/analysis/batch/close/<int:bid>', methods=['POST'])
@lab_required
def batch_close(bid):
    """Багцыг хаана.

    Багц нь хамтын ажил тул лабораторийн хэн ч хааж болно — үүсгэсэн
    химич ээлжээ дуусгаад явсан бол багц өлгөөтэй үлдэхгүй. Хаах нь
    зөвхөн товчлолыг арилгана, өгөгдөлд огт нөлөөлөхгүй.
    """
    conn = get_db()
    b = conn.execute('SELECT * FROM work_batch WHERE id=?', (bid,)).fetchone()
    if not b:
        conn.close()
        return jsonify({'ok': False, 'error': 'Багц олдсонгүй'}), 404
    conn.execute("UPDATE work_batch SET status='closed', closed_at=? WHERE id=?",
                 (datetime.now().isoformat(), bid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/analysis/row/delete', methods=['POST'])
@lab_required
def row_delete():
    """Давталтын мөрийг устгана (андуурч нэмсэн эсвэл хуулагдсан үед).

    Зөвхөн ДАВТАЛТ (is_duplicate>=2) устгагдана — үндсэн ба зэрэгцээ мөр нь
    хүснэгтийн байнгын мөр тул устгахгүй, утгыг нь цэвэрлэж болно.
    Баталгаажсан мөрийг зөвхөн ахлах/админ устгана.
    """
    data = request.get_json() or {}
    rid = data.get('receipt_id')
    row = data.get('row_num')
    dup = data.get('is_duplicate')
    try:
        rid, row, dup = int(rid), int(row), int(dup)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'Буруу утга'}), 400
    if dup < 2:
        return jsonify({'ok': False,
                        'error': 'Зөвхөн давталтын мөрийг устгана'}), 400

    conn = get_db()
    ent = conn.execute("""SELECT id, row_status FROM sample_entries
                          WHERE receipt_id=? AND row_num=? AND is_duplicate=?""",
                       (rid, row, dup)).fetchone()
    if not ent:
        conn.close()
        return jsonify({'ok': False, 'error': 'Мөр олдсонгүй'}), 404
    if ent['row_status'] == 'approved' and session.get('role') not in ('admin', 'senior'):
        conn.close()
        return jsonify({'ok': False,
                        'error': 'Баталгаажсан мөрийг зөвхөн ахлах устгана'}), 403

    conn.execute('DELETE FROM sample_entries WHERE id=?', (ent['id'],))
    # Үлдсэн давталтуудыг залгуулж дугаарлана (давталт 1, 2… тасрахгүй байхын
    # тулд). UNIQUE(receipt_id,row_num,is_duplicate)-тай мөргөлдөхгүйн тулд
    # өсөх дарааллаар шилжүүлнэ — өмнөх дугаар нь тухай бүрдээ сул болно.
    rest = [r['is_duplicate'] for r in conn.execute(
        """SELECT is_duplicate FROM sample_entries
           WHERE receipt_id=? AND row_num=? AND is_duplicate>? ORDER BY is_duplicate""",
        (rid, row, dup))]
    for d in rest:
        conn.execute("""UPDATE sample_entries SET is_duplicate=?
                        WHERE receipt_id=? AND row_num=? AND is_duplicate=?""",
                     (d - 1, rid, row, d))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'shifted': len(rest)})


@app.route('/analysis/row/done-all', methods=['POST'])
@lab_required
def row_done_all():
    """Бүх мөрийг нэгэн зэрэг done/undo хийнэ"""
    data = request.get_json()
    rid = data.get('receipt_id')
    rows = data.get('rows', [])
    conn = get_db()
    for r in rows:
        row = r.get('row_num')
        dup = r.get('is_duplicate', 0)
        action = r.get('action', 'done')
        if action == 'done':
            existing = conn.execute(
                "SELECT id FROM sample_entries WHERE receipt_id=? AND row_num=? AND is_duplicate=?",
                (rid, row, dup)
            ).fetchone()
            if existing:
                conn.execute("""UPDATE sample_entries SET row_status='done', done_by=?, done_at=?
                               WHERE receipt_id=? AND row_num=? AND is_duplicate=?""",
                            (session['user_id'], datetime.now().isoformat(), rid, row, dup))
            else:
                conn.execute("""INSERT INTO sample_entries(receipt_id,row_num,is_duplicate,row_status,done_by,done_at)
                               VALUES(?,?,?,'done',?,?)""",
                            (rid, row, dup, session['user_id'], datetime.now().isoformat()))
        else:
            conn.execute("""UPDATE sample_entries SET row_status='empty', done_by=NULL, done_at=NULL
                           WHERE receipt_id=? AND row_num=? AND is_duplicate=?""",
                        (rid, row, dup))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/analysis/row/approve', methods=['POST'])
@perm_required('can_approve')
def row_approve():
    """Ахлах химич мөрийг баталгаажуулна"""
    data = request.get_json()
    rid = data.get('receipt_id')
    rows = data.get('rows', [])  # [{row_num, is_duplicate}]
    approve_all = data.get('approve_all', False)
    conn = get_db()
    if approve_all:
        conn.execute("""UPDATE sample_entries SET row_status='approved', approved_by=?, approved_at=?
                       WHERE receipt_id=? AND row_status='done'""",
                    (session['user_id'], datetime.now().isoformat(), rid))
    else:
        for r in rows:
            conn.execute("""UPDATE sample_entries SET row_status='approved', approved_by=?, approved_at=?
                           WHERE receipt_id=? AND row_num=? AND is_duplicate=?""",
                        (session['user_id'], datetime.now().isoformat(), rid, r['row_num'], r['is_duplicate']))
    conn.commit()
    # Бүх үндсэн мөр баталгаажсан эсэхийг шалгана
    total = conn.execute("SELECT COUNT(*) FROM sample_entries WHERE receipt_id=? AND is_duplicate=0", (rid,)).fetchone()[0]
    approved = conn.execute("SELECT COUNT(*) FROM sample_entries WHERE receipt_id=? AND is_duplicate=0 AND row_status='approved'", (rid,)).fetchone()[0]
    if total > 0 and total == approved:
        conn.execute("UPDATE geo_samples SET status='done' WHERE id=(SELECT geo_sample_id FROM sample_receipt WHERE id=?)", (rid,))
        conn.execute("UPDATE sample_receipt SET prep_status='done' WHERE id=?", (rid,))
        # Энэ receipt QC ажил байвал internal_qc.status='done' болгоно
        conn.execute("UPDATE internal_qc SET status='done' WHERE receipt_id_1=? AND status='pending'", (rid,))
        conn.commit()
    all_approved = (total > 0 and total == approved)
    conn.close()
    return jsonify({'ok': True, 'all_approved': all_approved})

def result_access_denied(conn, receipt_id):
    """Үр дүн харах эрхийг шалгана. Эрхгүй бол (мессеж, хаяг) буцаана.

    Баяжуулагч зөвхөн 6000-6999 муж + can_view_result, геологч нь өөрийн
    view_ranges мужийг л харна. Үр дүнгийн ба АРХИВЫН хуудас хоёуланд
    хэрэглэгдэнэ — урьд нь архивын хуудсанд шалгалт огт байхгүй тул эрхгүй
    хэрэглэгч /archive/result/<id>-ээр дамжин үр дүнг харж чаддаг байв.
    """
    role = session.get('role')
    uid = session.get('user_id')
    if role == 'bayjuulach':
        u = conn.execute('SELECT can_view_result FROM users WHERE id=?', (uid,)).fetchone()
        if not u or not u['can_view_result']:
            return 'Үр дүн харах эрх байхгүй байна', url_for('dashboard')
        sr = conn.execute('SELECT lab_serial FROM sample_receipt WHERE id=?',
                          (receipt_id,)).fetchone()
        if not sr or not (6000 <= (sr['lab_serial'] or 0) <= 6999):
            return 'Энэ ажлыг харах эрх байхгүй байна', url_for('dashboard')
    elif role == 'geologist':
        _u = conn.execute('SELECT view_ranges FROM users WHERE id=?', (uid,)).fetchone()
        _vr = _u['view_ranges'] if _u else None
        if _vr is not None:
            _th = [int(x) for x in _vr.split(',') if x.strip().isdigit()]
            _sr = conn.execute('SELECT lab_serial FROM sample_receipt WHERE id=?',
                               (receipt_id,)).fetchone()
            if ((_sr['lab_serial'] or 0) if _sr else 0) // 1000 not in _th:
                return 'Энэ ажлыг харах эрх байхгүй байна', url_for('analysis')
    return None


@app.route('/analysis/result/<int:receipt_id>')
@login_required
def analysis_result(receipt_id):
    """Үр дүнгийн хуудас — бүх эрхэд харагдана"""
    lang = session.get('lang','mn')
    role = session.get('role')
    conn = get_db()
    _deny = result_access_denied(conn, receipt_id)
    if _deny:
        conn.close()
        flash(_deny[0], 'error')
        return redirect(_deny[1])
    if role in ('bayjuulach', 'geologist'):
        conn.execute('INSERT INTO result_view_log(user_id, receipt_id) VALUES(?,?)',
                     (session['user_id'], receipt_id))
        conn.commit()
    receipt = conn.execute("""
        SELECT sr.*, g.sample_name, g.sample_type, g.location,
               g.collected_date, g.quantity,
               ug.name as geo_name, up.name as prep_name
        FROM sample_receipt sr
        JOIN geo_samples g ON g.id=sr.geo_sample_id
        LEFT JOIN users ug ON ug.id=g.registered_by
        LEFT JOIN users up ON up.id=sr.received_by
        WHERE sr.id=?
    """, (receipt_id,)).fetchone()
    
    entries = conn.execute("""
        SELECT se.*, u1.name as done_name, u2.name as approved_name
        FROM sample_entries se
        LEFT JOIN users u1 ON u1.id=se.done_by
        LEFT JOIN users u2 ON u2.id=se.approved_by
        WHERE se.receipt_id=?
        ORDER BY se.row_num, se.is_duplicate
    """, (receipt_id,)).fetchall()

    crm_cert = None
    if receipt and receipt['sample_type'] == 'CRM':
        geo = conn.execute("""SELECT crm_mad, crm_aad, crm_vad, crm_sulfur, crm_cal, crm_g,
                                     crm_mad_unc, crm_aad_unc, crm_vad_unc, crm_sulfur_unc,
                                     crm_cal_unc, crm_g_unc
                              FROM geo_samples WHERE id=?""",
                           (receipt['geo_sample_id'],)).fetchone()
        crm_cert = dict(geo) if geo else None

    qc_tols = qc_tolerances(conn)   # холболт хаагдахаас өмнө уншина
    conn.close()

    # mt_result тооцоолох + display_name үүсгэх
    import re as _re
    _sname = receipt['sample_name'] or '' if receipt else ''
    _sparts = [s.strip() for s in _sname.split(';')] if ';' in _sname else None
    _m1 = _re.match(r'^(.*?)(\d+)\s*[-–]\s*(\d+)\s*$', _sname)
    _m2 = _re.match(r'^(.*?)(\d+)$', _sname) if not _m1 else None
    entries_list = []
    for e in entries:
        ed = dict(e)
        ed['mt_result'] = total_moisture(ed)
        # G индекс: гараар оруулаагүй бол жингээс бодно. Урьд нь зөвхөн g_val-ыг
        # уншдаг тул жингээр оруулсан дээж дэлгэц дээр "—" харагдаж, Excel-д
        # тоо гарч зөрдөг байсан.
        ed['g_val'] = lab_g_index(ed)
        rn = ed.get('row_num', 1)
        en = ed.get('sample_name') or ''
        if en and en != _sname:
            ed['display_name'] = en
        elif _sparts and len(_sparts) >= rn:
            ed['display_name'] = _sparts[rn - 1]
        elif _m1:
            ed['display_name'] = _m1.group(1) + str(int(_m1.group(2)) + rn - 1)
        elif _m2:
            ed['display_name'] = _m2.group(1) + str(int(_m2.group(2)) + rn - 1)
        else:
            ed['display_name'] = _sname
        entries_list.append(ed)
    apply_final_results(entries_list, qc_tols)   # давталтаас эцсийн үр дүнг сонгоно
    # Мөр бүрийн нэр (entry байхгүй мөрөнд fallback болгон ашиглана)
    _qty = (receipt['quantity'] or 1) if receipt else 1
    row_names = []
    for _rn in range(1, _qty + 1):
        if _sparts and len(_sparts) >= _rn:
            row_names.append(_sparts[_rn - 1])
        elif _m1:
            row_names.append(_m1.group(1) + str(int(_m1.group(2)) + _rn - 1))
        elif _m2:
            row_names.append(_m2.group(1) + str(int(_m2.group(2)) + _rn - 1))
        else:
            row_names.append(_sname)

    qc_row = request.args.get('qc_row', type=int)
    qc_id = request.args.get('qc_id', type=int)
    if qc_row:
        entries_list = [e for e in entries_list if e['row_num'] == qc_row]
    pending_approve = [e for e in entries_list if e['row_status'] == 'done' and e['is_duplicate'] == 0]
    return render_template('analysis/result.html',
        receipt=receipt, entries=entries_list, lang=lang, role=role, crm_cert=crm_cert,
        pending_approve=pending_approve, qc_row=qc_row, qc_id=qc_id, row_names=row_names)


def _restore_template_images(tmpl_path, buf):
    """openpyxl нь Pillow суугаагүй (эсвэл хуучин хувилбартай) орчинд template-ийн
    зургуудыг (logo) алдуулдаг. Гаралтын sheet1-д drawing холбоос байхгүй бол
    эх template-ийн drawing + media хэсгүүдийг zip түвшинд буцааж хуулна."""
    import zipfile, io, re
    import xml.etree.ElementTree as ET

    buf.seek(0)
    zout = zipfile.ZipFile(buf, 'r')
    try:
        s1 = zout.read('xl/worksheets/sheet1.xml').decode('utf-8')
    except KeyError:
        zout.close(); buf.seek(0); return buf
    if '<drawing ' in s1 or '<drawing/' in s1:
        zout.close(); buf.seek(0); return buf  # зургууд хадгалагдсан байна

    REL_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
    CT_NS = 'http://schemas.openxmlformats.org/package/2006/content-types'
    ztmpl = zipfile.ZipFile(tmpl_path, 'r')

    # Template sheet1 → drawing хэсгийг олох
    try:
        trels = ET.fromstring(ztmpl.read('xl/worksheets/_rels/sheet1.xml.rels'))
    except KeyError:
        ztmpl.close(); zout.close(); buf.seek(0); return buf
    drawing_part = None
    for rel in trels.findall(f'{{{REL_NS}}}Relationship'):
        if rel.get('Type', '').endswith('/drawing'):
            drawing_part = 'xl/' + rel.get('Target', '').replace('../', '')
            break
    if not drawing_part:
        ztmpl.close(); zout.close(); buf.seek(0); return buf

    drawing_xml = ztmpl.read(drawing_part)
    drawing_rels_part = 'xl/drawings/_rels/' + drawing_part.rsplit('/', 1)[-1] + '.rels'
    drawing_rels = ztmpl.read(drawing_rels_part).decode('utf-8')

    # Drawing-д хэрэглэгддэг media файлууд (нэрийг нь tmpl_ prefix-тэй болгож мөргөлдөөнөөс сэргийлнэ)
    media_map = {}  # эх нэр → шинэ нэр
    for tgt in re.findall(r'Target="(\.\./media/[^"]+)"', drawing_rels):
        old = 'xl/' + tgt.replace('../', '')
        new = 'xl/media/tmpl_' + old.rsplit('/', 1)[-1]
        media_map[old] = new
        drawing_rels = drawing_rels.replace(tgt, '../media/tmpl_' + old.rsplit('/', 1)[-1])

    new_drawing_part = 'xl/drawings/drawingTmplLogo.xml'
    new_drawing_rels_part = 'xl/drawings/_rels/drawingTmplLogo.xml.rels'
    rel_id = 'rIdTmplLogo'

    # sheet1.xml-д <drawing/> элемент нэмэх (schema дараалал: tableParts/extLst-ээс өмнө)
    ins = len(s1)
    for anchor in ('<tableParts', '<extLst', '</worksheet>'):
        i = s1.find(anchor)
        if i != -1:
            ins = min(ins, i)
    s1 = (s1[:ins]
          + f'<drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="{rel_id}"/>'
          + s1[ins:])

    # sheet1-ийн rels-д drawing relationship нэмэх (байхгүй бол шинээр үүсгэнэ)
    sheet_rels_part = 'xl/worksheets/_rels/sheet1.xml.rels'
    rel_elem = (f'<Relationship Id="{rel_id}" '
                f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
                f'Target="../drawings/drawingTmplLogo.xml"/>')
    try:
        srels = zout.read(sheet_rels_part).decode('utf-8')
        srels = srels.replace('</Relationships>', rel_elem + '</Relationships>')
    except KeyError:
        srels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                 f'<Relationships xmlns="{REL_NS}">' + rel_elem + '</Relationships>')

    # [Content_Types].xml — зургийн өргөтгөлүүд ба drawing override нэмэх
    ct = zout.read('[Content_Types].xml').decode('utf-8')
    ct_add = ''
    ext_types = {'jpeg': 'image/jpeg', 'jpg': 'image/jpeg', 'png': 'image/png',
                 'emf': 'image/x-emf', 'wdp': 'image/vnd.ms-photo'}
    for new in media_map.values():
        ext = new.rsplit('.', 1)[-1].lower()
        if ext in ext_types and f'Extension="{ext}"' not in ct + ct_add:
            ct_add += f'<Default Extension="{ext}" ContentType="{ext_types[ext]}"/>'
    ct_add += (f'<Override PartName="/{new_drawing_part}" '
               f'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')
    ct = ct.replace('</Types>', ct_add + '</Types>')

    # Шинэ zip угсрах
    out = io.BytesIO()
    replaced = {'xl/worksheets/sheet1.xml': s1.encode('utf-8'),
                sheet_rels_part: srels.encode('utf-8'),
                '[Content_Types].xml': ct.encode('utf-8')}
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as znew:
        for item in zout.infolist():
            data = replaced.pop(item.filename, None) or zout.read(item.filename)
            znew.writestr(item, data)
        for name, data in replaced.items():
            znew.writestr(name, data)
        znew.writestr(new_drawing_part, drawing_xml)
        znew.writestr(new_drawing_rels_part, drawing_rels.encode('utf-8'))
        for old, new in media_map.items():
            znew.writestr(new, ztmpl.read(old))
    ztmpl.close(); zout.close()
    out.seek(0)
    return out


@app.route('/analysis/export/<int:receipt_id>')
@login_required
def analysis_export(receipt_id):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    import io, os
    # Харилцагч эрх шалгах
    role = session.get('role')
    if role == 'guest':
        # Бусад экспортын маршруттай ижил — зочин файл татахгүй
        flash('Зочин горимд Excel татах боломжгүй.', 'error')
        return redirect(url_for('analysis_result', receipt_id=receipt_id))
    if role in ('geologist', 'bayjuulach'):
        conn_chk = get_db()
        u = conn_chk.execute("SELECT can_export FROM users WHERE id=?", (session['user_id'],)).fetchone()
        conn_chk.close()
        if not u or not u['can_export']:
            flash('Excel татах эрх байхгүй байна', 'error')
            return redirect(url_for('analysis_result', receipt_id=receipt_id))

    conn = get_db()
    receipt = conn.execute("""
        SELECT sr.*, g.sample_name, g.sample_type, g.location,
               g.collected_date, g.quantity,
               ug.name as geo_name, up.name as prep_name,
               uf.name as fm_op_name, um.name as mt_op_name
        FROM sample_receipt sr
        JOIN geo_samples g ON g.id=sr.geo_sample_id
        LEFT JOIN users ug ON ug.id=g.registered_by
        LEFT JOIN users up ON up.id=sr.received_by
        LEFT JOIN users uf ON uf.id=sr.fm_operator
        LEFT JOIN users um ON um.id=sr.mt_operator
        WHERE sr.id=?
    """, (receipt_id,)).fetchone()
    entries = conn.execute("""
        SELECT * FROM sample_entries WHERE receipt_id=? ORDER BY row_num, is_duplicate
    """, (receipt_id,)).fetchall()
    # ── Гарын үсгийн блок: ЖИНХЭНЭ гүйцэтгэгчид ────────────────────────
    # Урьд нь зөвхөн done_by (мөрийг "дууслаа" гэж дарсан хүн) уншигддаг
    # байсан тул шинжилгээг үнэхээр хийсэн химичүүд албан тайланд ОРОХГҮЙ,
    # оронд нь ✓ дарсан ганц хүн (заримдаа админ) бичигддэг байв.
    # Одоо шинжилгээ тус бүрийн гүйцэтгэгчээс (op_mt, op_mad …) цуглуулна.
    _op_cols = [c for c, _l, _f in ANALYSIS_OPS]
    _who = {}
    for _c in _op_cols:
        for r in conn.execute(f"""SELECT DISTINCT u.name, u.role
                                    FROM sample_entries se JOIN users u ON u.id=se.{_c}
                                   WHERE se.receipt_id=? AND se.{_c} IS NOT NULL""",
                              (receipt_id,)):
            _who[r['name']] = r['role']
    if not _who:      # хуучин бичлэгт гүйцэтгэгч тэмдэглэгдээгүй бол
        for r in conn.execute("""SELECT DISTINCT u.name, u.role
                                   FROM sample_entries se JOIN users u ON u.id=se.done_by
                                  WHERE se.receipt_id=? AND se.done_by IS NOT NULL""",
                              (receipt_id,)):
            _who[r['name']] = r['role']
    # Химич ба ахлах химичийг албан тушаалаар нь салгана
    chemist_names = sorted(n for n, ro in _who.items() if ro not in ('senior', 'admin'))
    senior_names  = sorted(n for n, ro in _who.items() if ro in ('senior', 'admin'))
    approver_names = [r['name'] for r in conn.execute("""
        SELECT DISTINCT u.name FROM sample_entries se JOIN users u ON u.id=se.approved_by
        WHERE se.receipt_id=? AND se.approved_by IS NOT NULL ORDER BY u.name""", (receipt_id,)).fetchall()]
    qc_tols = qc_tolerances(conn)   # холболт хаагдахаас өмнө уншина
    conn.close()

    # Template ачаалах
    tmpl = os.path.join(os.path.dirname(__file__), 'static', 'report_template.xlsx')
    wb = openpyxl.load_workbook(tmpl)
    ws = wb.worksheets[0]

    # Лого дахин нэмэх (template доторх wmf лого openpyxl-д хасагддаг тул)
    try:
        from openpyxl.drawing.image import Image as _XLImage
        # Тохиргооны логыг ашиглана — урьд нь logo.png шууд бичигдсэн байсан тул
        # өөр лаборатори өөрийн логог ачаалсан ч Excel дээр хуучин лого гардаг байв.
        _logo_path = logo_path()
        if _logo_path:
            _img = _XLImage(_logo_path)
            _img.width = 80
            _img.height = 115
            ws.add_image(_img, 'A1')
    except Exception:
        pass

    # ── Мэдээлэл нөхөх ──────────────────────────────────────
    TYPE_DISPLAY = {
        'PIT':'Уурхай\nPIT','STOCKPILE':'Овоолго\nStockpile','EXPORT':'Экспорт\nExport',
        'CONTROL':'Гааль\nControl','EQ_CONTROL':'Гадаад хяналт\nEQ control','DP':'Баяжуулах\nDP',
    }

    # A9: гарчиг template-ийн дагуу (дугаар нь C12-т бичигдэнэ)

    # C12: лаб дугаар
    ws['C12'] = receipt['lab_number']

    # H12: хүлээн авсан огноо
    ws['H12'] = receipt['received_date'] or ''

    # C13: дээжний төрөл
    ws['C13'] = TYPE_DISPLAY.get(receipt['sample_type'], receipt['sample_type'])

    # H13: шинжилгээ хийсэн огноо.
    # Урьд нь prep_done_at (ДЭЭЖ БЭЛТГЭЛ дууссан огноо) бичигддэг тул
    # "DATE SAMPLES ANALYSED" нь хүлээн авсан огноотой ижил гарч байв.
    # Одоо мөрүүдийн шинжилгээ дууссан хамгийн сүүлийн огноог авна.
    _an = conn2_analysed_date(entries) or receipt['prep_done_at'] or ''
    if _an and 'T' in str(_an):
        _an = str(_an).split('T')[0]
    ws['H13'] = str(_an)[:10]

    # H14: тайлан хэвлэсэн огноо
    ws['H14'] = datetime.now().strftime('%Y-%m-%d')

    # ── Өгөгдлийн мөрүүд (20-р мөрөөс) ─────────────────────
    _row_names, _, _ = sample_names_for(receipt)
    _rows = [dict(e) for e in entries]
    row_map = {(e['row_num'], e['is_duplicate']): e for e in _rows}

    def safe(e, field, dec=None):
        try:
            v = e[field]
            if v is None:
                return None
            return round(float(v), dec) if dec is not None else v
        except Exception:
            return None

    def calc_mt(e):
        # Үр дүнгийн хуудастай нэг ижил функц (зөрөх боломжгүй болгов)
        return total_moisture(e) if e else None

    def calc_g(e):
        try:
            if e['g_val'] is not None:
                return float(e['g_val'])
            gc = safe(e,'g_coke'); gt = safe(e,'g_tare')
            gs1 = safe(e,'g_sieve1'); gs2 = safe(e,'g_sieve2')
            if gc is not None and gt is not None and gs1 is not None and gs2 is not None:
                d = gc - gt
                if d > 0:
                    return 10 + (30 * (gs1 - gt) + 70 * (gs2 - gt)) / d
        except Exception:
            pass
        return None

    # Давталттай үед эцсийн үр дүнг сонгоно — үр дүнгийн хуудастай ижил дүрэм
    # (хамгийн ойрхон хоёрын дундаж). Mt-г энэ файлын өөрийн томьёогоор бодно.
    for _r in _rows:
        _r['mt_result'] = calc_mt(_r)
    apply_final_results(_rows, qc_tols)

    # Тоон формат — албан тайлангийн жишиг файлтай ижил (утга бүрэн нарийвчлалтай
    # хадгалагдаж, харагдац нь форматаар зохицуулагдана)
    COL_FMT = {3: '0.0', 4: '0.0', 5: '0.00', 6: '0.00', 7: '0.00', 8: '0.00',
               9: '0.00', 10: '0.00', 11: '0.00', 12: '0.00', 13: '0.00',
               14: '0', 15: '0', 16: '0', 17: '0.0'}

    data_row = 20
    for ri in range(1, (receipt['quantity'] or 1) + 1):
        e  = row_map.get((ri, 0))
        de = row_map.get((ri, 1))

        # Тооцоолсон үзүүлэлтүүд — үр дүнгийн хуудасны (result.html) томьёотой ижил
        mt  = safe(e, 'mt_result')   # давталтаас сонгогдсон эцсийн Mt
        mad = safe(e,'mad'); aad = safe(e,'aad'); vad = safe(e,'vad')
        sulfur = safe(e,'sulfur'); cal = safe(e,'cal_value')
        adb = vdb = vdaf = sdb = qb_ad = qnet_ar = None
        if mad is not None and (100 - mad) > 0:
            if aad:    adb = aad * 100 / (100 - mad)
            if vad:    vdb = vad * 100 / (100 - mad)
            if sulfur: sdb = sulfur * 100 / (100 - mad)
            if vad and aad is not None and (100 - mad - aad) > 0:
                vdaf = vad * 100 / (100 - mad - aad)
        if cal:
            qb_ad = cal / 4.1868
            if (sulfur is not None and mad is not None and aad is not None
                    and vad is not None and mt is not None and (100 - mad) > 0):
                qgr_ad_jg = cal - (sulfur * 94.1 + cal * 0.0016)
                _vdaf = vad * 100 / (100 - mad - aad) if (100 - mad - aad) > 0 else 0
                _adb = aad * 100 / (100 - mad)
                hdaf = 2.888 + 0.393 * (_vdaf ** 0.5) - 0.0023 * _adb
                had = hdaf * (100 - mad - aad) / 100
                qnet_ar = ((qgr_ad_jg - 206 * had) * ((100 - mt) / (100 - mad))
                           - 23 * mt) / 4.1868

        # Дээжийн нэр: sample_entries-д NULL байвал geo_samples-ийн жагсаалт
        # /муж-аас сэргээнэ. Урьд нь "Дээж 1, Дээж 2…" гэсэн утгагүй нэр
        # албан тайланд бичигддэг байв.
        _nm = safe(e, 'sample_name') or (
            _row_names[ri - 1] if len(_row_names) >= ri else None) or f'Дээж {ri}'
        vals = {1: ri, 2: _nm, 3: safe(e,'mass_kg'),
                4: mt, 5: safe(e,'mad'), 6: safe(e,'aad'), 7: adb, 8: safe(e,'vad'),
                9: vdb, 10: vdaf, 11: safe(e,'fc'), 12: safe(e,'sulfur'), 13: sdb,
                14: qb_ad, 15: qnet_ar, 16: calc_g(e), 17: safe(e,'fsi')}
        for col, v in vals.items():
            cell = ws.cell(data_row, col, v)
            if col in COL_FMT:
                cell.number_format = COL_FMT[col]
        data_row += 1

    # ── Гарын үсгийн блок (template мөр 196-199, дараа нь дээш шилжинэ) ────
    ws['B196'] = 'Шинжилгээ хийсэн: Химич\n/Analysed: Chemist/'
    for i, name in enumerate(chemist_names[:3]):
        ws.cell(196 + i, 7, f'/{name}/')
    ws['K196'] = 'Дээж бэлтгэсэн: Дээж бэлтгэгч\n/Sample prepared: Sample preparer/'
    preparers = [p for p in dict.fromkeys(
        [receipt['prep_operator'], receipt['fm_op_name'], receipt['mt_op_name']]) if p]
    for i, name in enumerate(preparers[:2]):
        ws.cell(196 + i, 16, f'/{name}/')
    ws['B199'] = 'Шинжилгээ хийсэн: Ахлах химич\n/Analysed: Senior Chemist/'
    # Ахлах химичийн нэр урьд нь ХААНА ч бичигддэггүй байсан тул тэр мөр
    # хоосон үлдэж, баруун талын "Хянсан"-ы нэр ахлах химич мэт харагддаг байв.
    for i, name in enumerate(senior_names[:2]):
        ws.cell(199 + i, 7, f'/{name}/')
    ws['J199'] = 'Хянсан: Лаборатори хариуцсан ахлах мэргэжилтэн\n/Checked: Senior Laboratory Specialist/ '
    if approver_names:
        ws['P199'] = f'/{approver_names[0]}/'

    # ── Ашиглагдаагүй мөрүүдийг устгах (template-д 20..189 = 170 мөр байдаг) ──
    n_rows = receipt['quantity'] or 1
    TMPL_ROWS = 170
    if n_rows < TMPL_ROWS:
        start = 20 + n_rows          # эхний устгах мөр
        count = TMPL_ROWS - n_rows
        cut_end = start + count      # footer-ийн хуучин эхлэл (мөр 190)
        # openpyxl delete_rows нь merged муж, мөрийн өндрийг шилжүүлдэггүй —
        # гараар зөөнө
        below = []
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= cut_end:
                below.append((rng.min_col, rng.min_row, rng.max_col, rng.max_row))
                ws.unmerge_cells(str(rng))
            elif rng.min_row >= start:
                ws.unmerge_cells(str(rng))
        heights = {r: d.height for r, d in list(ws.row_dimensions.items())
                   if r >= cut_end and d.height is not None}
        for r in [r for r in list(ws.row_dimensions) if r >= start]:
            del ws.row_dimensions[r]
        ws.delete_rows(start, count)
        for c1, r1, c2, r2 in below:
            ws.merge_cells(start_row=r1 - count, start_column=c1,
                           end_row=r2 - count, end_column=c2)
        for r, h in heights.items():
            ws.row_dimensions[r - count].height = h
        ws.print_area = f'A1:Q{200 - count}'


    output = io.BytesIO()
    wb.save(output)
    output = _restore_template_images(tmpl, output)
    output.seek(0)

    fname = f"result_{receipt['lab_number']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


# ── DEVICE USAGE ─────────────────────────────────────────

@app.route('/analysis/device-map', methods=['GET','POST'])
@admin_required
def device_map():
    """Шинжилгээний төрөл → тоног холбоос тохируулах"""
    conn = get_db()
    if request.method == 'POST':
        analysis_type = request.form.get('analysis_type')
        device_ids = request.form.getlist('device_ids[]')
        conn.execute("DELETE FROM analysis_device_map WHERE analysis_type=?", (analysis_type,))
        for did in device_ids:
            if did:
                conn.execute("""INSERT INTO analysis_device_map(analysis_type, device_id, is_active, updated_by, updated_at)
                               VALUES(?,?,1,?,?)""",
                            (analysis_type, did, session['user_id'], datetime.now().isoformat()))
        conn.commit()
        conn.close()
        return jsonify({'ok': True})

    devices = conn.execute("SELECT * FROM devices WHERE status='active' AND (stage='analysis' OR stage='both' OR stage IS NULL) ORDER BY name").fetchall()
    mapping = {}
    for r in conn.execute("SELECT * FROM analysis_device_map WHERE is_active=1").fetchall():
        mapping.setdefault(r['analysis_type'], []).append(r['device_id'])
    conn.close()
    return jsonify({'devices': [dict(d) for d in devices], 'mapping': mapping})

@app.route('/analysis/device-usage/start', methods=['POST'])
@login_required
def device_usage_start():
    """Химич шинжилгээ эхлэхэд тоногийн ашиглалт бүртгэнэ"""
    data = request.get_json()
    device_id = data.get('device_id')
    receipt_id = data.get('receipt_id')
    analysis_type = data.get('analysis_type')
    if not device_id:
        return jsonify({'ok': False})
    conn = get_db()
    # Хэрэв аль хэдийн нээлттэй байвал хаах
    conn.execute("""UPDATE device_usage_log 
                   SET ended_at=?, duration_min=ROUND((JULIANDAY(?)-JULIANDAY(started_at))*1440,1)
                   WHERE user_id=? AND device_id=? AND ended_at IS NULL""",
                (datetime.now().isoformat(), datetime.now().isoformat(),
                 session['user_id'], device_id))
    # Шинэ нээх
    cur = conn.execute("""INSERT INTO device_usage_log(device_id, user_id, receipt_id, analysis_type, started_at)
                          VALUES(?,?,?,?,?)""",
                      (device_id, session['user_id'], receipt_id, analysis_type, 
                       datetime.now().isoformat()))
    log_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'log_id': log_id})

@app.route('/analysis/device-usage/end', methods=['POST'])
@login_required  
def device_usage_end():
    """Химич дуусгахад тоногийн ашиглалт хаана"""
    data = request.get_json()
    device_id = data.get('device_id')
    sample_count = data.get('sample_count', 0)
    now = datetime.now().isoformat()
    conn = get_db()
    conn.execute("""UPDATE device_usage_log 
                   SET ended_at=?, sample_count=?,
                       duration_min=ROUND((JULIANDAY(?)-JULIANDAY(started_at))*1440,1)
                   WHERE user_id=? AND device_id=? AND ended_at IS NULL""",
                (now, sample_count, now, session['user_id'], device_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/analysis/device-usage/end-session', methods=['POST'])
@login_required
def device_usage_end_session():
    """Гарах үед бүх нээлттэй ашиглалтыг хаана"""
    now = datetime.now().isoformat()
    conn = get_db()
    conn.execute("""UPDATE device_usage_log
                   SET ended_at=?,
                       duration_min=ROUND((JULIANDAY(?)-JULIANDAY(started_at))*1440,1)
                   WHERE user_id=? AND ended_at IS NULL""",
                (now, now, session['user_id']))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/analysis/measure/multi', methods=['GET','POST'])
@lab_required
def analysis_measure_multi():
    """Олон lot-ыг нэг хуудсанд нэгтгэх"""
    lang = session.get('lang','mn')
    conn = get_db()
    
    if request.method == 'POST':
        # Сонгосон receipt_ids авах
        ids = request.form.getlist('receipt_ids')
        if not ids:
            flash('Дор хаяж нэг ажлын дугаар сонгоно уу!', 'error')
            return redirect(url_for('analysis'))
        return redirect(url_for('analysis_measure_multi') + '?ids=' + ','.join(ids))
    
    # GET — ids параметрээс авах
    ids_str = request.args.get('ids','')
    if not ids_str:
        return redirect(url_for('analysis'))

    ids = [int(i) for i in ids_str.split(',') if i.strip().isdigit()]

    # qc_rows параметр: "rid1:rn1,rid2:rn2" — зөвхөн тодорхой мөрийг харуулах
    qc_rows_str = request.args.get('qc_rows','')
    qc_row_map = {}  # {receipt_id: row_num}
    if qc_rows_str:
        for part in qc_rows_str.split(','):
            if ':' in part:
                r, n = part.split(':', 1)
                if r.strip().isdigit() and n.strip().isdigit():
                    qc_row_map[int(r.strip())] = int(n.strip())

    receipts = []
    for rid in ids:
        r = conn.execute("""
            SELECT sr.*, g.sample_name, g.sample_type, g.location,
                   g.collected_date, g.quantity,
                   g.notes as geo_notes,
                   ug.name as geo_name, up.name as prep_name
            FROM sample_receipt sr
            JOIN geo_samples g ON g.id=sr.geo_sample_id
            LEFT JOIN users ug ON ug.id=g.registered_by
            LEFT JOIN users up ON up.id=sr.received_by
            WHERE sr.id=?
        """, (rid,)).fetchone()
        if r:
            receipts.append(r)

    conn.close()
    if not receipts:
        return redirect(url_for('analysis'))

    qc_receipt_ids = [str(k) for k in qc_row_map.keys()]
    qc_id = request.args.get('qc_id', type=int)
    # QC хүлцэл — measure.html-тэй ижил эх сурвалжаас. Урьд нь энэ хуудсанд
    # утга нь кодод бэхлэгдсэн байсан тул тохиргооны өөрчлөлт хүчин төгөлдөр
    # болдоггүй, зөвхөн Mad/Aad/Vad шалгагддаг байв.
    conn2 = get_db()
    qc_map = {r['parameter']: r['tolerance']
              for r in conn2.execute("SELECT parameter, tolerance FROM qc_settings")}
    conn2.close()
    return render_template('analysis/measure_multi.html',
        receipts=receipts, lang=lang, qc_map=qc_map,
        ids=ids_str, qc_row_map=qc_row_map, qc_receipt_ids=qc_receipt_ids, qc_id=qc_id)

# ── LAB SETTINGS ────────────────────────────────────────
import json as _json

SETTINGS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance', 'settings.json')

def get_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, encoding='utf-8') as f:
            return _json.load(f)
    # Шинэ суулгацын анхдагч — саармаг лого. Байгаа суулгацууд settings.json
    # дотроо өөрийн логог зааж өгсөн байдаг тул энэ нь тэдэнд нөлөөлөхгүй.
    return {'lab_name':'Лабораторийн нэр','lab_name_en':'Laboratory',
            'lab_subtitle':'Лабораторийн удирдлагын систем','logo':'logo-default.png'}


def logo_path():
    """Тохиргоонд заасан логоны файлын бүтэн зам (Excel тайланд хэрэглэнэ)."""
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')
    for name in (get_settings().get('logo'), 'logo-default.png'):
        if name:
            p = os.path.join(base, name)
            if os.path.exists(p):
                return p
    return None

def save_settings(data):
    os.makedirs(os.path.dirname(SETTINGS_FILE), exist_ok=True)
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        _json.dump(data, f, ensure_ascii=False, indent=2)

@app.context_processor
def inject_settings():
    return dict(settings=get_settings(), weak_admin=WEAK_ADMIN_PASSWORD,
                weak_users=WEAK_PASSWORD_USERS, app_version=VERSION)


# Хуучин суулгацуудад анхдагч 'admin123' хэвээр үлдсэн байж болно.
# Эхлэхэд нэг удаа шалгаад, админд анхааруулга харуулна.
WEAK_ADMIN_PASSWORD = False
WEAK_PASSWORD_USERS = []          # ['Нэр (АЖИЛТНЫ ДУГААР)', ...]


def _check_weak_admin():
    """Анхдагч admin123 нууц үгтэй ИДЭВХТЭЙ хэрэглэгчдийг олно.

    Урьд нь зөвхөн admin дүртэй хүнийг шалгаж, хэн болохыг нь хэлдэггүй
    байсан тул "сольсон боловч анхааруулга арилахгүй" гэсэн байдалд
    ордог байв — үнэндээ өөр хэрэглэгч дээр үлдсэн байдаг.
    """
    global WEAK_ADMIN_PASSWORD, WEAK_PASSWORD_USERS
    try:
        conn = get_db()
        rows = conn.execute("""SELECT employee_id, name, role, password_hash
                               FROM users WHERE is_active=1""").fetchall()
        conn.close()
        WEAK_PASSWORD_USERS = [
            f"{r['name']} ({r['employee_id']})" for r in rows
            if r['password_hash'] and check_password(r['password_hash'], 'admin123')]
        WEAK_ADMIN_PASSWORD = bool(WEAK_PASSWORD_USERS)
        if WEAK_ADMIN_PASSWORD:
            app.logger.warning('Анхдагч admin123 нууц үгтэй хэрэглэгч: %s',
                               ', '.join(WEAK_PASSWORD_USERS))
    except Exception:
        app.logger.exception('Нууц үгийн шалгалт амжилтгүй')

@app.route('/lab-settings', methods=['GET','POST'])
@admin_required
def lab_settings():
    lang = session.get('lang','mn')
    s = get_settings()
    conn = get_db()
    qc_settings_list = conn.execute("SELECT * FROM qc_settings ORDER BY parameter").fetchall()

    if request.method == 'POST':
        action = request.form.get('action','lab_info')

        if action == 'lab_info':
            s['lab_name']     = request.form.get('lab_name', s['lab_name'])
            s['lab_name_en']  = request.form.get('lab_name_en', s['lab_name_en'])
            s['lab_subtitle'] = request.form.get('lab_subtitle', s['lab_subtitle'])
            logo = request.files.get('logo')
            if logo and logo.filename:
                ext = logo.filename.rsplit('.',1)[-1].lower()
                if ext in ('png','jpg','jpeg','webp'):
                    # Гит-д бүртгэлтэй static/logo.* файлыг дарж бичихгүй —
                    # эс бөгөөс тухайн сервер дээр git pull "local changes
                    # would be overwritten" гэж зогсдог. uploads нь gitignore-т.
                    updir = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                         'static', 'uploads')
                    os.makedirs(updir, exist_ok=True)
                    logo.save(os.path.join(updir, f'logo.{ext}'))
                    s['logo'] = f'uploads/logo.{ext}'
            save_settings(s)
            flash('Лабораторийн мэдээлэл хадгалагдлаа!', 'success')

        elif action == 'qc_settings':
            for q in qc_settings_list:
                key = f"tol_{q['parameter']}"
                val = request.form.get(key)
                if val:
                    conn.execute("UPDATE qc_settings SET tolerance=?, updated_by=? WHERE parameter=?",
                                (float(val), session['user_id'], q['parameter']))
            conn.commit()
            flash('QC тохиргоо хадгалагдлаа!', 'success')

        elif action == 'update_profile':
            photo = save_file(request.files.get('photo'), 'staff')
            conn.execute("UPDATE users SET position=?, phone=?, email=? WHERE id=?", (
                request.form.get('position'),
                request.form.get('phone'),
                request.form.get('email'),
                session['user_id']
            ))
            if photo:
                conn.execute("UPDATE users SET photo=? WHERE id=?", (photo, session['user_id']))
            conn.commit()
            flash('Профайл шинэчлэгдлээ!', 'success')

        elif action == 'change_password':
            user = conn.execute("SELECT * FROM users WHERE id=?", (session.get('user_id', 0),)).fetchone()
            old_pw = request.form.get('old_password','')
            new_pw = request.form.get('new_password','')
            confirm = request.form.get('confirm_password','')
            if not check_password(user['password_hash'], old_pw):
                flash('Хуучин нууц үг буруу!', 'error')
            elif new_pw != confirm:
                flash('Шинэ нууц үг таарахгүй байна!', 'error')
            elif len(new_pw) < 6:
                flash('Нууц үг хамгийн багадаа 6 тэмдэгт!', 'error')
            else:
                conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                           (hash_password(new_pw), session['user_id']))
                conn.commit()
                _check_weak_admin()
                flash('Нууц үг амжилттай солигдлоо!', 'success')

        conn.close()
        return redirect(url_for('lab_settings'))

    conn.close()
    _crm_conn = get_db()
    crm_materials = _crm_conn.execute("SELECT * FROM crm_materials ORDER BY crm_name").fetchall()
    now = datetime.now().isoformat()
    _crm_conn.execute("DELETE FROM guest_tokens WHERE expires_at < ?", (now,))
    _crm_conn.commit()
    guest_tokens = _crm_conn.execute("SELECT * FROM guest_tokens ORDER BY created_at DESC").fetchall()
    _crm_conn.close()
    _envconn = get_db()
    env_rooms = _envconn.execute(
        "SELECT * FROM env_rooms WHERE is_active=1 ORDER BY sort_order, name").fetchall()
    _envconn.close()
    return render_template('admin/settings.html', s=s, qc_settings=qc_settings_list, lang=lang,
                           crm_materials=crm_materials, today=date.today().isoformat(),
                           guest_tokens=guest_tokens, now=now,
                           env_rooms=env_rooms, env_lim=env_limits())

@app.route('/lab-settings/crm', methods=['POST'])
@admin_required
def lab_settings_crm():
    action = request.form.get('action')
    conn = get_db()
    try:
        if action == 'add':
            name = request.form.get('crm_name', '').strip()
            if not name:
                flash('CRM нэрийг оруулна уу', 'error')
                return redirect(url_for('lab_settings') + '?tab=crm')
            conn.execute("""INSERT INTO crm_materials (crm_name, aad_cert, aad_unc, vad_cert, vad_unc, sulfur_cert, sulfur_unc, cal_cert, cal_unc, g_cert, g_unc, notes, standard, manufacture_date, expiry_date, open_date)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                name,
                request.form.get('aad_cert') or None,
                request.form.get('aad_unc') or None,
                request.form.get('vad_cert') or None,
                request.form.get('vad_unc') or None,
                request.form.get('sulfur_cert') or None,
                request.form.get('sulfur_unc') or None,
                request.form.get('cal_cert') or None,
                request.form.get('cal_unc') or None,
                request.form.get('g_cert') or None,
                request.form.get('g_unc') or None,
                request.form.get('notes', '').strip() or None,
                request.form.get('standard', '').strip() or None,
                request.form.get('manufacture_date') or None,
                request.form.get('expiry_date') or None,
                request.form.get('open_date') or None,
            ))
            conn.commit()
            flash(f'CRM материал нэмэгдлээ: {name}', 'success')
        elif action == 'delete':
            mid = request.form.get('id')
            conn.execute("DELETE FROM crm_materials WHERE id=?", (mid,))
            conn.commit()
            flash('CRM материал устгагдлаа', 'success')
        elif action == 'deactivate':
            mid = request.form.get('id')
            conn.execute("UPDATE crm_materials SET is_active=0 WHERE id=?", (mid,))
            conn.commit()
            flash('CRM материал дуусгагдлаа', 'success')
    finally:
        conn.close()
    return redirect(url_for('lab_settings') + '?tab=crm')

@app.route('/lab-settings/crm/<int:mid>/edit', methods=['GET','POST'])
@admin_required
def crm_edit(mid):
    conn = get_db()
    mat = conn.execute("SELECT * FROM crm_materials WHERE id=?", (mid,)).fetchone()
    if not mat:
        conn.close()
        flash('CRM материал олдсонгүй', 'error')
        return redirect(url_for('lab_settings') + '?tab=crm')
    if request.method == 'POST':
        conn.execute("""UPDATE crm_materials SET
            crm_name=?, aad_cert=?, aad_unc=?, vad_cert=?, vad_unc=?,
            sulfur_cert=?, sulfur_unc=?, cal_cert=?, cal_unc=?,
            g_cert=?, g_unc=?, notes=?, standard=?,
            manufacture_date=?, expiry_date=?, open_date=?
            WHERE id=?""", (
            request.form.get('crm_name','').strip(),
            request.form.get('aad_cert') or None, request.form.get('aad_unc') or None,
            request.form.get('vad_cert') or None, request.form.get('vad_unc') or None,
            request.form.get('sulfur_cert') or None, request.form.get('sulfur_unc') or None,
            request.form.get('cal_cert') or None, request.form.get('cal_unc') or None,
            request.form.get('g_cert') or None, request.form.get('g_unc') or None,
            request.form.get('notes','').strip() or None,
            request.form.get('standard','').strip() or None,
            request.form.get('manufacture_date') or None,
            request.form.get('expiry_date') or None,
            request.form.get('open_date') or None,
            mid
        ))
        conn.commit(); conn.close()
        flash('CRM материал шинэчлэгдлээ', 'success')
        return redirect(url_for('lab_settings') + '?tab=crm')
    conn.close()
    return render_template('admin/crm_edit.html', mat=mat, today=date.today().isoformat())

# ── НӨӨЦЛӨЛТ ────────────────────────────────────────────
# WAL горимд shutil.copy2 нь АЮУЛТАЙ: хуулах агшинд WAL дотор байгаа
# гүйлгээ хуулбарт ороогүй үлдэж, эвдэрсэн сан үүсэж болно. SQLite-ийн
# online backup API нь ажиллаж байгаа сан дээр аюулгүй ажиллана.
INSTANCE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance')
BACKUP_KEEP_DAYS = 14


def make_backup():
    """Аюулгүй хуулбар үүсгээд замыг нь буцаана."""
    import sqlite3 as _sq
    src_path = os.path.join(INSTANCE_DIR, 'lab.db')
    if not os.path.exists(src_path):
        return None
    os.makedirs(INSTANCE_DIR, exist_ok=True)
    dest = os.path.join(INSTANCE_DIR,
                        f"lab_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
    src = _sq.connect(src_path, timeout=30)
    try:
        dst = _sq.connect(dest)
        try:
            src.backup(dst)
        finally:
            dst.close()
    finally:
        src.close()
    return dest


def prune_backups(keep_days=BACKUP_KEEP_DAYS):
    """Хуучирсан хуулбарыг цэвэрлэнэ — диск дүүрэхээс сэргийлнэ."""
    import glob as _glob
    cutoff, removed = _time.time() - keep_days * 86400, 0
    for f in _glob.glob(os.path.join(INSTANCE_DIR, 'lab_backup_*.db')):
        try:
            if os.path.getmtime(f) < cutoff:
                os.remove(f)
                removed += 1
        except OSError:
            pass
    return removed


def _backup_daily_loop():
    """Өдөрт нэг удаа автомат нөөцлөлт.

    Урьд нь хуулбарыг ЗӨВХӨН процесс эхлэхэд авдаг байсан тул сервер
    сар турш restart хийлгүй ажиллавал ганцхан хуулбар үүсдэг байв.
    """
    import glob as _glob
    while True:
        try:
            today = datetime.now().strftime('%Y%m%d')
            if not _glob.glob(os.path.join(INSTANCE_DIR, f'lab_backup_{today}_*.db')):
                path = make_backup()
                if path:
                    app.logger.info('Автомат нөөцлөлт: %s', os.path.basename(path))
                    prune_backups()
        except Exception:
            app.logger.exception('Автомат нөөцлөлт амжилтгүй боллоо')
        _time.sleep(3600)          # цаг тутам шалгана


@app.route('/backup')
@admin_required
def backup_db():
    """Одоогийн байдлаар аюулгүй хуулбар үүсгэж татуулна."""
    path = make_backup()
    if not path:
        flash('Мэдээллийн сан олдсонгүй', 'error')
        return redirect(url_for('lab_settings'))
    return send_file(path, as_attachment=True, download_name=os.path.basename(path))

# ── ШАЛГАЛТЫН ЗАГВАР (check templates) ──────────────────
@app.route('/check-templates')
@login_required
def check_templates_list():
    conn = get_db()
    rows = conn.execute("SELECT * FROM check_templates ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/check-templates/add', methods=['POST'])
@senior_required
def check_templates_add():
    conn = get_db()
    data = request.get_json()
    conn.execute("""INSERT INTO check_templates
        (name,param1,standard1,tolerance1,param2,standard2,tolerance2,
         param3,standard3,tolerance3,param4,standard4,tolerance4,created_by)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
        data.get('name'), data.get('param1'), data.get('standard1'), data.get('tolerance1'),
        data.get('param2'), data.get('standard2'), data.get('tolerance2'),
        data.get('param3'), data.get('standard3'), data.get('tolerance3'),
        data.get('param4'), data.get('standard4'), data.get('tolerance4'),
        session.get('user_id')
    ))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/check-templates/<int:tid>/delete', methods=['POST'])
@senior_required
def check_templates_delete(tid):
    conn = get_db()
    conn.execute("DELETE FROM check_templates WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/backup/list')
@admin_required
def backup_list():
    import glob, os
    inst = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance')
    files = []
    for f in sorted(glob.glob(os.path.join(inst, 'lab_backup_*.db')), reverse=True):
        stat = os.stat(f)
        files.append({
            'name': os.path.basename(f),
            'size': stat.st_size,
            'mtime': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M')
        })
    return jsonify(files)

@app.route('/backup/download/<filename>')
@admin_required
def backup_download(filename):
    import re
    if not re.match(r'^lab_backup_[\d_]+\.db$', filename):
        return 'Invalid', 400
    inst = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance')
    path = os.path.join(inst, filename)
    if not os.path.exists(path):
        return 'Not found', 404
    return send_file(path, as_attachment=True, download_name=filename)

@app.route('/backup/create', methods=['POST'])
@admin_required
def backup_create():
    path = make_backup()
    if not path:
        return jsonify({'ok': False, 'error': 'Мэдээллийн сан олдсонгүй'}), 404
    prune_backups()
    return jsonify({'ok': True, 'name': os.path.basename(path)})

# ── ОРЧНЫ ХЯНАЛТ (чийг / дулаан) ────────────────────────
ENV_SLOTS = [('start', 'Шинжилгээ эхлэхэд'), ('end', 'Шинжилгээ төгсгөлд')]
ENV_LIMIT_DEFAULTS = {'env_temp_min': 15.0, 'env_temp_max': 30.0,
                      'env_hum_min': 20.0,  'env_hum_max': 80.0}


def env_limits():
    """Бүх өрөөнд нэг ижил зөвшөөрөгдөх хязгаар (тохиргооны файлд)"""
    s = get_settings()
    out = {}
    for k, dflt in ENV_LIMIT_DEFAULTS.items():
        try:
            out[k] = float(s.get(k, dflt))
        except (TypeError, ValueError):
            out[k] = dflt
    return out


def env_check(temp, hum, lim):
    """Хязгаараас гарсан үзүүлэлтийн жагсаалт"""
    bad = []
    if temp is not None and not (lim['env_temp_min'] <= temp <= lim['env_temp_max']):
        bad.append('температур')
    if hum is not None and not (lim['env_hum_min'] <= hum <= lim['env_hum_max']):
        bad.append('чийг')
    return bad


@app.route('/env')
@lab_required
def env_page():
    """Өдрийн орчны бүртгэл — өрөө тус бүрт эхлэл ба төгсгөл"""
    lang = session.get('lang', 'mn')
    day = request.args.get('date') or date.today().isoformat()
    conn = get_db()
    rooms = conn.execute(
        "SELECT * FROM env_rooms WHERE is_active=1 ORDER BY sort_order, name").fetchall()
    rows = conn.execute("""
        SELECT r.*, u.name as by_name FROM env_readings r
        LEFT JOIN users u ON u.id=r.recorded_by
        WHERE r.reading_date=?
    """, (day,)).fetchall()
    conn.close()
    readings = {(r['room_id'], r['slot']): dict(r) for r in rows}
    lim = env_limits()
    for key, r in readings.items():
        r['bad'] = env_check(r['temperature'], r['humidity'], lim)
    return render_template('device/env.html', rooms=rooms, readings=readings,
                           day=day, slots=ENV_SLOTS, lim=lim, lang=lang,
                           today=date.today().isoformat())


@app.route('/env/save', methods=['POST'])
@lab_required
def env_save():
    day = request.form.get('date') or date.today().isoformat()
    room_id = request.form.get('room_id', type=int)
    slot = request.form.get('slot')
    if not room_id or slot not in ('start', 'end'):
        flash('Буруу хүсэлт.', 'error')
        return redirect(url_for('env_page', date=day))

    def num(k):
        v = (request.form.get(k) or '').strip()
        try:
            return float(v) if v != '' else None
        except ValueError:
            return None
    temp, hum = num('temperature'), num('humidity')
    notes = (request.form.get('notes') or '').strip()
    lim = env_limits()
    bad = env_check(temp, hum, lim)
    if bad and not notes:
        flash(f"{', '.join(bad).capitalize()} хязгаараас гарсан — тайлбар бичнэ үү.", 'error')
        return redirect(url_for('env_page', date=day))

    conn = get_db()
    if temp is None and hum is None:
        conn.execute("DELETE FROM env_readings WHERE room_id=? AND reading_date=? AND slot=?",
                     (room_id, day, slot))
        msg = 'Бүртгэл хоослов.'
    else:
        conn.execute("""
            INSERT INTO env_readings(room_id, reading_date, slot, temperature, humidity,
                                     recorded_by, recorded_at, notes)
            VALUES(?,?,?,?,?,?,?,?)
            ON CONFLICT(room_id, reading_date, slot) DO UPDATE SET
                temperature=excluded.temperature, humidity=excluded.humidity,
                recorded_by=excluded.recorded_by, recorded_at=excluded.recorded_at,
                notes=excluded.notes
        """, (room_id, day, slot, temp, hum, session.get('user_id'),
              datetime.now().strftime('%H:%M'), notes))
        msg = ('Хадгаллаа. ⚠️ ' + ', '.join(bad) + ' хязгаараас гарсан.') if bad else 'Хадгаллаа.'
    conn.commit()
    conn.close()
    flash(msg, 'error' if bad else 'success')
    return redirect(url_for('env_page', date=day))


# Хугацааны нэгтгэл: SQLite-ийн илэрхийлэл + харагдах нэр
ENV_PERIODS = {
    'week':     ("strftime('%Y-W%W', reading_date)",  'Долоо хоног'),
    'month':    ("strftime('%Y-%m', reading_date)",   'Сар'),
    'halfyear': ("strftime('%Y', reading_date) || '-H' || "
                 "(CASE WHEN CAST(strftime('%m', reading_date) AS INTEGER)<=6 "
                 "THEN 1 ELSE 2 END)",                'Хагас жил'),
    'year':     ("strftime('%Y', reading_date)",      'Жил'),
}


def env_stats_data(period='month'):
    """Орчны хяналтын нэгтгэл — долоо хоног / сар / хагас жил / жил.

    Тайлангийн хуудас болон бусад хуудсууд ижил функц ашиглана.
    """
    if period not in ENV_PERIODS:
        period = 'month'
    expr = ENV_PERIODS[period][0]
    lim = env_limits()
    conn = get_db()
    rooms = conn.execute(
        "SELECT * FROM env_rooms WHERE is_active=1 ORDER BY sort_order, name").fetchall()
    # Өрөө × хугацаа тус бүрийн дундаж, хамгийн бага/их, хэтэрсэн тоо
    agg = conn.execute(f"""
        SELECT {expr} AS period, room_id,
               COUNT(*) AS n,
               ROUND(AVG(temperature),1) AS t_avg,
               MIN(temperature) AS t_min, MAX(temperature) AS t_max,
               ROUND(AVG(humidity),1)    AS h_avg,
               MIN(humidity) AS h_min, MAX(humidity) AS h_max,
               SUM(CASE WHEN temperature IS NOT NULL
                         AND (temperature < ? OR temperature > ?) THEN 1 ELSE 0 END) AS t_bad,
               SUM(CASE WHEN humidity IS NOT NULL
                         AND (humidity < ? OR humidity > ?) THEN 1 ELSE 0 END) AS h_bad
        FROM env_readings
        WHERE temperature IS NOT NULL OR humidity IS NOT NULL
        GROUP BY period, room_id
        ORDER BY period DESC
    """, (lim['env_temp_min'], lim['env_temp_max'],
          lim['env_hum_min'], lim['env_hum_max'])).fetchall()
    # Хугацаа тус бүрийн бүх өрөөний нийлбэр дүн
    totals = conn.execute(f"""
        SELECT {expr} AS period, COUNT(*) AS n,
               ROUND(AVG(temperature),1) AS t_avg, ROUND(AVG(humidity),1) AS h_avg,
               SUM(CASE WHEN (temperature IS NOT NULL AND (temperature < ? OR temperature > ?))
                          OR (humidity IS NOT NULL AND (humidity < ? OR humidity > ?))
                        THEN 1 ELSE 0 END) AS bad
        FROM env_readings
        WHERE temperature IS NOT NULL OR humidity IS NOT NULL
        GROUP BY period ORDER BY period DESC LIMIT 24
    """, (lim['env_temp_min'], lim['env_temp_max'],
          lim['env_hum_min'], lim['env_hum_max'])).fetchall()
    # Хамгийн сүүлийн бүртгэлүүд (жинхэнэ бүртгэлийн хуудас)
    recent = conn.execute("""
        SELECT r.*, u.name AS by_name, m.name AS room_name
        FROM env_readings r
        LEFT JOIN users u ON u.id=r.recorded_by
        LEFT JOIN env_rooms m ON m.id=r.room_id
        ORDER BY r.reading_date DESC, m.sort_order, r.slot LIMIT 120
    """).fetchall()
    conn.close()

    by_period = {}
    for a in agg:
        by_period.setdefault(a['period'], {})[a['room_id']] = dict(a)
    recent_list = []
    for r in recent:
        d = dict(r)
        d['bad'] = env_check(d['temperature'], d['humidity'], lim)
        recent_list.append(d)
    return dict(env_rooms=rooms, env_by_period=by_period, env_totals=totals,
                env_periods=[t['period'] for t in totals], env_recent=recent_list,
                env_period=period, env_period_name=ENV_PERIODS[period][1],
                env_period_opts=ENV_PERIODS, env_slots=dict(ENV_SLOTS), env_lim=lim)


@app.route('/env/export')
@lab_required
def env_export():
    """Сарын орчны бүртгэлийг Excel болгон татах"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    month = request.args.get('month') or date.today().strftime('%Y-%m')
    conn = get_db()
    rooms = conn.execute(
        "SELECT * FROM env_rooms WHERE is_active=1 ORDER BY sort_order, name").fetchall()
    rows = conn.execute("""
        SELECT r.*, u.name as by_name FROM env_readings r
        LEFT JOIN users u ON u.id=r.recorded_by
        WHERE substr(r.reading_date,1,7)=? ORDER BY r.reading_date
    """, (month,)).fetchall()
    conn.close()
    lim = env_limits()
    data = {(r['reading_date'], r['room_id'], r['slot']): r for r in rows}
    days = sorted({r['reading_date'] for r in rows})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Орчны хяналт'
    ws['A1'] = f'ОРЧНЫ ХЯНАЛТЫН БҮРТГЭЛ — {month}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = (f"Зөвшөөрөгдөх хязгаар: температур {lim['env_temp_min']}–{lim['env_temp_max']}°C, "
                f"чийг {lim['env_hum_min']}–{lim['env_hum_max']}%")
    ws['A2'].font = Font(size=10, italic=True)

    hdr = Font(bold=True, size=10)
    ws.cell(4, 1, 'Огноо').font = hdr
    col = 2
    for room in rooms:
        for _, slot_name in ENV_SLOTS:
            c = ws.cell(4, col, f"{room['name']}\n{slot_name}")
            c.font = hdr
            c.alignment = Alignment(wrap_text=True, horizontal='center')
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
            col += 1
    ws.column_dimensions['A'].width = 12
    warn = PatternFill('solid', fgColor='FFF2CC')

    r_i = 5
    for day in days:
        ws.cell(r_i, 1, day)
        col = 2
        for room in rooms:
            for slot, _ in ENV_SLOTS:
                rec = data.get((day, room['id'], slot))
                if rec:
                    t = rec['temperature']
                    h = rec['humidity']
                    txt = f"{t if t is not None else '—'}°C / {h if h is not None else '—'}%"
                    if rec['by_name']:
                        txt += f"\n{rec['by_name']}"
                    cell = ws.cell(r_i, col, txt)
                    cell.alignment = Alignment(wrap_text=True, horizontal='center')
                    if env_check(t, h, lim):
                        cell.fill = warn
                col += 1
        r_i += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True,
                     download_name=f'orchin_{month}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/lab-settings/env-rooms', methods=['POST'])
@admin_required
def env_rooms_save():
    """Өрөө нэмэх / устгах + бүх өрөөнд хамаарах хязгаар"""
    action = request.form.get('action')
    conn = get_db()
    if action == 'add':
        nm = (request.form.get('name') or '').strip()
        if nm:
            n = conn.execute("SELECT COALESCE(MAX(sort_order),0)+1 FROM env_rooms").fetchone()[0]
            conn.execute("INSERT INTO env_rooms(name, sort_order) VALUES(?,?)", (nm, n))
            flash(f'"{nm}" өрөө нэмэгдлээ.', 'success')
    elif action == 'remove':
        rid = request.form.get('room_id', type=int)
        # Бүртгэл байгаа өрөөг устгахгүй — түүх хадгална, зөвхөн нуана
        conn.execute("UPDATE env_rooms SET is_active=0 WHERE id=?", (rid,))
        flash('Өрөө хаагдлаа (бүртгэлийн түүх хадгалагдана).', 'success')
    elif action == 'limits':
        s = get_settings()
        for k in ENV_LIMIT_DEFAULTS:
            v = (request.form.get(k) or '').strip()
            if v != '':
                try:
                    s[k] = float(v)
                except ValueError:
                    pass
        save_settings(s)
        flash('Орчны хязгаар хадгалагдлаа.', 'success')
    conn.commit()
    conn.close()
    return redirect(url_for('lab_settings'))


@app.route('/lab-settings/usage-clear', methods=['POST'])
@admin_required
def usage_clear():
    """Тоног төхөөрөмж ашигласан цагийн бүх бүртгэлийг цэвэрлэнэ.

    Буцаах боломжгүй тул устгахын өмнө автоматаар нөөцөлнө. Ашиглагдаж
    байгаа (дуусаагүй) бүртгэл ч цэвэрлэгдэх тул түгжигдсэн тоног
    төхөөрөмж чөлөөлөгдөнө.
    """
    try:
        _p = make_backup()
        bk_name = os.path.basename(_p) if _p else None
    except Exception:
        app.logger.exception('Устгахын өмнөх нөөцлөлт амжилтгүй')
        bk_name = None

    conn = get_db()
    counts = {}
    for tbl in ('usage_logs', 'device_usage_log'):
        try:
            counts[tbl] = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
            conn.execute(f"DELETE FROM {tbl}")
        except Exception:
            counts[tbl] = 0
    conn.commit()
    conn.close()

    msg = (f"Ашиглалтын цаг цэвэрлэгдлээ — {counts.get('usage_logs',0)} ашиглалтын "
           f"бүртгэл, {counts.get('device_usage_log',0)} шинжилгээний ашиглалт устгав.")
    if bk_name:
        msg += f' Нөөц хуулбар: {bk_name}'
    flash(msg, 'success')
    return redirect(url_for('lab_settings'))


@app.route('/backup/delete/<filename>', methods=['POST'])
@admin_required
def backup_delete(filename):
    import re
    if not re.match(r'^lab_backup_[\d_]+\.db$', filename):
        return jsonify({'ok': False}), 400
    inst = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance')
    path = os.path.join(inst, filename)
    if os.path.exists(path):
        os.remove(path)
    return jsonify({'ok': True})


def _startup():
    init_db()
    from models import init_analysis_db
    init_analysis_db()
    ensure_tables()
    _check_weak_admin()
    # Өдөр тутмын автомат нөөцлөлт — процессын хажуугаар байнга ажиллана
    import threading
    threading.Thread(target=_backup_daily_loop, daemon=True,
                     name='backup-daily').start()

_startup()

if __name__ == '__main__':
    print('Систем эхэллээ!')
    print('Браузерт нэвтрэх: http://localhost:5000')
    try:
        # Production сервер: олон хэрэглэгчийн ачааллыг Flask-ийн dev серверээс
        # хамаагүй сайн даана (хүсэлтийн дараалал, тогтмол thread pool)
        from waitress import serve
        serve(app, host='0.0.0.0', port=5000, threads=8)
    except ImportError:
        app.run(debug=False, host='0.0.0.0', port=5000)
